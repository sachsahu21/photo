"""
Image Scanner - Scans folders for images, extracts metadata,
detects duplicates, and saves results to Excel.

Usage:
    python image_scanner.py                        # Scans home directory
    python image_scanner.py /path/to/scan          # Scans specific folder
    python image_scanner.py /path1 /path2 --out results.xlsx

Requirements:
    pip install pillow openpyxl tqdm
    pip install pillow[exif]   # For full EXIF support
"""

import os
import sys
import hashlib
import argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict

try:
    from PIL import Image
    from PIL.ExifTags import TAGS, GPSTAGS
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    print("⚠️  Pillow not installed. Install with: pip install pillow\n   Metadata extraction will be limited.")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Install with: pip install openpyxl\n   Will fall back to CSV output.")

try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False


# ── Supported image extensions ─────────────────────────────────────────────────
IMAGE_EXTENSIONS = {
    '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif',
    '.webp', '.heic', '.heif', '.raw', '.cr2', '.nef', '.arw',
    '.dng', '.orf', '.rw2', '.pef', '.svg', '.ico', '.psd',
    '.avif', '.jfif', '.exr', '.hdr'
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def file_hash(filepath, chunk=65536):
    """MD5 hash of file contents for duplicate detection."""
    h = hashlib.md5()
    try:
        with open(filepath, 'rb') as f:
            while True:
                data = f.read(chunk)
                if not data:
                    break
                h.update(data)
        return h.hexdigest()
    except (PermissionError, OSError):
        return None


def get_gps_coords(exif_data):
    """Extract GPS lat/lon from EXIF as a readable string."""
    try:
        gps_info = exif_data.get('GPSInfo', {})
        if not gps_info:
            return None, None

        def to_decimal(vals):
            d, m, s = vals
            return float(d) + float(m) / 60 + float(s) / 3600

        lat = to_decimal(gps_info.get(2, (0, 0, 0)))
        lon = to_decimal(gps_info.get(4, (0, 0, 0)))
        lat_ref = gps_info.get(1, 'N')
        lon_ref = gps_info.get(3, 'E')

        if lat_ref == 'S':
            lat = -lat
        if lon_ref == 'W':
            lon = -lon

        return round(lat, 6), round(lon, 6)
    except Exception:
        return None, None


def extract_metadata(filepath):
    """Return a dict of image metadata."""
    meta = {
        'width': None, 'height': None, 'mode': None,
        'date_taken': None, 'camera_make': None, 'camera_model': None,
        'focal_length': None, 'aperture': None, 'iso': None,
        'exposure_time': None, 'gps_lat': None, 'gps_lon': None,
        'color_space': None, 'dpi': None, 'has_exif': False,
        'error': None
    }

    if not PIL_AVAILABLE:
        meta['error'] = 'Pillow not installed'
        return meta

    try:
        with Image.open(filepath) as img:
            meta['width'] = img.width
            meta['height'] = img.height
            meta['mode'] = img.mode

            # DPI
            dpi = img.info.get('dpi')
            if dpi:
                meta['dpi'] = f"{int(dpi[0])}x{int(dpi[1])}"

            # EXIF
            exif_raw = None
            if hasattr(img, '_getexif') and img._getexif():
                exif_raw = img._getexif()
            elif hasattr(img, 'getexif'):
                exif_raw = img.getexif()

            if exif_raw:
                meta['has_exif'] = True
                exif = {}
                gps_data = {}

                for tag_id, value in exif_raw.items():
                    tag = TAGS.get(tag_id, tag_id)
                    if tag == 'GPSInfo' and isinstance(value, dict):
                        for gps_tag_id, gps_val in value.items():
                            gps_tag = GPSTAGS.get(gps_tag_id, gps_tag_id)
                            gps_data[gps_tag_id] = gps_val
                        exif['GPSInfo'] = gps_data
                    else:
                        exif[tag] = value

                meta['camera_make'] = str(exif.get('Make', '') or '').strip() or None
                meta['camera_model'] = str(exif.get('Model', '') or '').strip() or None
                meta['color_space'] = str(exif.get('ColorSpace', '') or '').strip() or None

                # Date taken
                for dt_tag in ('DateTimeOriginal', 'DateTime', 'DateTimeDigitized'):
                    dt_val = exif.get(dt_tag)
                    if dt_val:
                        try:
                            meta['date_taken'] = datetime.strptime(str(dt_val), '%Y:%m:%d %H:%M:%S')
                        except ValueError:
                            meta['date_taken'] = str(dt_val)
                        break

                # Focal length
                fl = exif.get('FocalLength')
                if fl:
                    try:
                        meta['focal_length'] = f"{float(fl):.1f}mm"
                    except Exception:
                        meta['focal_length'] = str(fl)

                # Aperture (FNumber)
                fn = exif.get('FNumber')
                if fn:
                    try:
                        meta['aperture'] = f"f/{float(fn):.1f}"
                    except Exception:
                        meta['aperture'] = str(fn)

                # ISO
                iso = exif.get('ISOSpeedRatings') or exif.get('PhotographicSensitivity')
                if iso:
                    meta['iso'] = str(iso)

                # Exposure time
                exp = exif.get('ExposureTime')
                if exp:
                    try:
                        fv = float(exp)
                        if fv < 1:
                            meta['exposure_time'] = f"1/{round(1/fv)}s"
                        else:
                            meta['exposure_time'] = f"{fv}s"
                    except Exception:
                        meta['exposure_time'] = str(exp)

                # GPS
                lat, lon = get_gps_coords(exif)
                meta['gps_lat'] = lat
                meta['gps_lon'] = lon

    except Exception as e:
        meta['error'] = str(e)[:120]

    return meta


def scan_images(scan_roots, progress=True):
    """Walk directories and collect image file records."""
    records = []
    roots = [Path(r).expanduser().resolve() for r in scan_roots]

    # Count files first for progress bar
    print(f"\n🔍 Scanning: {', '.join(str(r) for r in roots)}")
    all_files = []
    for root in roots:
        for dirpath, _, filenames in os.walk(root, followlinks=False):
            for fn in filenames:
                fp = Path(dirpath) / fn
                if fp.suffix.lower() in IMAGE_EXTENSIONS:
                    all_files.append(fp)

    print(f"   Found {len(all_files)} image files. Extracting metadata...\n")

    iterator = tqdm(all_files, unit='img') if (TQDM_AVAILABLE and progress) else all_files

    hash_map = defaultdict(list)  # md5 → [filepath, ...]

    for fp in iterator:
        stat = fp.stat()
        size_bytes = stat.st_size
        file_md5 = file_hash(fp)
        meta = extract_metadata(fp)

        record = {
            'filename': fp.name,
            'folder': str(fp.parent),
            'full_path': str(fp),
            'extension': fp.suffix.lower().lstrip('.').upper(),
            'size_bytes': size_bytes,
            'size_kb': round(size_bytes / 1024, 1),
            'file_modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
            'md5_hash': file_md5,
            **meta
        }
        records.append(record)

        if file_md5:
            hash_map[file_md5].append(str(fp))

    # Mark duplicates
    duplicate_groups = {}
    group_id = 1
    for md5, paths in hash_map.items():
        if len(paths) > 1:
            for p in paths:
                duplicate_groups[p] = group_id
            group_id += 1

    for r in records:
        r['is_duplicate'] = 'YES' if r['full_path'] in duplicate_groups else 'No'
        r['duplicate_group'] = duplicate_groups.get(r['full_path'], '')

    dup_count = sum(1 for r in records if r['is_duplicate'] == 'YES')
    print(f"\n✅ Scanned {len(records)} images. Duplicates found: {dup_count}")
    return records, hash_map


# ── Excel Output ───────────────────────────────────────────────────────────────

HEADER_STYLE = {'font': Font(bold=True, color='FFFFFF', size=11),
                'fill': PatternFill('solid', start_color='2E4057'),
                'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)}

DUP_FILL = PatternFill('solid', start_color='FFD6D6')
ALT_FILL = PatternFill('solid', start_color='F5F7FA')

COLUMNS = [
    ('filename',        'Filename',         28),
    ('folder',          'Folder',           45),
    ('extension',       'Format',           10),
    ('size_kb',         'Size (KB)',         11),
    ('width',           'Width (px)',        11),
    ('height',          'Height (px)',       11),
    ('mode',            'Color Mode',       12),
    ('dpi',             'DPI',              10),
    ('date_taken',      'Date Taken',       20),
    ('camera_make',     'Camera Make',      16),
    ('camera_model',    'Camera Model',     18),
    ('focal_length',    'Focal Length',     13),
    ('aperture',        'Aperture',         11),
    ('iso',             'ISO',               8),
    ('exposure_time',   'Exposure',         11),
    ('gps_lat',         'GPS Lat',          11),
    ('gps_lon',         'GPS Lon',          11),
    ('has_exif',        'Has EXIF',         10),
    ('is_duplicate',    'Duplicate?',       12),
    ('duplicate_group', 'Dup. Group',       10),
    ('md5_hash',        'MD5 Hash',         34),
    ('file_modified',   'File Modified',    20),
    ('full_path',       'Full Path',        55),
    ('error',           'Read Error',       30),
]


def thin_border():
    side = Side(style='thin', color='CCCCCC')
    return Border(left=side, right=side, top=side, bottom=side)


def write_excel(records, out_path):
    wb = openpyxl.Workbook()

    # ── Sheet 1: All Images ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'All Images'
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 32

    for col_idx, (key, label, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_STYLE['font']
        cell.fill = HEADER_STYLE['fill']
        cell.alignment = HEADER_STYLE['alignment']
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, rec in enumerate(records, 2):
        is_dup = rec.get('is_duplicate') == 'YES'
        fill = DUP_FILL if is_dup else (ALT_FILL if row_idx % 2 == 0 else None)
        for col_idx, (key, _, _) in enumerate(COLUMNS, 1):
            val = rec.get(key, '')
            if isinstance(val, bool):
                val = 'Yes' if val else 'No'
            if isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d %H:%M:%S')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical='center')
            if fill:
                cell.fill = fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Sheet 2: Duplicates Only ───────────────────────────────────────────────
    ws2 = wb.create_sheet('Duplicates')
    ws2.freeze_panes = 'A2'
    ws2.row_dimensions[1].height = 32

    dup_cols = [('duplicate_group','Group',8), ('filename','Filename',28),
                ('folder','Folder',45), ('extension','Format',10),
                ('size_kb','Size (KB)',11), ('width','Width (px)',11),
                ('height','Height (px)',11), ('date_taken','Date Taken',20),
                ('md5_hash','MD5 Hash',34), ('full_path','Full Path',55)]

    for col_idx, (_, label, width) in enumerate(dup_cols, 1):
        cell = ws2.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_STYLE['font']
        cell.fill = PatternFill('solid', start_color='8B0000')
        cell.alignment = HEADER_STYLE['alignment']
        cell.border = thin_border()
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    dups = sorted([r for r in records if r.get('is_duplicate') == 'YES'],
                  key=lambda x: (x.get('duplicate_group', 0), x['full_path']))

    prev_group = None
    alt = False
    for row_idx, rec in enumerate(dups, 2):
        grp = rec.get('duplicate_group')
        if grp != prev_group:
            alt = not alt
            prev_group = grp
        fill = PatternFill('solid', start_color='FFE8E8') if alt else PatternFill('solid', start_color='FFF5F5')
        for col_idx, (key, _, _) in enumerate(dup_cols, 1):
            val = rec.get(key, '')
            if isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d %H:%M:%S')
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical='center')
            cell.fill = fill
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(dup_cols))}1"

    # ── Sheet 3: Summary ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Summary')
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 20

    summary_title_font = Font(bold=True, size=14, color='2E4057')
    label_font = Font(bold=True, size=11)
    accent_fill = PatternFill('solid', start_color='EBF2FF')

    ws3['A1'] = '📊 Image Scan Summary'
    ws3['A1'].font = summary_title_font
    ws3['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws3['A2'].font = Font(italic=True, color='888888')

    from collections import Counter
    ext_counts = Counter(r['extension'] for r in records)
    total = len(records)
    dup_total = sum(1 for r in records if r['is_duplicate'] == 'YES')
    dup_groups = len(set(r['duplicate_group'] for r in records if r['duplicate_group'] != ''))
    with_exif = sum(1 for r in records if r.get('has_exif'))
    with_gps = sum(1 for r in records if r.get('gps_lat'))

    rows = [
        ('', ''),
        ('GENERAL', ''),
        ('Total Images Found', total),
        ('Total Folders Scanned', len(set(r['folder'] for r in records))),
        ('Total Size (MB)', round(sum(r['size_bytes'] for r in records) / 1024 / 1024, 1)),
        ('', ''),
        ('DUPLICATES', ''),
        ('Duplicate Files', dup_total),
        ('Duplicate Groups', dup_groups),
        ('', ''),
        ('METADATA', ''),
        ('Files with EXIF', with_exif),
        ('Files with GPS', with_gps),
        ('', ''),
        ('BY FORMAT', ''),
    ]

    for ext, cnt in sorted(ext_counts.items(), key=lambda x: -x[1]):
        rows.append((f'  .{ext.lower()}', cnt))

    for row_idx, (label, value) in enumerate(rows, 4):
        a = ws3.cell(row=row_idx, column=1, value=label)
        b = ws3.cell(row=row_idx, column=2, value=value)
        if label in ('GENERAL', 'DUPLICATES', 'METADATA', 'BY FORMAT'):
            a.font = Font(bold=True, color='FFFFFF')
            a.fill = PatternFill('solid', start_color='2E4057')
            b.fill = PatternFill('solid', start_color='2E4057')
        elif label:
            a.font = label_font
            if row_idx % 2 == 0:
                a.fill = accent_fill
                b.fill = accent_fill

    wb.save(out_path)
    print(f"\n💾 Excel saved → {out_path}")


def write_csv(records, out_path):
    import csv
    keys = [c[0] for c in COLUMNS]
    headers = [c[1] for c in COLUMNS]
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=keys, extrasaction='ignore')
        f.write(','.join(headers) + '\n')
        for rec in records:
            row = {}
            for k in keys:
                v = rec.get(k, '')
                if isinstance(v, datetime):
                    v = v.strftime('%Y-%m-%d %H:%M:%S')
                if isinstance(v, bool):
                    v = 'Yes' if v else 'No'
                row[k] = v
            w.writerow(row)
    print(f"\n💾 CSV saved → {out_path}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Scan folders for images, extract metadata, detect duplicates.')
    parser.add_argument('paths', nargs='*', default=[str(Path.home())],
                        help='Folders to scan (default: home directory)')
    parser.add_argument('--out', default='image_scan_results.xlsx',
                        help='Output file (.xlsx or .csv)')
    parser.add_argument('--no-progress', action='store_true',
                        help='Disable progress bar')
    args = parser.parse_args()

    # Validate paths
    for p in args.paths:
        if not Path(p).exists():
            print(f"❌ Path not found: {p}")
            sys.exit(1)

    records, hash_map = scan_images(args.paths, progress=not args.no_progress)

    if not records:
        print("No images found.")
        return

    out = args.out
    if out.endswith('.csv') or not OPENPYXL_AVAILABLE:
        if not out.endswith('.csv'):
            out = out.replace('.xlsx', '.csv') if '.xlsx' in out else out + '.csv'
        write_csv(records, out)
    else:
        write_excel(records, out)

    # Print quick stats
    dup_count = sum(1 for r in records if r['is_duplicate'] == 'YES')
    dup_groups = len(set(r['duplicate_group'] for r in records if r['duplicate_group']))
    print(f"""
┌─────────────────────────────────────┐
│          SCAN COMPLETE              │
├─────────────────────────────────────┤
│  Total images    : {len(records):<18}│
│  Duplicate files : {dup_count:<18}│
│  Duplicate groups: {dup_groups:<18}│
│  Output          : {out[:18]:<18}│
└─────────────────────────────────────┘
""")


if __name__ == '__main__':
    main()
