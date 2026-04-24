
# ============================================================
# FILE: src/excel_writer.py  (ENHANCED with new columns/sheets)
# ============================================================
"""Excel Report Generation - Enhanced with thumbnails, analytics, clusters."""

import csv
import logging
from pathlib import Path
from datetime import datetime
from collections import Counter

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XlImage
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

logger = logging.getLogger(__name__)


class ExcelWriter:
    ALL_COLS = [
        ('filename', 'Filename', 30), ('folder', 'Folder', 50),
        ('extension', 'Format', 10), ('file_type', 'Type', 8),
        ('size_mb', 'Size (MB)', 12),
        ('width', 'Width (px)', 12), ('height', 'Height (px)', 12),
        ('mode', 'Color Mode', 12), ('dpi', 'DPI', 10),
        ('date_taken', 'Date Taken', 20),
        ('camera_make', 'Camera Make', 16), ('camera_model', 'Camera Model', 18),
        ('focal_length', 'Focal Length', 14), ('aperture', 'Aperture', 12),
        ('iso', 'ISO', 8), ('exposure_time', 'Exposure', 12),
        ('gps_lat', 'GPS Lat', 12), ('gps_lon', 'GPS Lon', 12),
        ('location_name', 'Location', 25), ('location_country', 'Country', 10),
        ('has_exif', 'Has EXIF', 10),
        ('blur_score', 'Blur Score', 12), ('quality_rating', 'Quality', 12),
        ('quality_score', 'Quality %', 11), ('quality_issues', 'Issues', 35),
        ('is_blurry', 'Blurry?', 10),
        ('face_count', 'Faces', 8), ('face_category', 'Face Type', 14),
        ('auto_tags', 'Auto Tags', 30), ('primary_tag', 'Primary Tag', 14),
        ('cluster_label', 'Cluster', 14),
        ('video_duration_fmt', 'Duration', 12), ('video_fps', 'FPS', 8),
        ('video_codec', 'Codec', 12), ('video_bitrate_kbps', 'Bitrate (kbps)', 14),
        ('is_duplicate', 'Duplicate?', 12), ('duplicate_group', 'Dup Group', 12),
        ('is_best_in_group', 'Best?', 8), ('recommendation', 'Recommendation', 20),
        ('delete_flag', 'DELETE? (Yes/No)', 16),
        ('md5_hash', 'MD5 Hash', 36), ('file_modified', 'File Modified', 20),
        ('full_path', 'Full Path', 60), ('error', 'Read Error', 30),
    ]

    DUP_COLS = [
        ('duplicate_group', 'Group', 12), ('is_best_in_group', 'Best?', 8),
        ('recommendation', 'Recommendation', 20), ('filename', 'Filename', 30),
        ('folder', 'Folder', 50), ('extension', 'Format', 10),
        ('file_type', 'Type', 8), ('size_mb', 'Size (MB)', 12),
        ('quality_score', 'Quality %', 11), ('blur_score', 'Blur Score', 12),
        ('face_count', 'Faces', 8), ('auto_tags', 'Tags', 25),
        ('width', 'Width (px)', 12), ('height', 'Height (px)', 12),
        ('video_duration_fmt', 'Duration', 12),
        ('date_taken', 'Date Taken', 20),
        ('delete_flag', 'DELETE? (Yes/No)', 16),
        ('md5_hash', 'MD5 Hash', 36), ('full_path', 'Full Path', 60),
    ]

    BLUR_COLS = [
        ('filename', 'Filename', 30), ('folder', 'Folder', 50),
        ('blur_score', 'Blur Score', 12), ('quality_rating', 'Quality', 12),
        ('quality_score', 'Quality %', 11), ('quality_issues', 'Issues', 35),
        ('face_count', 'Faces', 8),
        ('width', 'Width (px)', 12), ('height', 'Height (px)', 12),
        ('size_mb', 'Size (MB)', 12), ('date_taken', 'Date Taken', 20),
        ('delete_flag', 'DELETE? (Yes/No)', 16), ('full_path', 'Full Path', 60),
    ]

    C_HDR = '2E4057'; C_DUP_H = '8B0000'; C_BLUR_H = 'CC6600'
    C_WF = 'FFFFFF'; C_ALT = 'F2F4F7'; C_DUP = 'FFD6D6'; C_BLUR = 'FFF0D0'
    C_G1 = 'FFE0E0'; C_G2 = 'FFF0F0'; C_BEST = '006400'
    C_SEC = '2E4057'; C_ACC = 'EBF2FF'; C_BRD = 'CCCCCC'

    def __init__(self, config):
        out = config.get('output', {})
        self.output_folder = Path(out.get('output_folder', './reports'))
        self.output_folder.mkdir(parents=True, exist_ok=True)
        self.prefix = out.get('filename_prefix', 'image_scan')
        sh = out.get('sheets', {})
        self.do_all = sh.get('all_images', True)
        self.do_blur = sh.get('blurry_images', True)
        self.do_dup = sh.get('duplicates', True)
        self.do_sum = sh.get('summary', True)
        self.do_qual = sh.get('quality_report', True)
        self.do_analytics = sh.get('analytics', True)
        self.do_clusters = sh.get('clusters', True)
        self.embed_thumbs = config.get('thumbnails', {}).get('embed_in_excel', False)
        self.config = config

    def write(self, records, scan_folder, analytics_data=None):
        if not XLSX_OK:
            return self._csv_fb(records) or ''

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        op = self.output_folder / f"{self.prefix}_{Path(scan_folder).name}_{ts}.xlsx"
        wb = openpyxl.Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)

        try:
            if self.do_sum:
                print("  Writing Summary..."); self._summary(wb, records, scan_folder); wb.save(op); print("  ✓")
            if self.do_all:
                print("  Writing All Images..."); self._all(wb, records); wb.save(op); print("  ✓")
            if self.do_blur:
                print("  Writing Blurry..."); self._blurry(wb, records); wb.save(op); print("  ✓")
            if self.do_dup:
                print("  Writing Duplicates..."); self._dups(wb, records); wb.save(op); print("  ✓")
            if self.do_qual:
                print("  Writing Quality..."); self._quality(wb, records); wb.save(op); print("  ✓")
            if self.do_analytics and analytics_data:
                print("  Writing Analytics..."); self._analytics_sheet(wb, analytics_data); wb.save(op); print("  ✓")
            if self.do_clusters:
                clustered = [r for r in records if r.get('cluster_label')]
                if clustered:
                    print("  Writing Clusters..."); self._clusters_sheet(wb, records); wb.save(op); print("  ✓")
            return str(op)
        except Exception as e:
            logger.error(f"Excel error: {e}", exc_info=True)
            try:
                wb.save(op)
            except Exception:
                pass
            fb = self._csv_fb(records)
            if op.exists():
                return str(op)
            if fb:
                return fb
            raise

    def _hdr(self, c, bg=None):
        c.font = Font(bold=True, color=self.C_WF, size=11)
        c.fill = PatternFill('solid', fgColor=bg or self.C_HDR)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = self._brd()

    def _brd(self):
        s = Side(style='thin', color=self.C_BRD)
        return Border(left=s, right=s, top=s, bottom=s)

    def _sv(self, v):
        if v is None: return ''
        if isinstance(v, datetime): return v.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(v, bool): return 'Yes' if v else 'No'
        if isinstance(v, float): return round(v, 2)
        if isinstance(v, tuple): return 'x'.join(str(x) for x in v)
        if isinstance(v, (list, dict)): return str(v)
        if isinstance(v, str): return ''.join(ch for ch in v if ord(ch) >= 32 or ch in '\n\t')
        return v

    def _all(self, wb, records):
        ws = wb.create_sheet('All Images')
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 35
        for ci, (_, l, w) in enumerate(self.ALL_COLS, 1):
            c = ws.cell(row=1, column=ci, value=l); self._hdr(c)
            ws.column_dimensions[get_column_letter(ci)].width = w
        af = PatternFill('solid', fgColor=self.C_ALT)
        df = PatternFill('solid', fgColor=self.C_DUP)
        bf = PatternFill('solid', fgColor=self.C_BLUR)
        for ri, rec in enumerate(records, 2):
            fl = bf if rec.get('is_blurry') is True else df if str(rec.get('is_duplicate','')).upper()=='YES' else af if ri%2==0 else None
            for ci, (k, _, _) in enumerate(self.ALL_COLS, 1):
                c = ws.cell(row=ri, column=ci, value=self._sv(rec.get(k, '')))
                c.border = self._brd(); c.alignment = Alignment(vertical='center')
                if fl: c.fill = fl
        ws.auto_filter.ref = f"A1:{get_column_letter(len(self.ALL_COLS))}{len(records)+1}"

    def _blurry(self, wb, records):
        ws = wb.create_sheet('Blurry Images')
        ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 35
        for ci, (_, l, w) in enumerate(self.BLUR_COLS, 1):
            c = ws.cell(row=1, column=ci, value=l); self._hdr(c, self.C_BLUR_H)
            ws.column_dimensions[get_column_letter(ci)].width = w
        bl = sorted([r for r in records if r.get('is_blurry') is True],
                    key=lambda x: x.get('blur_score', 0) if isinstance(x.get('blur_score'), (int, float)) else 0)
        fl = PatternFill('solid', fgColor=self.C_BLUR)
        for ri, rec in enumerate(bl, 2):
            for ci, (k, _, _) in enumerate(self.BLUR_COLS, 1):
                c = ws.cell(row=ri, column=ci, value=self._sv(rec.get(k, '')))
                c.border = self._brd(); c.alignment = Alignment(vertical='center'); c.fill = fl
        if bl: ws.auto_filter.ref = f"A1:{get_column_letter(len(self.BLUR_COLS))}{len(bl)+1}"

    def _dups(self, wb, records):
        ws = wb.create_sheet('Duplicates')
        ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 35
        for ci, (_, l, w) in enumerate(self.DUP_COLS, 1):
            c = ws.cell(row=1, column=ci, value=l); self._hdr(c, self.C_DUP_H)
            ws.column_dimensions[get_column_letter(ci)].width = w
        dups = sorted([r for r in records if str(r.get('is_duplicate','')).upper()=='YES'],
                      key=lambda x: (str(x.get('duplicate_group','')), x.get('full_path','')))
        pg = None; af = False
        f1 = PatternFill('solid', fgColor=self.C_G1); f2 = PatternFill('solid', fgColor=self.C_G2)
        bf = Font(bold=True, color=self.C_BEST); nf = Font()
        for ri, rec in enumerate(dups, 2):
            g = rec.get('duplicate_group','')
            if g != pg: af = not af; pg = g
            fl = f1 if af else f2
            ib = str(rec.get('is_best_in_group','')).lower() == 'yes'
            for ci, (k, _, _) in enumerate(self.DUP_COLS, 1):
                c = ws.cell(row=ri, column=ci, value=self._sv(rec.get(k, '')))
                c.border = self._brd(); c.alignment = Alignment(vertical='center')
                c.fill = fl; c.font = bf if ib else nf
        if dups: ws.auto_filter.ref = f"A1:{get_column_letter(len(self.DUP_COLS))}{len(dups)+1}"

    def _quality(self, wb, records):
        ws = wb.create_sheet('Quality Report')
        ws.column_dimensions['A'].width = 38; ws.column_dimensions['B'].width = 20; ws.column_dimensions['C'].width = 15
        ws['A1'] = 'Quality Analysis'; ws['A1'].font = Font(bold=True, size=16, color=self.C_HDR)
        qs = [r['quality_score'] for r in records if isinstance(r.get('quality_score'), (int, float))]
        bs = [r['blur_score'] for r in records if isinstance(r.get('blur_score'), (int, float))]
        aq = sum(qs)/len(qs) if qs else 0
        rows = [('','',''), ('QUALITY STATS','',''),
                ('Analyzed', len(qs), ''), ('Average', f"{aq:.1f}%", ''),
                ('Best', f"{max(qs):.1f}%" if qs else 'N/A', ''),
                ('Worst', f"{min(qs):.1f}%" if qs else 'N/A', ''),
                ('','',''), ('DISTRIBUTION','Count','%')]
        tq = len(qs) or 1
        for label, lo, hi in [('Excellent',80,101),('Good',60,80),('Fair',40,60),('Poor',0,40)]:
            n = sum(1 for s in qs if lo <= s < hi)
            rows.append((f'{label} ({lo}-{hi-1 if hi<101 else 100}%)', n, f"{n/tq*100:.1f}%"))

        # Face stats
        fc = [r.get('face_count', 0) for r in records if isinstance(r.get('face_count'), int)]
        if any(f > 0 for f in fc):
            rows.extend([('','',''), ('FACE DETECTION','Count','')])
            cats = Counter(r.get('face_category','') for r in records if r.get('face_category'))
            for cat, cnt in cats.most_common():
                if cat: rows.append((cat, cnt, ''))

        secs = {'QUALITY STATS','DISTRIBUTION','FACE DETECTION','TAGS','CLUSTERS'}
        hf = PatternFill('solid', fgColor=self.C_SEC); af = PatternFill('solid', fgColor=self.C_ACC)
        for ri, rd in enumerate(rows, 4):
            a = ws.cell(row=ri, column=1, value=rd[0])
            b = ws.cell(row=ri, column=2, value=rd[1])
            c = ws.cell(row=ri, column=3, value=rd[2])
            if rd[0] in secs:
                for x in (a,b,c): x.font = Font(bold=True, color=self.C_WF); x.fill = hf; x.border = self._brd()
            elif rd[0]:
                a.font = Font(bold=True, size=11)
                if ri%2==0:
                    for x in (a,b,c): x.fill = af

    def _summary(self, wb, records, sf):
        ws = wb.create_sheet('Summary', 0)
        ws.column_dimensions['A'].width = 35; ws.column_dimensions['B'].width = 30
        ws['A1'] = 'Image Scanner Summary'; ws['A1'].font = Font(bold=True, size=16, color=self.C_HDR)
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"; ws['A2'].font = Font(italic=True, color='888888')
        ws['A3'] = f"Scanned: {sf}"; ws['A3'].font = Font(italic=True, color='888888')

        t = len(records); imgs = sum(1 for r in records if r.get('file_type')=='image')
        vids = sum(1 for r in records if r.get('file_type')=='video')
        dc = sum(1 for r in records if str(r.get('is_duplicate','')).upper()=='YES')
        dg = len(set(r.get('duplicate_group') for r in records if r.get('duplicate_group') and str(r.get('duplicate_group')).strip()))
        bc = sum(1 for r in records if r.get('is_blurry') is True)
        ts = sum(r.get('size_mb',0) or 0 for r in records)
        uf = len(set(r.get('folder','') for r in records))
        ec = Counter(r.get('extension','?') for r in records)
        fc_total = sum(r.get('face_count', 0) for r in records if isinstance(r.get('face_count'), int))
        geo_count = sum(1 for r in records if r.get('location_name'))
        tag_count = sum(1 for r in records if r.get('auto_tags'))
        cluster_count = len(set(r.get('cluster_label') for r in records if r.get('cluster_label')))

        from .utils import format_duration
        vd = [r.get('video_duration_sec') for r in records if isinstance(r.get('video_duration_sec'),(int,float)) and r.get('video_duration_sec',0)>0]
        tvd = format_duration(sum(vd)) if vd else 'N/A'

        rows = [('',''), ('GENERAL',''), ('Total Files', t), ('Images', imgs), ('Videos', vids),
                ('Folders', uf), ('Total Size', f"{ts:.1f} MB ({ts/1024:.2f} GB)"),
                ('',''), ('VIDEO',''), ('Total Duration', tvd), ('Videos with Duration', len(vd)),
                ('',''), ('QUALITY',''), ('Blurry', bc), ('Duplicates', dc), ('Dup Groups', dg),
                ('',''), ('ENRICHMENT',''), ('Faces Detected', fc_total), ('Geolocated', geo_count),
                ('Auto-tagged', tag_count), ('Clusters', cluster_count),
                ('',''), ('BY FORMAT','')]
        for ext, cnt in sorted(ec.items(), key=lambda x: -x[1]):
            rows.append((f"  .{ext.lower()}", cnt))

        secs = {'GENERAL','VIDEO','QUALITY','ENRICHMENT','BY FORMAT'}
        hf = PatternFill('solid', fgColor=self.C_SEC); af = PatternFill('solid', fgColor=self.C_ACC)
        for ri, (l, v) in enumerate(rows, 5):
            a = ws.cell(row=ri, column=1, value=l); b = ws.cell(row=ri, column=2, value=v)
            if l in secs:
                for x in (a,b): x.font = Font(bold=True, color=self.C_WF, size=11); x.fill = hf; x.border = self._brd()
            elif l:
                a.font = Font(bold=True, size=11)
                if ri%2==0: a.fill = af; b.fill = af

    def _analytics_sheet(self, wb, data):
        ws = wb.create_sheet('Analytics')
        ws.column_dimensions['A'].width = 40; ws.column_dimensions['B'].width = 20; ws.column_dimensions['C'].width = 15
        ws['A1'] = 'Storage Analytics'; ws['A1'].font = Font(bold=True, size=16, color=self.C_HDR)
        rows = [('','',''), ('OVERVIEW','',''),
                ('Total Files', data.get('total_files',0), ''),
                ('Total Size', data.get('total_size_human',''), ''),
                ('Images', data.get('images',0), f"{data.get('image_size_mb',0):.0f} MB"),
                ('Videos', data.get('videos',0), f"{data.get('video_size_mb',0):.0f} MB"),
                ('Dup Savings', f"{data.get('duplicate_savings_mb',0):.1f} MB", ''),
                ('','',''), ('SIZE BY FORMAT','Size (MB)',''),]
        for fmt, sz in data.get('size_by_format', {}).items():
            rows.append((f"  .{fmt.lower()}", f"{sz:.1f}", ''))
        rows.extend([('','',''), ('BY YEAR','Count','Size (MB)')])
        for yr, cnt in data.get('by_year', {}).items():
            rows.append((str(yr), cnt, f"{data.get('size_by_year',{}).get(yr,0):.1f}"))
        rows.extend([('','',''), ('RESOLUTION','Count','')])
        for res, cnt in data.get('resolution_distribution', {}).items():
            rows.append((res, cnt, ''))
        rows.extend([('','',''), ('TOP 10 LARGEST','Size (MB)','')])
        for f in data.get('top_10_largest', []):
            rows.append((f.get('filename',''), f.get('size_mb',0), ''))

        secs = {'OVERVIEW','SIZE BY FORMAT','BY YEAR','RESOLUTION','TOP 10 LARGEST'}
        hf = PatternFill('solid', fgColor=self.C_SEC); af = PatternFill('solid', fgColor=self.C_ACC)
        for ri, rd in enumerate(rows, 3):
            a = ws.cell(row=ri, column=1, value=rd[0])
            b = ws.cell(row=ri, column=2, value=rd[1])
            c = ws.cell(row=ri, column=3, value=rd[2])
            if rd[0] in secs:
                for x in (a,b,c): x.font = Font(bold=True, color=self.C_WF); x.fill = hf; x.border = self._brd()
            elif rd[0]:
                a.font = Font(bold=True, size=11)
                if ri%2==0:
                    for x in (a,b,c): x.fill = af

    def _clusters_sheet(self, wb, records):
        ws = wb.create_sheet('Clusters')
        ws.column_dimensions['A'].width = 15; ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15; ws.column_dimensions['D'].width = 50
        ws['A1'] = 'Image Clusters'; ws['A1'].font = Font(bold=True, size=16, color=self.C_HDR)
        cols = [('cluster_label','Cluster',15),('filename','Filename',30),('quality_score','Quality %',12),
                ('auto_tags','Tags',30),('full_path','Full Path',50)]
        for ci, (_, l, w) in enumerate(cols, 1):
            c = ws.cell(row=3, column=ci, value=l); self._hdr(c)
            ws.column_dimensions[get_column_letter(ci)].width = w
        clustered = sorted([r for r in records if r.get('cluster_label')],
                          key=lambda x: (x.get('cluster_label',''), x.get('full_path','')))
        pg = None; af = False
        f1 = PatternFill('solid', fgColor='E8F0FE'); f2 = PatternFill('solid', fgColor='F8F9FA')
        for ri, rec in enumerate(clustered, 4):
            g = rec.get('cluster_label','')
            if g != pg: af = not af; pg = g
            fl = f1 if af else f2
            for ci, (k, _, _) in enumerate(cols, 1):
                c = ws.cell(row=ri, column=ci, value=self._sv(rec.get(k, '')))
                c.border = self._brd(); c.fill = fl

    def _csv_fb(self, records):
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            p = self.output_folder / f"{self.prefix}_{ts}.csv"
            hs = [l for _, l, _ in self.ALL_COLS]; ks = [k for k, _, _ in self.ALL_COLS]
            with open(p, 'w', newline='', encoding='utf-8-sig') as f:
                w = csv.writer(f); w.writerow(hs)
                for r in records: w.writerow([self._sv(r.get(k,'')) for k in ks])
            return str(p)
        except Exception:
            return None

