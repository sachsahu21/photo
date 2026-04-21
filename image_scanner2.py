"""
Image Scanner
-------------
1. Set SCAN_FOLDER below to your folder path
2. Run: python image_scanner.py
"""

import os
import hashlib
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter

# ──────────────────────────────────────────────
#  ✏️  EDIT THIS — your folder path
# ──────────────────────────────────────────────
SCAN_FOLDER = r"C:\Users\issuser\Desktop\Sachin\hdd"   # Windows example
# SCAN_FOLDER = "/Users/yourname/Photos"       # Mac/Linux example

#  Output file (saved in the same folder as this script)
OUTPUT_FILE = "meta"
# ──────────────────────────────────────────────


IMAGE_EXTENSIONS = {
    '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif',
    '.webp', '.heic', '.heif', '.raw', '.cr2', '.nef', '.arw',
    '.dng', '.orf', '.rw2', '.pef', '.svg', '.ico', '.psd',
    '.avif', '.jfif'
}

# ── Install check ──────────────────────────────
try:
    from PIL import Image
    from PIL.ExifTags import TAGS, GPSTAGS
    PIL_OK = True
except ImportError:
    PIL_OK = False
    print("WARNING: Pillow not found. Run: pip install pillow")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("WARNING: openpyxl not found. Run: pip install openpyxl")


# ── File hash for duplicate detection ─────────
def file_hash(filepath):
    h = hashlib.md5()
    try:
        with open(filepath, 'rb') as f:
            while chunk := f.read(65536):
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return None


# ── GPS extraction ─────────────────────────────
def get_gps(exif_data):
    try:
        gps = exif_data.get('GPSInfo', {})
        if not gps:
            return None, None
        def to_dec(v):
            return float(v[0]) + float(v[1]) / 60 + float(v[2]) / 3600
        lat = to_dec(gps.get(2, (0, 0, 0)))
        lon = to_dec(gps.get(4, (0, 0, 0)))
        if gps.get(1) == 'S': lat = -lat
        if gps.get(3) == 'W': lon = -lon
        return round(lat, 6), round(lon, 6)
    except Exception:
        return None, None


# ── Metadata extraction ────────────────────────
def get_metadata(filepath):
    meta = {
        'width': None, 'height': None, 'mode': None, 'dpi': None,
        'date_taken': None, 'camera_make': None, 'camera_model': None,
        'focal_length': None, 'aperture': None, 'iso': None,
        'exposure_time': None, 'gps_lat': None, 'gps_lon': None,
        'has_exif': False, 'error': None
    }
    if not PIL_OK:
        return meta
    try:
        with Image.open(filepath) as img:
            meta['width']  = img.width
            meta['height'] = img.height
            meta['mode']   = img.mode
            dpi = img.info.get('dpi')
            if dpi:
                meta['dpi'] = f"{int(dpi[0])}x{int(dpi[1])}"

            exif_raw = None
            if hasattr(img, '_getexif'):
                exif_raw = img._getexif()
            elif hasattr(img, 'getexif'):
                exif_raw = img.getexif()

            if exif_raw:
                meta['has_exif'] = True
                exif = {}
                for tag_id, val in exif_raw.items():
                    tag = TAGS.get(tag_id, tag_id)
                    if tag == 'GPSInfo' and isinstance(val, dict):
                        exif['GPSInfo'] = val
                    else:
                        exif[tag] = val

                meta['camera_make']  = (str(exif.get('Make',  '') or '')).strip() or None
                meta['camera_model'] = (str(exif.get('Model', '') or '')).strip() or None

                for dt_tag in ('DateTimeOriginal', 'DateTime', 'DateTimeDigitized'):
                    dt = exif.get(dt_tag)
                    if dt:
                        try:
                            meta['date_taken'] = datetime.strptime(str(dt), '%Y:%m:%d %H:%M:%S')
                        except Exception:
                            meta['date_taken'] = str(dt)
                        break

                fl = exif.get('FocalLength')
                if fl:
                    try: meta['focal_length'] = f"{float(fl):.1f}mm"
                    except: meta['focal_length'] = str(fl)

                fn = exif.get('FNumber')
                if fn:
                    try: meta['aperture'] = f"f/{float(fn):.1f}"
                    except: meta['aperture'] = str(fn)

                iso = exif.get('ISOSpeedRatings') or exif.get('PhotographicSensitivity')
                if iso:
                    meta['iso'] = str(iso)

                exp = exif.get('ExposureTime')
                if exp:
                    try:
                        fv = float(exp)
                        meta['exposure_time'] = f"1/{round(1/fv)}s" if fv < 1 else f"{fv}s"
                    except: meta['exposure_time'] = str(exp)

                meta['gps_lat'], meta['gps_lon'] = get_gps(exif)

    except Exception as e:
        meta['error'] = str(e)[:120]

    return meta


# ── Scan all images ────────────────────────────
def scan(folder):
    folder = Path(folder).expanduser().resolve()
    if not folder.exists():
        print(f"ERROR: Folder not found: {folder}")
        return []

    print(f"\nScanning: {folder}")

    all_files = [
        Path(dirpath) / fn
        for dirpath, _, filenames in os.walk(folder)
        for fn in filenames
        if Path(fn).suffix.lower() in IMAGE_EXTENSIONS
    ]

    print(f"Found {len(all_files)} images. Extracting metadata...\n")

    records  = []
    hash_map = defaultdict(list)

    for i, fp in enumerate(all_files, 1):
        print(f"  [{i}/{len(all_files)}] {fp.name}", end='\r')
        stat = fp.stat()
        md5  = file_hash(fp)
        meta = get_metadata(fp)

        rec = {
            'filename':      fp.name,
            'folder':        str(fp.parent),
            'full_path':     str(fp),
            'extension':     fp.suffix.lower().lstrip('.').upper(),
            'size_mb':       round(stat.st_size / (1024 * 1024), 2),
            'file_modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
            'md5_hash':      md5,
            **meta
        }
        records.append(rec)
        if md5:
            hash_map[md5].append(str(fp))

    # Mark duplicates
    dup_groups = {}
    gid = 1
    for md5, paths in hash_map.items():
        if len(paths) > 1:
            for p in paths:
                dup_groups[p] = gid
            gid += 1

    for r in records:
        r['is_duplicate']    = 'YES' if r['full_path'] in dup_groups else 'No'
        r['duplicate_group'] = dup_groups.get(r['full_path'], '')

    print()
    return records


# ── Excel writer ───────────────────────────────
COLUMNS = [
    ('filename',        'Filename',        28),
    ('folder',          'Folder',          45),
    ('extension',       'Format',          10),
    ('size_mb',         'Size (MB)',        11),
    ('width',           'Width (px)',       11),
    ('height',          'Height (px)',      11),
    ('mode',            'Color Mode',      12),
    ('dpi',             'DPI',             10),
    ('date_taken',      'Date Taken',      20),
    ('camera_make',     'Camera Make',     16),
    ('camera_model',    'Camera Model',    18),
    ('focal_length',    'Focal Length',    13),
    ('aperture',        'Aperture',        11),
    ('iso',             'ISO',              8),
    ('exposure_time',   'Exposure',        11),
    ('gps_lat',         'GPS Lat',         11),
    ('gps_lon',         'GPS Lon',         11),
    ('has_exif',        'Has EXIF',        10),
    ('is_duplicate',    'Duplicate?',      12),
    ('duplicate_group', 'Dup. Group',      10),
    ('md5_hash',        'MD5 Hash',        34),
    ('file_modified',   'File Modified',   20),
    ('full_path',       'Full Path',       55),
    ('error',           'Read Error',      25),
]

def _border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(color='2E4057'):
    return {
        'font':      Font(bold=True, color='FFFFFF', size=11),
        'fill':      PatternFill('solid', start_color=color),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

def write_excel(records, out_path):
    
    parent = Path(SCAN_FOLDER).name
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path(out_path) / f"image_scan_{parent}_{ts}.xlsx"
    
    wb = openpyxl.Workbook()

    # ── Sheet 1: All Images ──────────────────────
    ws = wb.active
    ws.title = 'All Images'
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 32

    for ci, (_, label, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=label)
        h = _hdr()
        c.font, c.fill, c.alignment = h['font'], h['fill'], h['alignment']
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    alt_fill = PatternFill('solid', start_color='F5F7FA')
    dup_fill = PatternFill('solid', start_color='FFD6D6')

    for ri, rec in enumerate(records, 2):
        fill = dup_fill if rec['is_duplicate'] == 'YES' else (alt_fill if ri % 2 == 0 else None)
        for ci, (key, _, _) in enumerate(COLUMNS, 1):
            val = rec.get(key, '')
            if isinstance(val, datetime): val = val.strftime('%Y-%m-%d %H:%M:%S')
            if isinstance(val, bool):     val = 'Yes' if val else 'No'
            if isinstance(val, str):
                val = ''.join(ch for ch in val if ord(ch) >= 32)

            # c = ws.cell(row=ri, column=ci, value=val)
            try:
                c = ws.cell(row=ri, column=ci, value=val)
            except Exception as e:
                print(f"\nERROR at row {ri}, column {key}: {repr(val)}")
                val = str(val)
                val = ''.join(ch for ch in val if ord(ch) >= 32)
                c = ws.cell(row=ri, column=ci, value=val)

            c.border = _border()
            c.alignment = Alignment(vertical='center')
            if fill: c.fill = fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Sheet 2: Duplicates ──────────────────────
    ws2 = wb.create_sheet('Duplicates')
    ws2.freeze_panes = 'A2'
    ws2.row_dimensions[1].height = 32

    dup_cols = [
        ('duplicate_group', 'Group',    8),
        ('filename',        'Filename', 28),
        ('folder',          'Folder',   45),
        ('extension',       'Format',   10),
        ('size_mb',         'Size (MB)',11),
        ('md5_hash',        'MD5 Hash', 34),
        ('full_path',       'Full Path',55),
    ]
    for ci, (_, label, width) in enumerate(dup_cols, 1):
        c = ws2.cell(row=1, column=ci, value=label)
        h = _hdr('8B0000')
        c.font, c.fill, c.alignment = h['font'], h['fill'], h['alignment']
        c.border = _border()
        ws2.column_dimensions[get_column_letter(ci)].width = width

    dups = sorted([r for r in records if r['is_duplicate'] == 'YES'],
                  key=lambda x: (x.get('duplicate_group', 0), x['full_path']))

    prev, alt_flag = None, False
    f1 = PatternFill('solid', start_color='FFE8E8')
    f2 = PatternFill('solid', start_color='FFF5F5')
    for ri, rec in enumerate(dups, 2):
        grp = rec.get('duplicate_group')
        if grp != prev:
            alt_flag = not alt_flag
            prev = grp
        fill = f1 if alt_flag else f2
        for ci, (key, _, _) in enumerate(dup_cols, 1):
            val = rec.get(key, '')
            if isinstance(val, datetime): val = val.strftime('%Y-%m-%d %H:%M:%S')
            c = ws2.cell(row=ri, column=ci, value=val)
            c.border = _border()
            c.alignment = Alignment(vertical='center')
            c.fill = fill
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(dup_cols))}1"

    # ── Sheet 3: Summary ─────────────────────────
    ws3 = wb.create_sheet('Summary')
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 20

    ws3['A1'] = 'Image Scan Summary'
    ws3['A1'].font = Font(bold=True, size=14, color='2E4057')
    ws3['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws3['A2'].font = Font(italic=True, color='888888')
    ws3['A3'] = f"Scanned: {SCAN_FOLDER}"
    ws3['A3'].font = Font(italic=True, color='888888')

    total     = len(records)
    dup_count = sum(1 for r in records if r['is_duplicate'] == 'YES')
    dup_grps  = len(set(r['duplicate_group'] for r in records if r['duplicate_group']))
    with_exif = sum(1 for r in records if r.get('has_exif'))
    with_gps  = sum(1 for r in records if r.get('gps_lat'))
    ext_cnts  = Counter(r['extension'] for r in records)

    rows = [
        ('', ''),
        ('GENERAL', ''),
        ('Total Images Found',    total),
        ('Total Folders Scanned', len(set(r['folder'] for r in records))),
        ('Total Size (MB)',        round(sum(r.get('size_mb', 0) for r in records) , 1)),
        ('', ''),
        ('DUPLICATES', ''),
        ('Duplicate Files',  dup_count),
        ('Duplicate Groups', dup_grps),
        ('', ''),
        ('METADATA', ''),
        ('Files with EXIF', with_exif),
        ('Files with GPS',  with_gps),
        ('', ''),
        ('BY FORMAT', ''),
        *[(f'  .{ext.lower()}', cnt) for ext, cnt in sorted(ext_cnts.items(), key=lambda x: -x[1])]
    ]

    hdr_fill = PatternFill('solid', start_color='2E4057')
    acc_fill = PatternFill('solid', start_color='EBF2FF')
    for ri, (label, value) in enumerate(rows, 5):
        a = ws3.cell(row=ri, column=1, value=label)
        b = ws3.cell(row=ri, column=2, value=value)
        if label in ('GENERAL', 'DUPLICATES', 'METADATA', 'BY FORMAT'):
            a.font = Font(bold=True, color='FFFFFF')
            a.fill = hdr_fill
            b.fill = hdr_fill
        elif label:
            a.font = Font(bold=True, size=11)
            if ri % 2 == 0:
                a.fill = acc_fill
                b.fill = acc_fill

    wb.save(out_path)
    print(f"Saved -> {out_path}")


# ── Main ───────────────────────────────────────
if __name__ == '__main__':

    if not PIL_OK or not OPENPYXL_OK:
        print("\nPlease install missing packages and re-run.")
        exit(1)

    records = scan(SCAN_FOLDER)

    if not records:
        print("No images found. Check your SCAN_FOLDER path.")
        exit(0)

    out = Path(__file__).parent/OUTPUT_FILE
    write_excel(records, str(out))

    dup_count = sum(1 for r in records if r['is_duplicate'] == 'YES')
    dup_grps  = len(set(r['duplicate_group'] for r in records if r['duplicate_group']))

    print(f"""
+--------------------------------------+
|          SCAN COMPLETE               |
+--------------------------------------+
  Total images    : {len(records)}
  Duplicate files : {dup_count}
  Duplicate groups: {dup_grps}
  Output file     : {out}
+--------------------------------------+
""")
