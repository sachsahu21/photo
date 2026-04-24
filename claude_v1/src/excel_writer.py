import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class ExcelWriter:
    COLUMNS = [
        ('filename',        'Filename',         28),
        ('folder',          'Folder',           45),
        ('file_type',       'Type',             10),
        ('extension',       'Format',           10),
        ('size_mb',         'Size (MB)',         11),
        ('width',           'Width (px)',        11),
        ('height',          'Height (px)',       11),
        ('mode',            'Color Mode',        12),
        ('dpi',             'DPI',               10),
        ('date_taken',      'Date Taken',        20),
        ('camera_make',     'Camera Make',       16),
        ('camera_model',    'Camera Model',      18),
        ('focal_length',    'Focal Length',      13),
        ('aperture',        'Aperture',          11),
        ('iso',             'ISO',                8),
        ('exposure_time',   'Exposure',          11),
        ('gps_lat',         'GPS Lat',           11),
        ('gps_lon',         'GPS Lon',           11),
        ('has_exif',        'Has EXIF',          10),
        ('blur_score',      'Blur Score',        11),
        ('quality_rating',  'Quality',           12),
        ('quality_score',   'Quality %',         10),
        ('quality_issues',  'Issues',            30),
        ('is_blurry',       'Blurry?',           10),
        ('is_duplicate',    'Duplicate?',        12),
        ('duplicate_group', 'Dup. Group',        10),
        ('is_best_in_group','Best?',              8),
        ('recommendation',  'Recommendation',    18),
        ('delete_flag',     'DELETE? (Yes/No)',  15),
        ('md5_hash',        'MD5 Hash',          34),
        ('file_modified',   'File Modified',     20),
        ('full_path',       'Full Path',         55),
        ('error',           'Read Error',        25),
    ]

    def __init__(self, config: Dict):
        self.config = config
        out = config.get('output', {})
        self.output_folder = Path(out.get('output_folder', './output'))
        self.output_folder.mkdir(parents=True, exist_ok=True)

    def write(self, records: List[Dict], scan_folder: str) -> str:
        parent = Path(scan_folder).name
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = self.output_folder / f"image_scan_{parent}_{ts}.xlsx"

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        print("  Writing All Images...")
        self._all_images(wb, records);      wb.save(out_path)
        print("  Writing Blurry Images...")
        self._blurry(wb, records);          wb.save(out_path)
        print("  Writing Duplicates...")
        self._duplicates(wb, records);      wb.save(out_path)
        print("  Writing Quality Report...")
        self._quality_report(wb, records);  wb.save(out_path)
        print("  Writing Summary...")
        self._summary(wb, records, scan_folder); wb.save(out_path)

        logger.info(f"Excel saved: {out_path}")
        return str(out_path)

    # ── sheets ──────────────────────────────────────────────
    def _all_images(self, wb, records):
        ws = wb.create_sheet('All Images', 0)
        ws.freeze_panes = 'A2'
        self._headers(ws, self.COLUMNS)
        af = PatternFill('solid', start_color='F5F7FA')
        df = PatternFill('solid', start_color='FFD6D6')
        bf = PatternFill('solid', start_color='FFE8B6')
        for ri, rec in enumerate(records, 2):
            fill = (bf if rec.get('is_blurry') is True
                    else df if rec.get('is_duplicate') == 'YES'
                    else af if ri % 2 == 0 else None)
            for ci, (key, _, _) in enumerate(self.COLUMNS, 1):
                c = ws.cell(ri, ci, self._fmt(rec.get(key)))
                c.border = self._bdr(); c.alignment = Alignment(vertical='center')
                if fill: c.fill = fill
        ws.auto_filter.ref = f"A1:{get_column_letter(len(self.COLUMNS))}1"

    def _blurry(self, wb, records):
        cols = [
            ('filename','Filename',28),('folder','Folder',45),
            ('blur_score','Blur Score',12),('quality_rating','Quality',12),
            ('quality_score','Quality %',10),('quality_issues','Issues',30),
            ('width','Width (px)',11),('height','Height (px)',11),
            ('size_mb','Size (MB)',11),('date_taken','Date Taken',20),
            ('delete_flag','DELETE? (Yes/No)',15),('full_path','Full Path',55),
        ]
        ws = wb.create_sheet('Blurry Images')
        ws.freeze_panes = 'A2'
        self._headers(ws, cols, 'FF8C00')
        fill = PatternFill('solid', start_color='FFE8B6')
        recs = sorted([r for r in records if r.get('is_blurry') is True],
                      key=lambda x: x.get('blur_score') or 0)
        for ri, rec in enumerate(recs, 2):
            for ci, (key, _, _) in enumerate(cols, 1):
                c = ws.cell(ri, ci, self._fmt(rec.get(key)))
                c.border = self._bdr(); c.fill = fill
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    def _duplicates(self, wb, records):
        cols = [
            ('duplicate_group','Group',8),('is_best_in_group','Best?',8),
            ('recommendation','Recommendation',18),('filename','Filename',28),
            ('folder','Folder',45),('extension','Format',10),
            ('size_mb','Size (MB)',11),('quality_score','Quality %',10),
            ('delete_flag','DELETE? (Yes/No)',15),('md5_hash','MD5 Hash',34),
            ('full_path','Full Path',55),
        ]
        ws = wb.create_sheet('Duplicates')
        ws.freeze_panes = 'A2'
        self._headers(ws, cols, '8B0000')
        f1 = PatternFill('solid', start_color='FFE8E8')
        f2 = PatternFill('solid', start_color='FFF5F5')
        dups = sorted([r for r in records if r.get('is_duplicate') == 'YES'],
                      key=lambda x: (x.get('duplicate_group') or 0, x.get('full_path','')))
        prev, alt = None, False
        for ri, rec in enumerate(dups, 2):
            grp = rec.get('duplicate_group')
            if grp != prev: alt = not alt; prev = grp
            fill = f1 if alt else f2
            for ci, (key, _, _) in enumerate(cols, 1):
                c = ws.cell(ri, ci, self._fmt(rec.get(key)))
                c.border = self._bdr(); c.fill = fill
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    def _quality_report(self, wb, records):
        ws = wb.create_sheet('Quality Report')
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws['A1'] = 'Quality Analysis Report'
        ws['A1'].font = Font(bold=True, size=14, color='2E4057')

        img_recs = [r for r in records if r.get('file_type') == 'image']
        scores = [r.get('quality_score', 0) for r in img_recs if isinstance(r.get('quality_score'), (int, float))]
        avg = sum(scores)/len(scores) if scores else 0

        hf = PatternFill('solid', start_color='2E4057')
        rows = [
            ('',''),('QUALITY STATISTICS',''),
            ('Average Quality Score', f"{avg:.1f}%"),
            ('Highest Quality', f"{max(scores):.1f}%" if scores else 'N/A'),
            ('Lowest Quality',  f"{min(scores):.1f}%" if scores else 'N/A'),
            ('',''),('QUALITY DISTRIBUTION',''),
            ('Excellent (80-100%)', sum(1 for s in scores if s >= 80)),
            ('Good (60-79%)',       sum(1 for s in scores if 60 <= s < 80)),
            ('Fair (40-59%)',       sum(1 for s in scores if 40 <= s < 60)),
            ('Poor (0-39%)',        sum(1 for s in scores if s < 40)),
        ]
        for ri, (label, val) in enumerate(rows, 3):
            a = ws.cell(ri, 1, label); b = ws.cell(ri, 2, val)
            if label in ('QUALITY STATISTICS','QUALITY DISTRIBUTION'):
                a.font = Font(bold=True, color='FFFFFF'); a.fill = hf; b.fill = hf
            elif label:
                a.font = Font(bold=True)

    def _summary(self, wb, records, scan_folder):
        ws = wb.create_sheet('Summary', 0)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws['A1'] = 'Image Scan Summary'
        ws['A1'].font = Font(bold=True, size=14, color='2E4057')
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, color='888888')
        ws['A3'] = f"Scanned: {scan_folder}"
        ws['A3'].font = Font(italic=True, color='888888')

        total   = len(records)
        images  = [r for r in records if r.get('file_type') == 'image']
        videos  = [r for r in records if r.get('file_type') == 'video']
        dups    = sum(1 for r in records if r.get('is_duplicate') == 'YES')
        dup_grps= len(set(r.get('duplicate_group') for r in records if r.get('duplicate_group')))
        blurry  = sum(1 for r in records if r.get('is_blurry') is True)
        exif    = sum(1 for r in records if r.get('has_exif'))
        gps     = sum(1 for r in records if r.get('gps_lat'))
        ext_cnt = Counter(r.get('extension','?') for r in records)
        scores  = [r.get('quality_score',0) for r in images if isinstance(r.get('quality_score'),(int,float))]
        avg_q   = sum(scores)/len(scores) if scores else 0
        tot_mb  = sum(r.get('size_mb',0) or 0 for r in records)

        hf = PatternFill('solid', start_color='2E4057')
        af = PatternFill('solid', start_color='EBF2FF')
        rows = [
            ('',''),('GENERAL',''),
            ('Total Files',          total),
            ('Images',               len(images)),
            ('Videos',               len(videos)),
            ('Total Folders',        len(set(r.get('folder','') for r in records))),
            ('Total Size (MB)',       round(tot_mb, 1)),
            ('Average Quality Score', f"{avg_q:.1f}%"),
            ('',''),('QUALITY ISSUES',''),
            ('Blurry Images',   blurry),
            ('Duplicate Files', dups),
            ('Duplicate Groups',dup_grps),
            ('',''),('METADATA',''),
            ('Files with EXIF', exif),
            ('Files with GPS',  gps),
            ('',''),('BY FORMAT',''),
            *[(f'  .{e.lower()}', c) for e, c in sorted(ext_cnt.items(), key=lambda x: -x[1])]
        ]
        for ri, (label, val) in enumerate(rows, 5):
            a = ws.cell(ri, 1, label); b = ws.cell(ri, 2, val)
            if label in ('GENERAL','QUALITY ISSUES','METADATA','BY FORMAT'):
                a.font = Font(bold=True, color='FFFFFF'); a.fill = hf; b.fill = hf
            elif label:
                a.font = Font(bold=True, size=11)
                if ri % 2 == 0: a.fill = af; b.fill = af

    # ── helpers ─────────────────────────────────────────────
    def _headers(self, ws, cols, color='2E4057'):
        ws.row_dimensions[1].height = 32
        for ci, (_, label, width) in enumerate(cols, 1):
            c = ws.cell(1, ci, label)
            c.font      = Font(bold=True, color='FFFFFF', size=11)
            c.fill      = PatternFill('solid', start_color=color)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border    = self._bdr()
            ws.column_dimensions[get_column_letter(ci)].width = width

    def _bdr(self):
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    def _fmt(self, val):
        if isinstance(val, datetime): return val.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(val, bool):     return 'Yes' if val else 'No'
        if isinstance(val, str):      return ''.join(c for c in val if ord(c) >= 32)
        return val
