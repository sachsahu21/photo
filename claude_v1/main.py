"""
Photo Library Scanner - Main Entry Point
"""
import sys, logging
from pathlib import Path
from typing import Optional

from src.config_manager import ConfigManager
from src.scanner import ImageScanner
from src.duplicate_handler import DuplicateHandler
from src.organizer import ImageOrganizer
from src.excel_writer import ExcelWriter
from src.main_utils import setup_logging, save_backup, load_backup


def print_banner():
    print("""
╔══════════════════════════════════════════════╗
║   Photo Library Scanner  v2.0               ║
║   Scan · Detect · Organise                  ║
╚══════════════════════════════════════════════╝""")

def print_menu():
    print("""
┌──────────────────────────────────────────────┐
│  1. Scan & Extract Metadata                  │
│  1b. Resume Excel write (from backup)        │
│  2. Delete Marked Files (from Excel)         │
│  3. Organise Images by Date                  │
│  4. Full Workflow  (1 → 2 → 3)               │
│  5. Exit                                     │
└──────────────────────────────────────────────┘""")


# ─────────────────────────────────────────────
# TASK 1
# ─────────────────────────────────────────────
def task_1(config, logger) -> Optional[str]:
    print("\n" + "="*50)
    print("TASK 1: SCAN & EXTRACT METADATA")
    print("="*50)
    records = []
    try:
        scanner     = ImageScanner(config.to_dict())
        scan_folder = config.get('scan.folder_path')
        records     = scanner.scan(scan_folder)

        if not records:
            print("⚠ No files found"); return None

        save_backup(records)

        dup_handler = DuplicateHandler(
            selection_criteria=config.get('duplicates.selection_criteria',
                                          ['quality','resolution','date','size'])
        )
        records = dup_handler.mark_duplicates(records)
        save_backup(records)

        blurry = sum(1 for r in records if r.get('is_blurry') is True)
        dups   = sum(1 for r in records if r.get('is_duplicate') == 'YES')
        grps   = len(set(r.get('duplicate_group') for r in records if r.get('duplicate_group')))

        print(f"\n✓ Scanned:    {len(records)} files")
        print(f"✓ Blurry:     {blurry}")
        print(f"✓ Duplicates: {dups} in {grps} groups")

        writer     = ExcelWriter(config.to_dict())
        excel_path = writer.write(records, scan_folder)
        print(f"\n✓ Excel: {excel_path}")
        return excel_path

    except Exception as e:
        logger.error(f"Task 1 error: {e}", exc_info=True)
        print(f"✗ Error: {e}")
        if records:
            save_backup(records)
        return None


# ─────────────────────────────────────────────
# TASK 1b – resume from backup
# ─────────────────────────────────────────────
def task_1b(config, logger) -> Optional[str]:
    print("\n" + "="*50)
    print("TASK 1b: RESUME EXCEL WRITE FROM BACKUP")
    print("="*50)
    records = load_backup()
    if not records: return None
    try:
        writer = ExcelWriter(config.to_dict())
        path   = writer.write(records, config.get('scan.folder_path'))
        print(f"✓ Excel: {path}")
        return path
    except Exception as e:
        logger.error(f"Task 1b error: {e}", exc_info=True)
        print(f"✗ Error: {e}"); return None


# ─────────────────────────────────────────────
# TASK 2 – delete
# ─────────────────────────────────────────────
def task_2(excel_path: str, logger):
    print("\n" + "="*50)
    print("TASK 2: DELETE MARKED FILES")
    print("="*50)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['All Images']

        col = {}
        for ci, cell in enumerate(ws[1], 1):
            if cell.value:
                col[str(cell.value).strip().lower()] = ci

        del_col  = col.get('delete? (yes/no)')
        path_col = col.get('full path')
        name_col = col.get('filename')

        if not del_col:
            print("✗ 'DELETE? (Yes/No)' column not found"); return

        deleted = errors = 0
        for ri in range(2, ws.max_row + 1):
            flag = ws.cell(ri, del_col).value
            if str(flag).strip().lower() in ('yes','true','1'):
                fp   = ws.cell(ri, path_col).value if path_col else None
                name = ws.cell(ri, name_col).value if name_col else fp
                if not fp:
                    print(f"⚠ No path for row {ri}"); continue
                try:
                    p = Path(fp)
                    if p.exists():
                        p.unlink(); deleted += 1
                        print(f"✓ Deleted: {name}")
                    else:
                        print(f"⚠ Not found: {name}")
                except Exception as e:
                    errors += 1; print(f"✗ Error: {name} – {e}")

        print(f"\n✓ Deleted: {deleted}  Errors: {errors}")

    except Exception as e:
        logger.error(f"Task 2 error: {e}", exc_info=True)
        print(f"✗ Error: {e}")


# ─────────────────────────────────────────────
# TASK 3 – organise
# ─────────────────────────────────────────────
def task_3(excel_path: str, config, logger):
    print("\n" + "="*50)
    print("TASK 3: ORGANISE IMAGES BY DATE")
    print("="*50)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['All Images']

        col_map = {cell.value: ci for ci, cell in enumerate(ws[1], 1) if cell.value}
        records = []
        for ri in range(2, ws.max_row + 1):
            records.append({hdr: ws.cell(ri, ci).value for hdr, ci in col_map.items()})

        organizer = ImageOrganizer(config.to_dict())
        movements = organizer.organize(records)

        ok  = sum(1 for m in movements if m['status'] == 'Success')
        err = sum(1 for m in movements if 'Error' in m['status'])
        print(f"\n✓ Organised: {ok}  Errors: {err}")
        print(f"✓ Output:    {config.get('organization.output_folder')}")

    except Exception as e:
        logger.error(f"Task 3 error: {e}", exc_info=True)
        print(f"✗ Error: {e}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    print_banner()
    config = ConfigManager()
    logger = setup_logging(config)
    logger.info("Started")

    while True:
        print_menu()
        choice = input("Choice (1/1b/2/3/4/5): ").strip().lower()

        if choice == '1':
            task_1(config, logger)

        elif choice == '1b':
            task_1b(config, logger)

        elif choice == '2':
            ep = input("Excel path: ").strip()
            if Path(ep).exists(): task_2(ep, logger)
            else: print(f"✗ Not found: {ep}")

        elif choice == '3':
            ep = input("Excel path: ").strip()
            if Path(ep).exists(): task_3(ep, config, logger)
            else: print(f"✗ Not found: {ep}")

        elif choice == '4':
            ep = task_1(config, logger)
            if ep:
                input("\nPress Enter for Task 2 (delete)...")
                task_2(ep, logger)
                input("\nPress Enter for Task 3 (organise)...")
                task_3(ep, config, logger)

        elif choice == '5':
            print("✓ Goodbye!"); break
        else:
            print("✗ Invalid choice")

        input("\nPress Enter to continue...")

if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n✗ Interrupted")
        sys.exit(1)
