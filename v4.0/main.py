"""Image Scanner v4.0"""

import sys
import logging
import pickle
from pathlib import Path

from src.config_manager import ConfigManager
from src.scanner import ImageScanner
from src.duplicate_handler import DuplicateHandler
from src.organizer import ImageOrganizer
from src.excel_writer import ExcelWriter

BK = 'records-backup.pkl'


def setup_log(config):
    lf = config.get('logging.file', './logs/image-scanner.log')
    Path(lf).parent.mkdir(parents=True, exist_ok=True)
    handlers = [logging.FileHandler(lf, encoding='utf-8')]
    if config.get('logging.console', True):
        handlers.append(logging.StreamHandler())
    logging.basicConfig(
        level=getattr(logging, str(config.get('logging.level', 'INFO')).upper(), logging.INFO),
        format='%(asctime)s-%(name)s-%(levelname)s-%(message)s',
        handlers=handlers
    )
    return logging.getLogger(__name__)


def save_bk(records):
    try:
        with open(BK, 'wb') as f:
            pickle.dump(records, f, protocol=pickle.HIGHEST_PROTOCOL)
        print('  Backup: ' + str(len(records)) + ' records')
    except Exception as e:
        print('  Backup failed: ' + str(e))


def load_bk():
    if not Path(BK).exists():
        print('  No backup')
        return None
    try:
        with open(BK, 'rb') as f:
            r = pickle.load(f)
        print('  Loaded ' + str(len(r)) + ' records')
        return r
    except Exception as e:
        print('  Load error: ' + str(e))
        return None


def task_1(config, logger):
    print('\n  TASK 1: SCAN')
    print('  ' + '=' * 50)
    records = None
    try:
        scanner = ImageScanner(config.to_dict())
        sf = config.get('scan.folder_path')
        print('  Scanning: ' + str(sf))
        records = scanner.scan(sf)
        if not records:
            print('  No files!')
            return None
        print('  Found ' + str(len(records)) + ' files')
        save_bk(records)
        if config.get('duplicates.enabled', True):
            print('  Detecting duplicates...')
            records = DuplicateHandler(
                hash_algorithm=config.get('duplicates.hash_algorithm', 'md5'),
                selection_criteria=config.get(
                    'duplicates.selection_criteria',
                    ['quality', 'resolution', 'date', 'size']
                )
            ).mark_duplicates(records)
        if config.get('similar_detection.enabled', False):
            print('  Detecting similar...')
            from src.similar_detector import SimilarDetector
            sd = SimilarDetector(config.to_dict())
            records = sd.compute_hashes(records)
            sg = sd.find_similar(records)
            records = sd.mark_similar(records, sg)
            sc = sum(1 for r in records if str(r.get('is_similar', '')).upper() == 'YES')
            print('  Similar: ' + str(sc) + ' in ' + str(len(sg)) + ' groups')
        if config.get('clustering.enabled', False):
            try:
                from src.image_clusterer import ImageClusterer
                records = ImageClusterer(
                    n_clusters=config.get('clustering.n_clusters', 10)
                ).cluster(records)
            except Exception:
                pass
        save_bk(records)
        ad = None
        if config.get('analytics.enabled', True):
            try:
                from src.analytics import StorageAnalytics
                ad = StorageAnalytics().analyze(records)
            except Exception:
                pass
        dc = sum(1 for r in records if str(r.get('is_duplicate', '')).upper() == 'YES')
        bc = sum(1 for r in records if r.get('is_blurry') is True)
        sc = sum(1 for r in records if str(r.get('is_similar', '')).upper() == 'YES')
        print('  Blurry: ' + str(bc) + ' | Dups: ' + str(dc) + ' | Similar: ' + str(sc))
        if config.get('comparison.enabled', True):
            try:
                from src.comparison_generator import ComparisonGenerator
                pg = ComparisonGenerator(
                    config.get('comparison.output_folder', './comparisons')
                ).generate(records)
                if pg:
                    print('  Comparisons: ' + str(len(pg)))
            except Exception:
                pass
        print('\n  Generating Excel...')
        ep = ExcelWriter(config.to_dict()).write(records, sf, analytics_data=ad)
        print('  Excel: ' + str(ep))
        return ep
    except Exception as e:
        logger.error('Task 1: %s', e, exc_info=True)
        print('  Error: ' + str(e))
        if records:
            save_bk(records)
        return None


def task_1b(config, logger):
    print('\n  TASK 1B: RESUME')
    records = load_bk()
    if not records:
        return None
    ad = None
    try:
        from src.analytics import StorageAnalytics
        ad = StorageAnalytics().analyze(records)
    except Exception:
        pass
    ep = ExcelWriter(config.to_dict()).write(
        records, config.get('scan.folder_path'), analytics_data=ad
    )
    print('  Excel: ' + str(ep))
    return ep


def task_2(ep, logger):
    print('\n  TASK 2: DELETE')
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ep)
        sn = next((n for n in ['Duplicates', 'All Images'] if n in wb.sheetnames), None)
        if not sn:
            print('  No sheet')
            return
        ws = wb[sn]
        dci = fci = nci = None
        for ci, c in enumerate(ws[1], 1):
            if not c.value:
                continue
            h = str(c.value).strip().lower()
            if 'delete' in h:
                dci = ci
            elif 'full path' in h or h == 'path':
                fci = ci
            elif h in ('filename', 'file'):
                nci = ci
        if not dci or not fci:
            print('  Columns not found')
            return
        td = []
        for ri in range(2, ws.max_row + 1):
            v = ws.cell(row=ri, column=dci).value
            if v and str(v).strip().lower() in ('yes', 'true', '1'):
                td.append((
                    ws.cell(row=ri, column=fci).value,
                    ws.cell(row=ri, column=nci).value if nci else '?'
                ))
        if not td:
            print('  Nothing to delete')
            return
        print('  ' + str(len(td)) + ' files marked')
        if input('  Confirm? (yes/no): ').strip().lower() != 'yes':
            return
        ok = er = nf = 0
        for fp, fn in td:
            if not fp:
                er += 1
                continue
            try:
                p = Path(fp)
                if p.exists():
                    p.unlink()
                    ok += 1
                else:
                    nf += 1
            except Exception:
                er += 1
        print('  Deleted: ' + str(ok) + ' | Missing: ' + str(nf) + ' | Errors: ' + str(er))
    except Exception as e:
        print('  Error: ' + str(e))


def task_3(ep, config, logger):
    print('\n  TASK 3: ORGANIZE')
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ep)
        sn = next((n for n in ['All Images', 'Duplicates'] if n in wb.sheetnames), None)
        if not sn:
            print('  No sheet')
            return
        ws = wb[sn]
        cm = {str(c.value).strip(): ci for ci, c in enumerate(ws[1], 1) if c.value}
        h2k = {
            'Filename': 'filename', 'File': 'filename', 'Folder': 'folder',
            'Full Path': 'full_path', 'Path': 'full_path',
            'Date Taken': 'date_taken', 'Date': 'date_taken',
            'File Modified': 'file_modified',
            'DELETE?': 'delete_flag', 'Dup?': 'is_duplicate',
            'Type': 'file_type', 'Format': 'extension', 'Fmt': 'extension',
            'Size MB': 'size_mb', 'Size': 'size_mb',
            'Width': 'width', 'W': 'width',
            'Height': 'height', 'H': 'height',
            'EXIF': 'has_exif', 'Metadata': 'metadata_status',
        }
        records = []
        for ri in range(2, ws.max_row + 1):
            rec = {}
            for h, ci in cm.items():
                v = ws.cell(row=ri, column=ci).value
                rec[h2k.get(h, h.lower().replace(' ', '_'))] = v
                rec[h] = v
            records.append(rec)
        print('  Loaded ' + str(len(records)) + ' records')
        mvs = ImageOrganizer(config.to_dict()).organize(records)
        s = sum(1 for m in mvs if m['status'] == 'Success')
        e = sum(1 for m in mvs if 'Error' in m['status'])
        print('  Done: ' + str(s) + ' success, ' + str(e) + ' errors')
    except Exception as e:
        print('  Error: ' + str(e))


def task_7(config, logger):
    print('\n  TASK 7: CONVERT FOLDER STRUCTURE')
    print('  ' + '=' * 50)
    source = config.get('organization.output_folder', './organized_images')
    print('  Source folder: ' + str(source))
    if not Path(source).exists():
        print('  Error: Folder not found!')
        return
    print('')
    print('  Structures:')
    print('')
    print('    1. flat')
    print('       Organized/')
    print('         2022-01-30-020pic-beach/')
    print('         2022-03-00-012pic/')
    print('         2024-03-screenshots-045pic/')
    print('         undated-005pic/')
    print('')
    print('    2. year')
    print('       Organized/')
    print('         2022/')
    print('           2022-01-30-020pic-beach/')
    print('           2022-03-00-012pic/')
    print('         2024/')
    print('           2024-03-screenshots-045pic/')
    print('         undated-005pic/')
    print('')
    print('    3. year-month-date')
    print('       Organized/')
    print('         2022/')
    print('           01-Jan/')
    print('             2022-01-30-020pic-beach/')
    print('           03-Mar/')
    print('             2022-03-00-012pic/')
    print('         2024/')
    print('           03-Mar/')
    print('             2024-03-screenshots-045pic/')
    print('         undated-005pic/')
    print('')
    ch = input('  Select target (1/2/3): ').strip()
    targets = {'1': 'flat', '2': 'year', '3': 'year-month-date'}
    target = targets.get(ch)
    if not target:
        print('  Invalid choice')
        return
    print('')
    print('  Target: ' + target)
    out = input('  Output folder (Enter=in-place): ').strip()
    target_folder = out if out else None
    print('')
    if input('  Confirm? (yes/no): ').strip().lower() != 'yes':
        print('  Cancelled')
        return
    print('')
    ImageOrganizer.convert_structure(source, target, target_folder)


def task_8(config, logger):
    print('\n  TASK 8: MERGE SAME-DATE FOLDERS')
    print('  ' + '=' * 50)
    source = config.get('organization.output_folder', './organized_images')
    print('  Source folder: ' + str(source))
    if not Path(source).exists():
        print('  Error: Folder not found!')
        return
    detected = ImageOrganizer.detect_structure(source)
    print('  Detected structure: ' + detected)
    override = input('  Structure override [flat/year/year-month-date] (Enter=detected): ').strip().lower()
    structure = override if override in ('flat', 'year', 'year-month-date') else detected
    print('  Using structure: ' + structure)
    print('')
    print('  This will find folders with same date like:')
    print('    2012-11-00-001pic-singapore')
    print('    2012-11-00-002pic')
    print('  And merge them into:')
    print('    2012-11-00-003pic-singapore')
    print('')
    if input('  Confirm? (yes/no): ').strip().lower() != 'yes':
        print('  Cancelled')
        return
    print('')
    ImageOrganizer.merge_duplicate_dates(source, structure)


def main():
    try:
        print('')
        print('  ========================================================')
        print('       IMAGE SCANNER v4.0')
        print('  ========================================================')
        config = ConfigManager()
        logger = setup_log(config)
        print('  Config: ' + str(config.config_path))
        print('  Scan: ' + str(config.get('scan.folder_path')))
        print('  Structure: ' + str(config.get('organization.folder_structure', 'year')))
        ft = []
        for k, n in [
            ('duplicates.enabled', 'duplicates'),
            ('similar_detection.enabled', 'similar'),
            ('blur_detection.enabled', 'blur'),
            ('face_detection.enabled', 'faces'),
            ('processing.fast_mode', 'FAST'),
        ]:
            if config.get(k):
                ft.append(n)
        if ft:
            print('  Features: ' + ', '.join(ft))
        if not config.validate():
            if input('  Continue? (yes/no): ').strip().lower() != 'yes':
                return
        last = None
        while True:
            print('')
            print('  1.  Scan & Extract')
            print('  1b. Resume Excel')
            print('  2.  Delete Marked')
            print('  3.  Organize')
            print('  4.  Full (1>2>3)')
            print('  5.  Web Dashboard')
            print('  6.  Comparisons')
            print('  7.  Convert Folder Structure')
            print('  8.  Merge Same-Date Folders')
            print('  0.  Exit')
            print('')
            ch = input('  Choice: ').strip().lower()
            if ch == '1':
                last = task_1(config, logger)
            elif ch == '1b':
                last = task_1b(config, logger)
            elif ch == '2':
                p = input('  Excel (Enter=last): ').strip() or last
                if p and Path(p).exists():
                    task_2(p, logger)
                else:
                    print('  Not found')
            elif ch == '3':
                p = input('  Excel (Enter=last): ').strip() or last
                if p and Path(p).exists():
                    task_3(p, config, logger)
                else:
                    print('  Not found')
            elif ch == '4':
                last = task_1(config, logger)
                if last:
                    if input('  Delete? (yes/no): ').strip().lower() == 'yes':
                        task_2(last, logger)
                    if input('  Organize? (yes/no): ').strip().lower() == 'yes':
                        task_3(last, config, logger)
            elif ch == '5':
                try:
                    import subprocess
                    ap = Path(__file__).parent / 'web' / 'streamlit_app.py'
                    if ap.exists():
                        subprocess.Popen([sys.executable, '-m', 'streamlit', 'run', str(ap)])
                        print('  http://localhost:8501')
                    else:
                        print('  Not found')
                except Exception as e:
                    print('  Error: ' + str(e))
            elif ch == '6':
                recs = load_bk()
                if recs:
                    try:
                        from src.comparison_generator import ComparisonGenerator
                        pg = ComparisonGenerator(
                            config.get('comparison.output_folder', './comparisons')
                        ).generate(recs)
                        print('  Generated ' + str(len(pg)) + ' pages')
                    except Exception as e:
                        print('  Error: ' + str(e))
            elif ch == '7':
                task_7(config, logger)
            elif ch == '8':
                task_8(config, logger)
            elif ch == '0':
                print('\n  Bye!')
                break
            else:
                print('  Invalid')
            input('\n  Enter to continue...')
    except KeyboardInterrupt:
        print('\n  Interrupted')
        sys.exit(1)
    except Exception as e:
        print('\n  Fatal: ' + str(e))
        sys.exit(1)


if __name__ == '__main__':
    main()

