# import required modules
import csv
from pathlib import Path
from typing import Set
from datetime import datetime

from src.config_manager import ConfigManager

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def collect_files(base_path: Path, recursive: bool, extensions: dict) -> Set[str]:
    """Return a set of file paths (as strings) matching the given extensions.
    extensions dict looks like {'images': ['jpg', 'jpeg'], 'videos': ['mp4', ...]}
    """
    ext_set = {f".{e.lower()}" for cat in extensions.values() for e in cat}
    if recursive:
        files = {str(p) for p in base_path.rglob("*") if p.is_file() and p.suffix.lower() in ext_set}
    else:
        files = {str(p) for p in base_path.iterdir() if p.is_file() and p.suffix.lower() in ext_set}
    return files


def load_global_checkpoint(ck_path: Path) -> Set[str]:
    if not ck_path.is_file():
        return set()
    try:
        import json
        data = json.loads(ck_path.read_text(encoding='utf-8'))
        return set(data.get('processed', []))
    except Exception:
        return set()


def human_readable_size(num_bytes: int) -> str:
    """Convert bytes to a human-readable string."""
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if num_bytes < 1024:
            return f"{num_bytes:.2f} {unit}"
        num_bytes /= 1024
    return f"{num_bytes:.2f} PB"


def _safe_size(path_str: str) -> int:
    try:
        return Path(path_str).stat().st_size
    except Exception:
        return 0


def _style_header(ws, row: int = 1):
    fill = PatternFill("solid", fgColor="BDD7EE")
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')


def main():
    cfg = ConfigManager().to_dict()
    workspace_root = Path(cfg['workspace'].get('_resolved_root') or cfg['workspace']['root'])
    scan_cfg = cfg['scan']
    scan_path = Path(scan_cfg['folder_path'])
    recursive = bool(scan_cfg.get('recursive', True))
    extensions = scan_cfg.get('extensions', {})

    all_files = collect_files(scan_path, recursive, extensions)
    ck_path = Path(
        cfg.get('processing', {}).get('global_checkpoint_file')
        or workspace_root / "global_checkpoint.json"
    )
    processed = load_global_checkpoint(ck_path)

    # Intersect checkpoint with actual files — removes ghost entries from old runs
    scanned = all_files & processed
    pending = all_files - processed

    total_count = len(all_files)
    scanned_count = len(scanned)
    pending_count = len(pending)

    pending_size_bytes = sum(_safe_size(f) for f in pending)
    scanned_size_bytes = sum(_safe_size(f) for f in scanned)
    pending_size_human = human_readable_size(pending_size_bytes)
    scanned_size_human = human_readable_size(scanned_size_bytes)

    # Per-directory breakdown
    dir_stats: dict = {}
    for f in all_files:
        d = str(Path(f).parent)
        if d not in dir_stats:
            dir_stats[d] = {'total': 0, 'scanned': 0, 'pending': 0, 'pending_bytes': 0}
        dir_stats[d]['total'] += 1
        if f in scanned:
            dir_stats[d]['scanned'] += 1
        else:
            dir_stats[d]['pending'] += 1
            dir_stats[d]['pending_bytes'] += _safe_size(f)

    reports_dir = workspace_root / "folder_analysis"
    reports_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # --- CSV: backward-compatible flat list ---
    csv_path = reports_dir / ("scanned_pending_report_" + ts + ".csv")
    with csv_path.open('w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["status", "full_path"])
        for f in sorted(scanned):
            writer.writerow(["scanned", f])
        for f in sorted(pending):
            writer.writerow(["pending", f])

    # --- XLSX: multi-sheet detailed report ---
    xlsx_path = reports_dir / ("scanned_pending_report_" + ts + ".xlsx")
    wb = Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["Metric", "Value"])
    ws1.append(["Scan Folder", str(scan_path)])
    ws1.append(["Recursive Scan", str(recursive)])
    ws1.append(["Report Generated", ts])
    ws1.append([])
    ws1.append(["Total Files Found", total_count])
    ws1.append(["Scanned (processed)", scanned_count])
    ws1.append(["Pending (not yet scanned)", pending_count])
    ws1.append([])
    ws1.append(["Scanned Size", scanned_size_human])
    ws1.append(["Pending Size", pending_size_human])
    ws1.append([])
    ws1.append(["Directories Scanned", len([d for d, s in dir_stats.items() if s['scanned'] > 0])])
    ws1.append(["Directories with Pending", len([d for d, s in dir_stats.items() if s['pending'] > 0])])
    ws1.append(["Total Directories", len(dir_stats)])
    _style_header(ws1)
    ws1.column_dimensions['A'].width = 32
    ws1.column_dimensions['B'].width = 60

    # Sheet 2: By Directory
    ws2 = wb.create_sheet("By Directory")
    ws2.append(["Directory", "Total Files", "Scanned", "Pending", "Pending Size", "% Done"])
    for d, stats in sorted(dir_stats.items()):
        pct = round(stats['scanned'] / stats['total'] * 100, 1) if stats['total'] else 0
        ws2.append([
            d,
            stats['total'],
            stats['scanned'],
            stats['pending'],
            human_readable_size(stats['pending_bytes']),
            f"{pct}%",
        ])
    _style_header(ws2)
    ws2.column_dimensions['A'].width = 80
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws2.column_dimensions[col].width = 16

    # Sheet 3: Pending Files
    ws3 = wb.create_sheet("Pending Files")
    ws3.append(["Directory", "Filename", "Full Path", "Size (Bytes)", "Size (Readable)"])
    for f in sorted(pending):
        p = Path(f)
        sz = _safe_size(f)
        ws3.append([str(p.parent), p.name, f, sz, human_readable_size(sz)])
    _style_header(ws3)
    ws3.column_dimensions['A'].width = 60
    ws3.column_dimensions['B'].width = 40
    ws3.column_dimensions['C'].width = 80
    ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 16

    # Sheet 4: Scanned Files
    ws4 = wb.create_sheet("Scanned Files")
    ws4.append(["Directory", "Filename", "Full Path"])
    for f in sorted(scanned):
        p = Path(f)
        ws4.append([str(p.parent), p.name, f])
    _style_header(ws4)
    ws4.column_dimensions['A'].width = 60
    ws4.column_dimensions['B'].width = 40
    ws4.column_dimensions['C'].width = 80

    # Sheet 5: All Files
    ws5 = wb.create_sheet("All Files")
    ws5.append(["Status", "Directory", "Filename", "Full Path"])
    for f in sorted(scanned):
        p = Path(f)
        ws5.append(["scanned", str(p.parent), p.name, f])
    for f in sorted(pending):
        p = Path(f)
        ws5.append(["pending", str(p.parent), p.name, f])
    _style_header(ws5)
    ws5.column_dimensions['A'].width = 12
    ws5.column_dimensions['B'].width = 60
    ws5.column_dimensions['C'].width = 40
    ws5.column_dimensions['D'].width = 80

    wb.save(xlsx_path)

    print(f"  Total: {total_count}  |  Scanned: {scanned_count}  |  Pending: {pending_count}")
    print(f"  Pending size: {pending_size_human}")
    print(f"  CSV  : {csv_path}")
    print(f"  XLSX : {xlsx_path}")
    print(f"  Sheets: Summary | By Directory | Pending Files | Scanned Files | All Files")


if __name__ == "__main__":
    main()
