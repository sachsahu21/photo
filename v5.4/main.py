"""Image Scanner v4.1"""

import os
os.environ["OPENCV_LOG_LEVEL"] = "FATAL"

import csv
import sys
import logging
import pickle
from pathlib import Path
from datetime import datetime
from time import perf_counter

from src.config_manager import ConfigManager
from src.scanner import ImageScanner
from src.duplicate_handler import DuplicateHandler
from src.organizer import ImageOrganizer
from src.excel_writer import ExcelWriter
from src.face_indexer import FaceIndexer
from src.metadata_store import MetadataStore
from src.people_sync import sync_people_tags
from src.metadata_reconcile import reconcile_vault_paths, auto_reconcile_if_enabled
from src.vault_maintenance import (
    dedupe_vault,
    quarantine_generated_targets,
    quarantine_media_cascade,
    cleanup_untagged_orphans,
    save_last_excel_path,
    load_last_excel_path,
    rebuild_vault_index,
)


from src.workspace_paths import records_backup_path


def setup_log(config):
    lf = config.get('logging.file')
    if not lf:
        raise ValueError('logging.file not resolved; set workspace.root in config.yaml')
    Path(lf).parent.mkdir(parents=True, exist_ok=True)
    handlers = [logging.FileHandler(lf, encoding='utf-8')]
    if config.get('logging.console', True):
        handlers.append(logging.StreamHandler())
    logging.basicConfig(
        level=getattr(logging, str(config.get('logging.level', 'INFO')).upper(), logging.INFO),
        format='%(asctime)s-%(name)s-%(levelname)s-%(message)s',
        handlers=handlers
    )
    return logging.getLogger(__name__)


def save_bk(records, config):
    path = records_backup_path(config)
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, 'wb') as f:
            pickle.dump(records, f, protocol=pickle.HIGHEST_PROTOCOL)
        print('  Backup: ' + str(len(records)) + ' records -> ' + str(path))
    except Exception as e:
        print('  Backup failed: ' + str(e))


def load_bk(config):
    path = records_backup_path(config)
    if not path.exists():
        print('  No backup')
        return None
    try:
        with open(path, 'rb') as f:
            r = pickle.load(f)
        print('  Loaded ' + str(len(r)) + ' records from ' + str(path))
        return r
    except Exception as e:
        print('  Load error: ' + str(e))
        return None


def _norm_path(value):
    try:
        return str(Path(str(value)).expanduser().resolve()).lower()
    except Exception:
        return str(value or '').strip().lower()


def _record_path_set(records):
    paths = set()
    for rec in records or []:
        fp = rec.get('full_path') or rec.get('file_path')
        if fp:
            paths.add(_norm_path(fp))
    return paths


def _with_scan_defaults(records):
    for r in records or []:
        if not r.get('delete_flag'):
            r['delete_flag'] = 'No'
        r.setdefault('is_duplicate', 'No')
        r.setdefault('duplicate_group', '')
        r.setdefault('is_best_in_group', '')
        r.setdefault('recommendation', '')
        r.setdefault('is_similar', 'No')
        r.setdefault('similar_group', '')
        r.setdefault('similar_score', '')
        r.setdefault('similar_methods', '')
    return records


def _recompute_duplicate_metadata(config):
    if not config.get('duplicates.enabled', True):
        return []
    md_store = MetadataStore(config.to_dict())
    records = md_store.load_records(exclude_missing=False)
    if not records:
        return []
    records = DuplicateHandler(
        hash_algorithm=config.get('duplicates.hash_algorithm', 'md5'),
        selection_criteria=config.get(
            'duplicates.selection_criteria',
            ['quality', 'resolution', 'date', 'size']
        )
    ).mark_duplicates(_with_scan_defaults(records))
    return md_store.upsert_records(records)


def _make_batch_writer(config):
    try:
        interval = int(config.get('processing.metadata_flush_interval', 100) or 100)
    except (TypeError, ValueError):
        interval = 100
    interval = max(1, interval)
    cfg = config.to_dict()
    md_store = MetadataStore(cfg)
    pending = []
    count = 0

    def flush(force=False):
        nonlocal pending
        if not pending:
            return
        if not force and len(pending) < interval:
            return
        batch = _with_scan_defaults(list(pending))
        md_store.upsert_records(batch)
        try:
            md_store.rebuild_index()
        except Exception:
            pass
        print('  Metadata batch saved: ' + str(len(batch)) + ' records')
        pending = []

    def on_record(rec):
        nonlocal count
        pending.append(rec)
        count += 1
        if count % interval == 0:
            flush(force=True)

    return on_record, lambda: flush(force=True)


def _export_missing_metadata_report(config, missing_paths):
    reports = Path(config.get('output.output_folder') or Path(config.get('workspace.root')) / 'reports')
    reports.mkdir(parents=True, exist_ok=True)
    out = reports / ('missing-metadata-' + datetime.now().strftime('%Y%m%d-%H%M%S') + '.csv')
    with open(out, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['full_path'])
        for p in sorted(missing_paths):
            writer.writerow([p])
    print('  Missing metadata report: ' + str(out))
    return out


def _missing_metadata_paths(records, config):
    expected = {}
    for rec in records or []:
        fp = rec.get('full_path') or rec.get('file_path')
        if fp:
            expected[_norm_path(fp)] = str(fp)
    if not expected:
        return []
    stored = _record_path_set(MetadataStore(config.to_dict()).load_records())
    return [original for key, original in expected.items() if key not in stored]


def _handle_missing_metadata(missing_paths, config, logger):
    if not missing_paths:
        print('  Metadata verification: all scanned files have metadata JSON')
        return
    print('  Metadata verification: ' + str(len(missing_paths)) + ' file(s) missing metadata JSON')
    while True:
        print('  1. Rescan missing files only')
        print('  2. Export missing metadata report')
        print('  3. Ignore for now')
        ch = input('  Choice (1/2/3): ').strip()
        if ch == '1':
            try:
                scanner = ImageScanner(config.to_dict())
                on_record, flush = _make_batch_writer(config)
                repaired = scanner.scan_files(missing_paths, batch_callback=on_record)
                flush()
                if repaired:
                    MetadataStore(config.to_dict()).upsert_records(_with_scan_defaults(repaired))
                    MetadataStore(config.to_dict()).rebuild_index()
                print('  Repaired metadata files: ' + str(len(repaired)))
            except Exception as e:
                logger.error('Missing metadata repair: %s', e, exc_info=True)
                print('  Repair error: ' + str(e))
            return
        if ch == '2':
            try:
                _export_missing_metadata_report(config, missing_paths)
            except Exception as e:
                print('  Report error: ' + str(e))
            return
        if ch == '3' or ch == '':
            print('  Ignored for now')
            return
        print('  Invalid choice')


def task_1(config, logger):
    print('\n  BUILD / REFRESH METADATA VAULT')
    print('  ' + '-' * 38)
    records = None
    try:
        scanner = ImageScanner(config.to_dict())
        sf = config.get('scan.folder_path')
        print('  Scanning: ' + str(sf))
        on_record, flush_batches = _make_batch_writer(config)
        records = scanner.scan(sf, batch_callback=on_record)
        flush_batches()
        if not records:
            print('  No files!')
            return None
        print('  Found ' + str(len(records)) + ' files')
        records = _with_scan_defaults(records)
        scanned_records = list(records)
        if config.get('duplicates.enabled', True):
            print('  Detecting duplicates...')
            records = DuplicateHandler(
                hash_algorithm=config.get('duplicates.hash_algorithm', 'md5'),
                selection_criteria=config.get(
                    'duplicates.selection_criteria',
                    ['quality', 'resolution', 'date', 'size']
                )
            ).mark_duplicates(records)
        else:
            for r in records:
                r['is_duplicate'] = 'No'
                r['duplicate_group'] = ''
                r['is_best_in_group'] = ''
                r['recommendation'] = ''
        if config.get('similar_detection.enabled', False):
            print('  Detecting similar...')
            from src.similar_detector import SimilarDetector
            sd = SimilarDetector(config.to_dict())
            records = sd.compute_hashes(records)
            sg = sd.find_similar(records)
            records = sd.mark_similar(records, sg)
        else:
            for r in records:
                r['is_similar'] = 'No'
                r['similar_group'] = ''
                r['similar_score'] = ''
                r['similar_methods'] = ''
        md_store = MetadataStore(config.to_dict())
        records = md_store.upsert_records(records)
        if config.get('duplicates.enabled', True):
            print('  Rechecking duplicates across full metadata vault...')
            records = _recompute_duplicate_metadata(config)
        if config.get('metadata.dedupe_after_scan', True):
            dedupe_vault(config.to_dict(), quiet=True)
            records = MetadataStore(config.to_dict()).load_records()
        print('  Metadata JSON: ' + str(len(records)) + ' records')
        save_bk(records, config)
        missing_paths = _missing_metadata_paths(scanned_records, config)
        _handle_missing_metadata(missing_paths, config, logger)
        return str(md_store.root)
    except Exception as e:
        logger.error('Task 1: %s', e, exc_info=True)
        print('  Error: ' + str(e))
        if records:
            save_bk(records, config)
        return None


def task_analyze_folders(config, logger):
    # Import required libraries locally to avoid unnecessary global dependencies
    import json
    from datetime import datetime
    try:
        from openpyxl import Workbook
    except ImportError:
        Workbook = None

    print('\n  ANALYZE FOLDER & FILE COUNTS')
    print('  ' + '-' * 38)
    default_path = config.get('scan.folder_path', '')
    p = input(f'  Folder path to analyze (Enter={default_path}): ').strip()
    if not p:
        p = default_path

    if not p or not Path(p).exists():
        print('  Error: Invalid or missing folder path.')
        return

    print('  Analyzing directory tree, this may take a moment...')
    try:
        scanner = ImageScanner(config.to_dict())
        stats = scanner.analyze_folders(p)

        # Pretty‑print to console
        print('\n  --- DIRECTORY BREAKDOWN ---')
        for folder_name, fstats in stats['top_level'].items():
            size_mb = fstats['size_bytes'] / (1024 * 1024)
            print(f"  > {folder_name}/")
            print(f"      Size:       {size_mb:.2f} MB")
            print(f"      Subfolders: {fstats['subfolders']}")
            print(f"      Images:     {fstats['images']}")
            print(f"      Videos:     {fstats['videos']}")
            print(f"      Others:     {fstats['others']}")
            print('')

        print('  --- OVERALL TOTALS ---')
        total_size_mb = stats['total_size_bytes'] / (1024 * 1024)
        print(f"  Total Size:       {total_size_mb:.2f} MB")
        print(f"  Total Folders:    {stats['total_folders']}")
        print(f"  Total Images:     {stats['total_images']}")
        print(f"  Total Videos:     {stats['total_videos']}")
        print(f"  Total Others:     {stats['total_others']}")

        # Save analysis to Excel in workspace/folder_analysis
        try:
            workspace_root = Path(config.get('workspace.root'))
            # Determine parent folder name from scanned path and include brackets
            parent_name = Path(p).name
            # Prepare output directory and filename
            analysis_dir = workspace_root / 'folder_analysis'
            analysis_dir.mkdir(parents=True, exist_ok=True)
            filename = f'folder_analysis_{parent_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            out_path = analysis_dir / filename

            if Workbook:
                wb = Workbook()
                # Summary sheet
                ws_summary = wb.active
                ws_summary.title = 'Summary'
                ws_summary.append(['Metric', 'Value'])
                ws_summary.append(['Total Size (MB)', f"{total_size_mb:.2f}"])
                ws_summary.append(['Total Folders', stats['total_folders']])
                ws_summary.append(['Total Images', stats['total_images']])
                ws_summary.append(['Total Videos', stats['total_videos']])
                ws_summary.append(['Total Others', stats['total_others']])

                # Details sheet per top-level folder
                ws_detail = wb.create_sheet(title='Details')
                ws_detail.append(['Full Path', 'Size (MB)', 'Subfolders', 'Images', 'Videos', 'Others'])
                root_path = Path(p).resolve()
                # Write top-level folder rows
                for fname, fstats in stats['top_level'].items():
                    size_mb = fstats['size_bytes'] / (1024 * 1024)
                    ws_detail.append([
                        str(root_path / fname),
                        round(size_mb, 2),
                        fstats['subfolders'],
                        fstats['images'],
                        fstats['videos'],
                        fstats['others']
                    ])
                # Write subfolder rows
                for sub_path, substats in stats['subfolders'].items():
                    size_mb = substats['size_bytes'] / (1024 * 1024)
                    ws_detail.append([
                        str(root_path / sub_path),
                        round(size_mb, 2),
                        '',  # subfolders count not tracked for deeper levels
                        substats['images'],
                        substats['videos'],
                        substats['others']
                    ])
                # Hierarchical view sheet
                ws_hier = wb.create_sheet(title='Hierarchy')
                ws_hier.append(['Path', 'Size (MB)', 'Subfolders', 'Images', 'Videos', 'Others'])
                # Add top-level entries
                for fname, fstats in stats['top_level'].items():
                    size_mb = fstats['size_bytes'] / (1024 * 1024)
                    ws_hier.append([fname, round(size_mb, 2), fstats['subfolders'], fstats['images'], fstats['videos'], fstats['others']])
                # Add subfolder entries
                for sub_path, substats in stats['subfolders'].items():
                    size_mb = substats['size_bytes'] / (1024 * 1024)
                    ws_hier.append([sub_path, round(size_mb, 2), '', substats['images'], substats['videos'], substats['others']])
                # Determine folder with maximum size
                if stats['top_level']:
                    max_folder = max(stats['top_level'].items(), key=lambda kv: kv[1]['size_bytes'])
                    max_name, max_stats = max_folder
                    ws_max = wb.create_sheet(title='MaxFolder')
                    ws_max.append(['Full Path', 'Size (MB)'])
                    root_path = Path(p)
                    max_full_path = str(root_path / max_name)
                    ws_max.append([
                        max_full_path,
                        round(max_stats['size_bytes'] / (1024 * 1024), 2)
                    ])
                wb.save(out_path)
                print(f'  Analysis saved to: {out_path}')
            else:
                # Fallback to JSON if openpyxl is unavailable
                json_path = analysis_dir / f'folder_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(stats, f, indent=2)
                print(f'  openpyxl not installed – analysis saved as JSON: {json_path}')
        except Exception as e:
            logger.error('Failed to save analysis: %s', e, exc_info=True)
            print('  Error saving analysis: ' + str(e))
    except Exception as e:
        logger.error('Analyze folders: %s', e, exc_info=True)
        print('  Error: ' + str(e))


def task_1b(config, logger):
    print('\n  GENERATE / REFRESH EXCEL')
    total_start = perf_counter()
    cfg = config.to_dict()
    if config.get('metadata.dedupe_before_excel', True):
        t = perf_counter()
        dedupe_vault(cfg, quiet=True)
        print('  Timing: pre-Excel vault dedupe ' + str(round(perf_counter() - t, 2)) + 's')
    else:
        print('  Skipping metadata dedupe before Excel; run option 23 if needed.')
    t = perf_counter()
    records = MetadataStore(cfg).load_records()
    print('  Timing: metadata load ' + str(round(perf_counter() - t, 2)) + 's')
    if not records:
        records = load_bk(config)
    if not records:
        return None
    missing = sum(1 for r in records if str(r.get('file_exists', '')).lower() != 'yes')
    if missing:
        print('  Note: ' + str(missing) + ' vault row(s) have no file on disk (see File Exists? column)')
    if config.get('workflow.reset_dup_sim_for_excel', False):
        for r in records:
            r['is_duplicate'] = 'No'
            r['is_similar'] = 'No'
    for r in records:
        if not r.get('delete_flag'):
            r['delete_flag'] = 'No'
    ad = None
    try:
        t = perf_counter()
        from src.analytics import StorageAnalytics
        ad = StorageAnalytics().analyze(records)
        print('  Timing: analytics ' + str(round(perf_counter() - t, 2)) + 's')
    except Exception:
        pass
    t = perf_counter()
    ep = ExcelWriter(cfg).write(
        records, config.get('scan.folder_path'), analytics_data=ad
    )
    print('  Timing: Excel write ' + str(round(perf_counter() - t, 2)) + 's')
    if ep:
        save_last_excel_path(config, ep)
        print('  Excel: ' + str(ep))
        if config.get('comparison.generate_after_excel', False):
            try:
                from src.comparison_generator import ComparisonGenerator
                out = config.get('comparison.output_folder')
                if out:
                    ComparisonGenerator(out).generate(records)
                    print('  Comparisons: ' + str(out))
            except Exception as e:
                print('  Comparison skipped: ' + str(e))
    print('  Total time: ' + str(round(perf_counter() - total_start, 2)) + 's')
    return ep


def task_reconcile_paths(config, logger):
    print('\n  RECONCILE VAULT PATHS')
    print('  ' + '-' * 38)
    try:
        stats = reconcile_vault_paths(config.to_dict())
        if stats.get('dedupe_removed'):
            print('  Dedupe removed: ' + str(stats['dedupe_removed']) + ' duplicate json')
    except Exception as e:
        logger.error('Reconcile: %s', e, exc_info=True)
        print('  Error: ' + str(e))


def task_dedupe_vault(config, logger):
    print('\n  DEDUPE METADATA FILES')
    print('  ' + '-' * 38)
    try:
        dedupe_vault(config.to_dict())
    except Exception as e:
        print('  Error: ' + str(e))


def _path_under(path, root):
    try:
        Path(path).resolve().relative_to(Path(root).resolve())
        return True
    except Exception:
        return False


def _generated_quarantine_target(path, workspace_root, protected_roots):
    p = Path(path).expanduser()
    if not p:
        return None
    try:
        rp = p.resolve()
        wr = Path(workspace_root).resolve()
    except Exception:
        return None
    if rp == wr or not _path_under(rp, wr):
        return None
    for root in protected_roots:
        if root and _path_under(rp, root):
            return None
    return rp


def task_fresh_restart_metadata(config, logger):
    print('\n  FRESH RESTART (CLEAR VAULT)')
    print('  ' + '-' * 38)
    try:
        workspace_root = Path(config.get('workspace.root')).expanduser().resolve()
        protected_roots = [
            config.get('scan.folder_path'),
            config.get('organization.output_folder'),
        ]
        targets = []
        metadata_root = config.get('metadata.root_folder')
        if metadata_root:
            targets.append(('metadata folder', metadata_root))
        checkpoint_file = config.get('processing.checkpoint_file')
        if checkpoint_file:
            targets.append(('checkpoints folder', str(Path(checkpoint_file).parent)))
        backup = records_backup_path(config)
        targets.append(('records backup', str(backup)))

        safe_targets = []
        for label, raw in targets:
            target = _generated_quarantine_target(raw, workspace_root, protected_roots)
            if target:
                safe_targets.append((label, target))
            else:
                print('  Skipped unsafe target: ' + label + ' -> ' + str(raw))

        if not safe_targets:
            print('  Nothing safe to quarantine.')
            return

        print('  This moves generated metadata state to quarantine:')
        for label, target in safe_targets:
            print('  - ' + label + ': ' + str(target))
        print('  Photos are not deleted from scan or organized folders.')
        if input('  Type DELETE METADATA to continue: ').strip() != 'DELETE METADATA':
            print('  Cancelled')
            return

        stats = quarantine_generated_targets(
            config.to_dict(),
            safe_targets,
            action='fresh-restart',
            manifest_prefix='fresh-restart-manifest',
        )

        Path(config.get('metadata.root_folder')).mkdir(parents=True, exist_ok=True)
        if checkpoint_file:
            Path(checkpoint_file).parent.mkdir(parents=True, exist_ok=True)
        print('  Quarantined generated target(s): ' + str(stats.get('moved', 0)))
        print('  Quarantine folder: ' + str(stats.get('run_root', '')))
        print('  Manifest: ' + str(stats.get('manifest_path', '')))
        print('  Starting fresh scan...')
        return task_1(config, logger)
    except Exception as e:
        logger.error('Fresh restart metadata: %s', e, exc_info=True)
        print('  Error: ' + str(e))


def _missing_value(value):
    if value is None:
        return True
    if value == '' or value == [] or value == {}:
        return True
    if isinstance(value, str) and value.strip().lower() in ('unknown', 'none', 'no exif'):
        return True
    return False


def _merge_missing_record(existing, fresh):
    merged = dict(existing or {})
    for key, value in (fresh or {}).items():
        if _missing_value(value):
            continue
        if key not in merged or _missing_value(merged.get(key)):
            merged[key] = value
    if fresh.get('last_seen_at'):
        merged['last_seen_at'] = fresh.get('last_seen_at')
    return merged


def task_enrich_metadata(config, logger):
    print('\n  ENRICH EXISTING METADATA')
    print('  ' + '-' * 38)
    try:
        md_store = MetadataStore(config.to_dict())
        existing = md_store.load_records(exclude_missing=False)
        if not existing:
            print('  No metadata records found. Run Build / Refresh first.')
            return
        _ENRICH_FIELDS = ('date_taken', 'camera_model', 'blur_score', 'width', 'height', 'gps_lat', 'gps_lon')

        def _is_complete(rec):
            return all(not _missing_value(rec.get(f)) for f in _ENRICH_FIELDS)

        by_path = {}
        files = []
        skipped_complete = 0
        for rec in existing:
            fp = str(rec.get('full_path') or '').strip()
            if fp and Path(fp).is_file():
                key = _norm_path(fp)
                if _is_complete(rec):
                    skipped_complete += 1
                    continue
                by_path[key] = rec
                files.append(fp)
        files = sorted(set(files))
        if not files:
            print('  No existing metadata records point to files on disk.')
            if skipped_complete:
                print('  All ' + str(skipped_complete) + ' records are already complete — nothing to enrich.')
            return
        total = len(files)
        print('  Files to enrich: ' + str(total) +
              (' (skipped ' + str(skipped_complete) + ' already complete)' if skipped_complete else ''))
        print('  Mode: fill missing fields only; global checkpoint is ignored.')
        if input('  Proceed? (yes/no): ').strip().lower() != 'yes':
            print('  Cancelled')
            return

        try:
            interval = int(config.get('processing.metadata_flush_interval', 100) or 100)
        except (TypeError, ValueError):
            interval = 100
        interval = max(1, interval)
        pending = []
        enriched = []

        def flush(force=False):
            nonlocal pending
            if not pending:
                return
            if not force and len(pending) < interval:
                return
            batch = _with_scan_defaults(list(pending))
            md_store.upsert_records(batch)
            try:
                md_store.rebuild_index()
            except Exception:
                pass
            pending = []

        def on_record(fresh):
            key = _norm_path(fresh.get('full_path') or fresh.get('file_path') or '')
            merged = _merge_missing_record(by_path.get(key, {}), fresh)
            pending.append(merged)
            enriched.append(merged)
            done = len(enriched)
            if done % interval == 0:
                flush(force=True)
                print('  Progress: ' + str(done) + ' / ' + str(total) + ' enriched')

        scanner = ImageScanner(config.to_dict())
        scanner.scan_files(files, batch_callback=on_record)
        flush(force=True)
        if enriched:
            md_store.upsert_records(_with_scan_defaults(enriched))
            md_store.rebuild_index()
            save_bk(md_store.load_records(exclude_missing=False), config)
        print('  Enriched metadata records: ' + str(len(enriched)))
    except Exception as e:
        logger.error('Enrich metadata: %s', e, exc_info=True)
        print('  Error: ' + str(e))


def task_cleanup_untagged(config, logger):
    print('\n  CLEANUP UNTAGGED SAMPLE FOLDERS')
    print('  ' + '-' * 38)
    try:
        cleanup_untagged_orphans(config.to_dict())
    except Exception as e:
        print('  Error: ' + str(e))


def task_2(ep, config, logger):
    print('\n  APPLY EXCEL DELETE ACTIONS (QUARANTINE)')
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ep)
        print('  Select sheet actions to apply:')
        print('  1. Duplicates sheet (default)')
        print('  2. Similar Images sheet')
        print('  3. All Images sheet')
        print('  4. All available sheets')
        sheet_choice = input('  Choice (Enter=1): ').strip() or '1'
        sheet_map = {
            '1': ['Duplicates'],
            '2': ['Similar Images'],
            '3': ['All Images'],
            '4': ['Duplicates', 'Similar Images', 'All Images'],
        }
        sheet_names = sheet_map.get(sheet_choice, ['Duplicates'])
        targets = []
        for sn in sheet_names:
            if sn not in wb.sheetnames:
                print('  Sheet not found, skipped: ' + sn)
                continue
            ws = wb[sn]
            hdr = {str(c.value).strip().lower(): ci for ci, c in enumerate(ws[1], 1) if c.value}
            fci = hdr.get('full path') or hdr.get('path')
            if not fci:
                continue
            dci = None
            mci = hdr.get('metadata json path')
            midci = hdr.get('media id')
            md5ci = hdr.get('md5 hash')
            groupci = hdr.get('dup group') or hdr.get('group') or hdr.get('duplicate group')
            for h, ci in hdr.items():
                if 'delete' in h:
                    dci = ci
            if not dci:
                print('  No DELETE? column, skipped: ' + sn)
                continue
            for ri in range(2, ws.max_row + 1):
                fp = ws.cell(row=ri, column=fci).value
                if not fp:
                    continue
                delete_flag = ws.cell(row=ri, column=dci).value if dci else ''
                delete_me = str(delete_flag).strip().lower() in ('yes', 'true', '1')
                if not delete_me:
                    continue
                mp = ws.cell(row=ri, column=mci).value if mci else ''
                mid = ws.cell(row=ri, column=midci).value if midci else ''
                md5 = ws.cell(row=ri, column=md5ci).value if md5ci else ''
                group = ws.cell(row=ri, column=groupci).value if groupci else ''
                targets.append({
                    'full_path': str(fp),
                    'metadata_json_path': str(mp or '').strip(),
                    'media_id': str(mid or '').strip(),
                    'md5_hash': str(md5 or '').strip(),
                    'duplicate_group': str(group or '').strip(),
                    'sheet': sn,
                })

        if not targets:
            print('  Nothing marked for quarantine')
            return
        uniq = {}
        for t in targets:
            uniq[t['full_path']] = t
        print('  ' + str(len(uniq)) + ' files marked')
        print('  Files will be moved to quarantine, not permanently deleted.')
        if input('  Confirm quarantine? (yes/no): ').strip().lower() != 'yes':
            return
        stats = quarantine_media_cascade(config.to_dict(), list(uniq.values()))
        print(
            '  Images moved: ' + str(stats['files_moved'])
            + ' | missing: ' + str(stats['files_missing'])
        )
        print(
            '  Metadata moved: ' + str(stats['json_moved'])
            + ' | missing metadata: ' + str(stats['json_missing'])
            + ' | face index rows: ' + str(stats['face_rows_removed'])
        )
        print('  Quarantine folder: ' + str(stats.get('quarantine_root', '')))
        print('  Manifest: ' + str(stats.get('manifest_path', '')))
        print('  Rechecking duplicates across remaining metadata...')
        _recompute_duplicate_metadata(config)
        dedupe_vault(config.to_dict(), quiet=True)
        auto_reconcile_if_enabled(config.to_dict(), 'after quarantine')
        records = MetadataStore(config.to_dict()).load_records()
        if records:
            save_bk(records, config)
    except Exception as e:
        print('  Error: ' + str(e))


def task_3(ep, config, logger):
    print('\n  ORGANIZE LIBRARY FROM EXCEL')
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ep)
        sn = next((n for n in ['All Images', 'Duplicates'] if n in wb.sheetnames), None)
        if not sn:
            print('  No sheet')
            return
        ws = wb[sn]
        cm = {str(c.value).strip(): ci for ci, c in enumerate(ws[1], 1) if c.value}
        h2k = {
            'Media ID': 'media_id',
            'Filename': 'filename', 'File': 'filename', 'Folder': 'folder',
            'Full Path': 'full_path', 'Path': 'full_path',
            'Date Taken': 'date_taken', 'Date': 'date_taken',
            'File Modified': 'file_modified',
            'Manual Date Override': 'manual_date_override',
            'Effective Organize Date': 'effective_organize_date',
            'Effective Date Source': 'effective_date_source',
            'Metadata JSON Path': 'metadata_json_path',
            'DELETE?': 'delete_flag', 'Dup?': 'is_duplicate',
            'Type': 'file_type', 'Format': 'extension', 'Fmt': 'extension',
            'Size MB': 'size_mb', 'Size': 'size_mb',
            'Width': 'width', 'W': 'width',
            'Height': 'height', 'H': 'height',
            'EXIF': 'has_exif', 'Metadata': 'metadata_status',
        }
        records = []
        for ri in range(2, ws.max_row + 1):
            rec = {}
            for h, ci in cm.items():
                v = ws.cell(row=ri, column=ci).value
                rec[h2k.get(h, h.lower().replace(' ', '_'))] = v
                rec[h] = v
            records.append(rec)
        print('  Loaded ' + str(len(records)) + ' records')

        # --- Selective directory filter ---
        dir_map = {}
        for rec in records:
            fp = str(rec.get('full_path') or rec.get('Full Path') or '').strip()
            if fp:
                d = str(Path(fp).parent)
                dir_map[d] = dir_map.get(d, 0) + 1
        dirs = sorted(dir_map.keys())
        if dirs:
            print('\n  Source directories found in Excel:')
            for i, d in enumerate(dirs, 1):
                print(f'  {i:3}.  {d}  ({dir_map[d]} files)')
            print('\n  Enter numbers to organize, comma-separated (e.g. 1,3)')
            sel = input('  Selection (Enter = ALL): ').strip()
            if sel:
                try:
                    chosen_indices = {int(x.strip()) for x in sel.split(',') if x.strip()}
                    chosen_dirs = {dirs[i - 1] for i in chosen_indices if 1 <= i <= len(dirs)}
                    before = len(records)
                    records = [
                        r for r in records
                        if str(Path(str(r.get('full_path') or r.get('Full Path') or '')).parent) in chosen_dirs
                    ]
                    print(f'  Filtered to {len(records)} of {before} files across {len(chosen_dirs)} director(ies)')
                except (ValueError, IndexError):
                    print('  Invalid selection — organizing ALL')
        # --- end filter ---

        mvs = ImageOrganizer(config.to_dict()).organize(records)
        s = sum(1 for m in mvs if m['status'] == 'Success')
        e = sum(1 for m in mvs if 'Error' in m['status'])
        print('  Done: ' + str(s) + ' success, ' + str(e) + ' errors')
        auto_reconcile_if_enabled(config.to_dict(), 'after organize')
        dedupe_vault(config.to_dict(), quiet=True)
        records = MetadataStore(config.to_dict()).load_records()
        if records:
            save_bk(records, config)
    except Exception as e:
        print('  Error: ' + str(e))


def task_5(config, logger):
    print('\n  BUILD / UPDATE FACE INDEX')
    print('  ' + '-' * 38)
    try:
        fd = str(config.get('faces.data_folder', '') or '').strip()
        if fd:
            Path(fd).mkdir(parents=True, exist_ok=True)
        fi = FaceIndexer(config.to_dict())
        files, faces = fi.build_or_update_index(recursive=True)
        print(f'  Updated: {files} files | {faces} faces')
        print(f'  DB: {fi.index_db}')
    except Exception as e:
        print('  Error: ' + str(e))


def task_6(config, logger):
    print('\n  SYNC PEOPLE TAGS')
    print('  ' + '-' * 38)
    try:
        md_store = MetadataStore(config.to_dict())
        records = md_store.load_records()
        if not records:
            records = load_bk(config)
        if not records:
            print('  No records. Run Step 1 first.')
            return None

        fi = FaceIndexer(config.to_dict())
        matches = fi.find_person()
        untagged_root = Path(config.get('faces.untagged_root'))
        try:
            ums = int(config.get('faces.untagged_max_samples', 1) or 1)
        except (TypeError, ValueError):
            ums = 1
        ums = max(1, min(20, ums))
        pick_q = config.get('faces.untagged_pick_best_quality', True)
        if isinstance(pick_q, str):
            pick_q = pick_q.strip().lower() in ('1', 'true', 'yes', 'on')
        export_ut = config.get('faces.export_untagged', True)
        if isinstance(export_ut, str):
            export_ut = export_ut.strip().lower() in ('1', 'true', 'yes', 'on')
        known, unknown = sync_people_tags(
            records,
            matches,
            untagged_root,
            export_untagged=bool(export_ut),
            seed_only_refresh=False,
            untagged_max_samples=ums,
            untagged_pick_best_quality=bool(pick_q),
            untagged_export_mode=str(config.get('faces.untagged_export_mode', 'full') or 'full'),
            config=config.to_dict(),
        )
        save_bk(records, config)
        print(f'  Known tagged: {known} | Unknown tagged: {unknown}')
        print(f'  Untagged samples folder: {untagged_root}')
        return True
    except Exception as e:
        print('  Error: ' + str(e))
        return None


def task_7(config, logger):
    print('\n  REFRESH SEED FEEDBACK')
    print('  ' + '-' * 38)
    try:
        md_store = MetadataStore(config.to_dict())
        records = md_store.load_records()
        if not records:
            records = load_bk(config)
        if not records:
            print('  No records. Run Step 1 first.')
            return None
        fi = FaceIndexer(config.to_dict())
        matches = fi.find_person()
        untagged_root = Path(config.get('faces.untagged_root'))
        known, unknown = sync_people_tags(
            records, matches, untagged_root, export_untagged=False, seed_only_refresh=True,
            config=config.to_dict(),
        )
        save_bk(records, config)
        print(f'  Known re-tagged: {known}')
        return True
    except Exception as e:
        print('  Error: ' + str(e))
        return None


def task_convert_structure(config, logger):
    print('\n  CONVERT FOLDER STRUCTURE')
    print('  ' + '-' * 38)
    source = config.get('organization.output_folder', './organized_images')
    print('  Source folder: ' + str(source))
    if not Path(source).exists():
        print('  Error: Folder not found!')
        return
    print('')
    print('  1. flat   2. year   3. year-month-date')
    ch = input('  Select target (1/2/3): ').strip()
    targets = {'1': 'flat', '2': 'year', '3': 'year-month-date'}
    target = targets.get(ch)
    if not target:
        print('  Invalid choice')
        return
    out = input('  Output folder (Enter=in-place): ').strip()
    target_folder = out if out else None
    if input('  Confirm? (yes/no): ').strip().lower() != 'yes':
        print('  Cancelled')
        return
    ImageOrganizer.convert_structure(source, target, target_folder)
    auto_reconcile_if_enabled(config.to_dict(), 'after convert structure')


def task_merge_dates(config, logger):
    print('\n  MERGE DUPLICATE DATES')
    print('  ' + '-' * 38)
    source = config.get('organization.output_folder', './organized_images')
    print('  Source folder: ' + str(source))
    if not Path(source).exists():
        print('  Error: Folder not found!')
        return
    detected = ImageOrganizer.detect_structure(source)
    print('  Detected structure: ' + detected)
    override = input('  Structure override [flat/year/year-month-date] (Enter=detected): ').strip().lower()
    structure = override if override in ('flat', 'year', 'year-month-date') else detected
    print('  Using structure: ' + structure)
    if input('  Confirm? (yes/no): ').strip().lower() != 'yes':
        print('  Cancelled')
        return
    ImageOrganizer.merge_duplicate_dates(source, structure)
    auto_reconcile_if_enabled(config.to_dict(), 'after merge same-date')


def task_8(config, logger):
    return task_1b(config, logger)


def task_update_counts(config, logger):
    print('\n  UPDATE PICTURE COUNTS')
    print('  ' + '-' * 38)
    source = config.get('organization.output_folder', './organized_images')
    print('  Source folder: ' + str(source))
    if not Path(source).exists():
        print('  Error: Folder not found!')
        return
    if input('  Confirm? (yes/no): ').strip().lower() != 'yes':
        print('  Cancelled')
        return
    ImageOrganizer.update_all_pic_counts(source)


def _vault_line(config) -> str:
    lines = []
    art = config.artifact_summary() if hasattr(config, 'artifact_summary') else {}
    ws = art.get('workspace') or str(config.get('workspace.root', '') or '').strip()
    if ws:
        lines.append('  Workspace: ' + ws)
    rf = art.get('metadata') or str(config.get('metadata.root_folder', '') or '').strip()
    if rf:
        lines.append('  Metadata vault: ' + rf)
    lr = str(config.get('metadata.library_root', '') or '').strip()
    if lr:
        lines.append('  Library root: ' + lr)
    return '\n'.join(lines)


def main():
    try:
        print('')
        print('  ========================================================')
        print('       IMAGE SCANNER v5.0')
        print('  ========================================================')
        config = ConfigManager()
        if not config.validate():
            print('\n  Fix config.yaml and restart.')
            sys.exit(1)
        logger = setup_log(config)
        print('  Config: ' + str(config.config_path))
        art = config.artifact_summary()
        print('  Workspace: ' + str(art.get('workspace', '')))
        print('  Artifacts: metadata, reports, face_data, ... under workspace')
        print('  Scan: ' + str(config.get('scan.folder_path')))
        print('  Structure: ' + str(config.get('organization.folder_structure', 'year')))
        ft = []
        for k, n in [
            ('duplicates.enabled', 'duplicates'),
            ('similar_detection.enabled', 'similar'),
            ('blur_detection.enabled', 'blur'),
            ('face_detection.enabled', 'faces'),
            ('processing.fast_mode', 'FAST'),
        ]:
            if config.get(k):
                ft.append(n)
        if ft:
            print('  Features: ' + ', '.join(ft))
        last_excel = load_last_excel_path(config)
        while True:
            print('\n' + '=' * 56)
            print(_vault_line(config))
            print('')
            print('  1. ANALYSIS')
            print('  ' + '-' * 30)
            print('  11.  Analyze folder & file counts')
            print('  12.  Export scan progress report')

            print('')
            print('  2. METADATA')
            print('  ' + '-' * 30)
            print('  21.  Build / Refresh Metadata Vault')
            print('  22.  Enrich existing metadata')
            print('  23.  Reconcile vault paths')
            print('  24.  Dedupe metadata files')
            print('  25.  Fresh restart (clear vault)')

            print('')
            print('  3. EXCEL & DATA')
            print('  ' + '-' * 30)
            print('  31.  Generate / Refresh Excel')
            print('  32.  Apply delete actions (quarantine)')

            print('')
            print('  4. LIBRARY')
            print('  ' + '-' * 30)
            print('  41.  Organize from Excel')
            print('  42.  Convert folder structure')
            print('  43.  Merge duplicate dates')
            print('  44.  Update picture counts')

            print('')
            print('  5. FACES & PEOPLE')
            print('  ' + '-' * 30)
            print('  51.  Build / Update face index')
            print('  52.  Sync people tags')
            print('  53.  Refresh seed feedback')
            print('  54.  Cleanup untagged samples')

            print('')
            print('   0.  Exit')
            print('=' * 56)
            ch = input('  Choice: ').strip().lower()

            if ch == '11':
                task_analyze_folders(config, logger)
            elif ch == '12':
                from src import generate_report
                generate_report.main()
            elif ch == '21':
                task_1(config, logger)
            elif ch == '22':
                task_enrich_metadata(config, logger)
            elif ch == '23':
                task_reconcile_paths(config, logger)
            elif ch == '24':
                task_dedupe_vault(config, logger)
            elif ch == '25':
                task_fresh_restart_metadata(config, logger)
            elif ch == '31':
                last_excel = task_1b(config, logger) or last_excel
            elif ch == '32':
                p = input('  Excel (Enter=last): ').strip() or last_excel or load_last_excel_path(config) or ''
                if p and Path(p).exists():
                    task_2(p, config, logger)
                else:
                    print('  Not found')

            elif ch == '41':
                p = input('  Excel (Enter=last): ').strip() or last_excel or load_last_excel_path(config) or ''
                if p and Path(p).exists():
                    task_3(p, config, logger)
                else:
                    print('  Not found')
            elif ch == '42':
                task_convert_structure(config, logger)
            elif ch == '43':
                task_merge_dates(config, logger)
            elif ch == '44':
                task_update_counts(config, logger)
            elif ch == '51':
                task_5(config, logger)
            elif ch == '52':
                task_6(config, logger)
            elif ch == '53':
                task_7(config, logger)
            elif ch == '54':
                task_cleanup_untagged(config, logger)
            elif ch == '0':
                print('\n  Bye!')
                break
            else:
                print('  Invalid choice')
            
    except KeyboardInterrupt:
        print('\n  Interrupted')
        sys.exit(1)
    except Exception as e:
        print('\n  Fatal: ' + str(e))
        sys.exit(1)


if __name__ == '__main__':
    main()

