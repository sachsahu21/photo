
# usage instruction:

# # Install dependencies
# pip install pillow openpyxl opencv-python

# # Run the script
# python image_scanner.py

# # After Task 1:
# # 1. Open the generated Excel file
# # 2. Go to "Blurry Images" sheet
# # 3. Go to "Duplicates" sheet
# # 4. Mark files with "Yes" in "DELETE? (Yes/No)" column
# # 5. Save the Excel file

# # Uncomment Task 2 in the script and run again
# # delete_marked_files(excel_path)

# # Uncomment Task 3 and run again
# # organize_by_date(records, ORGANIZED_FOLDER)
# 
# 
# 

"""
Enhanced Image Scanner with Blur Detection & Organization
----------------------------------------------------------
1. Set SCAN_FOLDER below to your folder path
2. Run: python image_scanner.py
3. Follow the workflow for duplicate/blur handling and organization
"""

import os
import hashlib
import cv2
import numpy as np
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter
import shutil

# ──────────────────────────────────────────────
#  ✏️  EDIT THIS — your folder path
# ──────────────────────────────────────────────
SCAN_FOLDER = r"C:\Users\ISSUser\Desktop\Sachin\hdd\pic"
OUTPUT_FILE = "meta"
ORGANIZED_FOLDER = r"C:\Users\issuser\Desktop\Sachin\organized_photos"
# ──────────────────────────────────────────────

IMAGE_EXTENSIONS = {
    '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif',
    '.webp', '.heic', '.heif', '.raw', '.cr2', '.nef', '.arw',
    '.dng', '.orf', '.rw2', '.pef', '.svg', '.ico', '.psd',
    '.avif', '.jfif'
}

# ── Install check ──────────────────────────────
try:
    from PIL import Image
    from PIL.ExifTags import TAGS, GPSTAGS
    PIL_OK = True
except ImportError:
    PIL_OK = False
    print("WARNING: Pillow not found. Run: pip install pillow")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("WARNING: openpyxl not found. Run: pip install openpyxl")

try:
    import cv2
    CV2_OK = True
except ImportError:
    CV2_OK = False
    print("WARNING: opencv-python not found. Run: pip install opencv-python")


# ── BLUR DETECTION ─────────────────────────────
def detect_blur(filepath, threshold=100):
    """
    Detect if image is blurry using Laplacian variance method.
    Returns: (is_blurry: bool, blur_score: float, quality_rating: str)
    """
    try:
        if not CV2_OK:
            return None, None, "Unknown"

        img = cv2.imread(str(filepath))
        if img is None:
            return None, None, "Error"

        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        laplacian_var = cv2.Laplacian(gray, cv2.CV_64F).var()

        if laplacian_var < threshold * 0.5:
            return True, round(laplacian_var, 2), "Very Blurry"
        elif laplacian_var < threshold:
            return True, round(laplacian_var, 2), "Blurry"
        elif laplacian_var < threshold * 2:
            return False, round(laplacian_var, 2), "Fair"
        else:
            return False, round(laplacian_var, 2), "Sharp"

    except Exception as e:
        return None, None, f"Error: {str(e)[:30]}"


# ── QUALITY ASSESSMENT ─────────────────────────
def assess_image_quality(filepath, width, height, blur_score):
    """
    Comprehensive quality assessment.
    Returns: quality_score (0-100), issues_list
    """
    issues = []
    score = 100

    # Check resolution
    if width and height:
        megapixels = (width * height) / 1_000_000
        if megapixels < 1:
            issues.append("Low resolution")
            score -= 20
        elif megapixels < 2:
            issues.append("Below 2MP")
            score -= 10

    # Check blur
    if blur_score is not None:
        if blur_score < 50:
            issues.append("Very blurry")
            score -= 30
        elif blur_score < 100:
            issues.append("Slightly blurry")
            score -= 15

    # Check file size (corrupted files often have unusual sizes)
    try:
        size_mb = os.path.getsize(filepath) / (1024 * 1024)
        if size_mb < 0.05:
            issues.append("Suspiciously small")
            score -= 15
    except:
        pass

    return max(0, score), issues


# ── File hash for duplicate detection ─────────
def file_hash(filepath):
    h = hashlib.md5()
    try:
        with open(filepath, 'rb') as f:
            while chunk := f.read(65536):
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return None


# ── GPS extraction ─────────────────────────────
def get_gps(exif_data):
    try:
        gps = exif_data.get('GPSInfo', {})
        if not gps:
            return None, None
        def to_dec(v):
            return float(v[0]) + float(v[1]) / 60 + float(v[2]) / 3600
        lat = to_dec(gps.get(2, (0, 0, 0)))
        lon = to_dec(gps.get(4, (0, 0, 0)))
        if gps.get(1) == 'S': lat = -lat
        if gps.get(3) == 'W': lon = -lon
        return round(lat, 6), round(lon, 6)
    except Exception:
        return None, None


# ── Metadata extraction ────────────────────────
def get_metadata(filepath):
    meta = {
        'width': None, 'height': None, 'mode': None, 'dpi': None,
        'date_taken': None, 'camera_make': None, 'camera_model': None,
        'focal_length': None, 'aperture': None, 'iso': None,
        'exposure_time': None, 'gps_lat': None, 'gps_lon': None,
        'has_exif': False, 'error': None
    }
    if not PIL_OK:
        return meta
    try:
        with Image.open(filepath) as img:
            meta['width']  = img.width
            meta['height'] = img.height
            meta['mode']   = img.mode
            dpi = img.info.get('dpi')
            if dpi:
                meta['dpi'] = f"{int(dpi[0])}x{int(dpi[1])}"

            exif_raw = None
            if hasattr(img, '_getexif'):
                exif_raw = img._getexif()
            elif hasattr(img, 'getexif'):
                exif_raw = img.getexif()

            if exif_raw:
                meta['has_exif'] = True
                exif = {}
                for tag_id, val in exif_raw.items():
                    tag = TAGS.get(tag_id, tag_id)
                    if tag == 'GPSInfo' and isinstance(val, dict):
                        exif['GPSInfo'] = val
                    else:
                        exif[tag] = val

                meta['camera_make']  = (str(exif.get('Make',  '') or '')).strip() or None
                meta['camera_model'] = (str(exif.get('Model', '') or '')).strip() or None

                for dt_tag in ('DateTimeOriginal', 'DateTime', 'DateTimeDigitized'):
                    dt = exif.get(dt_tag)
                    if dt:
                        try:
                            meta['date_taken'] = datetime.strptime(str(dt), '%Y:%m:%d %H:%M:%S')
                        except Exception:
                            meta['date_taken'] = str(dt)
                        break

                fl = exif.get('FocalLength')
                if fl:
                    try: meta['focal_length'] = f"{float(fl):.1f}mm"
                    except: meta['focal_length'] = str(fl)

                fn = exif.get('FNumber')
                if fn:
                    try: meta['aperture'] = f"f/{float(fn):.1f}"
                    except: meta['aperture'] = str(fn)

                iso = exif.get('ISOSpeedRatings') or exif.get('PhotographicSensitivity')
                if iso:
                    meta['iso'] = str(iso)

                exp = exif.get('ExposureTime')
                if exp:
                    try:
                        fv = float(exp)
                        meta['exposure_time'] = f"1/{round(1/fv)}s" if fv < 1 else f"{fv}s"
                    except: meta['exposure_time'] = str(exp)

                meta['gps_lat'], meta['gps_lon'] = get_gps(exif)

    except Exception as e:
        meta['error'] = str(e)[:120]

    return meta


# ── Scan all images ────────────────────────────
def scan(folder):
    folder = Path(folder).expanduser().resolve()
    if not folder.exists():
        print(f"ERROR: Folder not found: {folder}")
        return []

    print(f"\nScanning: {folder}")

    all_files = [
        Path(dirpath) / fn
        for dirpath, _, filenames in os.walk(folder)
        for fn in filenames
        if Path(fn).suffix.lower() in IMAGE_EXTENSIONS
    ]

    print(f"Found {len(all_files)} images. Extracting metadata...\n")

    records  = []
    hash_map = defaultdict(list)

    for i, fp in enumerate(all_files, 1):
        print(f"  [{i}/{len(all_files)}] {fp.name}", end='\r')
        stat = fp.stat()
        md5  = file_hash(fp)
        meta = get_metadata(fp)

        # NEW: Blur detection
        is_blurry, blur_score, quality_rating = detect_blur(str(fp))

        rec = {
            'filename':      fp.name,
            'folder':        str(fp.parent),
            'full_path':     str(fp),
            'extension':     fp.suffix.lower().lstrip('.').upper(),
            'size_mb':       round(stat.st_size / (1024 * 1024), 2),
            'file_modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
            'md5_hash':      md5,
            'is_blurry':     is_blurry,
            'blur_score':    blur_score,
            'quality_rating': quality_rating,
            'delete_flag':   'No',  # NEW: User can set to 'Yes'
            **meta
        }

        # Quality assessment
        quality_score, issues = assess_image_quality(str(fp), meta['width'], meta['height'], blur_score)
        rec['quality_score'] = quality_score
        rec['quality_issues'] = '; '.join(issues) if issues else 'None'

        records.append(rec)
        if md5:
            hash_map[md5].append(str(fp))

    # Mark duplicates
    dup_groups = {}
    gid = 1
    for md5, paths in hash_map.items():
        if len(paths) > 1:
            for p in paths:
                dup_groups[p] = gid
            gid += 1

    for r in records:
        r['is_duplicate']    = 'YES' if r['full_path'] in dup_groups else 'No'
        r['duplicate_group'] = dup_groups.get(r['full_path'], '')

    print()
    return records


# ── Excel writer with new columns ──────────────
COLUMNS = [
    ('filename',        'Filename',        28),
    ('folder',          'Folder',          45),
    ('extension',       'Format',          10),
    ('size_mb',         'Size (MB)',        11),
    ('width',           'Width (px)',       11),
    ('height',          'Height (px)',      11),
    ('mode',            'Color Mode',      12),
    ('dpi',             'DPI',             10),
    ('date_taken',      'Date Taken',      20),
    ('camera_make',     'Camera Make',     16),
    ('camera_model',    'Camera Model',    18),
    ('focal_length',    'Focal Length',    13),
    ('aperture',        'Aperture',        11),
    ('iso',             'ISO',              8),
    ('exposure_time',   'Exposure',        11),
    ('gps_lat',         'GPS Lat',         11),
    ('gps_lon',         'GPS Lon',         11),
    ('has_exif',        'Has EXIF',        10),
    ('blur_score',      'Blur Score',      11),
    ('quality_rating',  'Quality',         12),
    ('quality_score',   'Quality %',       10),
    ('quality_issues',  'Issues',          30),
    ('is_blurry',       'Blurry?',         10),
    ('is_duplicate',    'Duplicate?',      12),
    ('duplicate_group', 'Dup. Group',      10),
    ('delete_flag',     'DELETE? (Yes/No)', 15),  # NEW: User editable
    ('md5_hash',        'MD5 Hash',        34),
    ('file_modified',   'File Modified',   20),
    ('full_path',       'Full Path',       55),
    ('error',           'Read Error',      25),
]

def _border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(color='2E4057'):
    return {
        'font':      Font(bold=True, color='FFFFFF', size=11),
        'fill':      PatternFill('solid', start_color=color),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }

def write_excel(records, out_path):

    parent = Path(SCAN_FOLDER).name
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path(out_path) / f"image_scan_{parent}_{ts}.xlsx"

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Images ──────────────────────
    ws = wb.active
    ws.title = 'All Images'
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 32

    for ci, (_, label, width) in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=label)
        h = _hdr()
        c.font, c.fill, c.alignment = h['font'], h['fill'], h['alignment']
        c.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    alt_fill = PatternFill('solid', start_color='F5F7FA')
    dup_fill = PatternFill('solid', start_color='FFD6D6')
    blur_fill = PatternFill('solid', start_color='FFE8B6')

    for ri, rec in enumerate(records, 2):
        if rec['is_blurry'] == True:
            fill = blur_fill
        elif rec['is_duplicate'] == 'YES':
            fill = dup_fill
        else:
            fill = alt_fill if ri % 2 == 0 else None

        for ci, (key, _, _) in enumerate(COLUMNS, 1):
            val = rec.get(key, '')
            if isinstance(val, datetime): val = val.strftime('%Y-%m-%d %H:%M:%S')
            if isinstance(val, bool):     val = 'Yes' if val else 'No'
            if isinstance(val, str):
                val = ''.join(ch for ch in val if ord(ch) >= 32)

            try:
                c = ws.cell(row=ri, column=ci, value=val)
            except Exception as e:
                print(f"\nERROR at row {ri}, column {key}: {repr(val)}")
                val = str(val)
                val = ''.join(ch for ch in val if ord(ch) >= 32)
                c = ws.cell(row=ri, column=ci, value=val)

            c.border = _border()
            c.alignment = Alignment(vertical='center')
            if fill: c.fill = fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Sheet 2: Blurry Images ──────────────────
    ws_blur = wb.create_sheet('Blurry Images')
    ws_blur.freeze_panes = 'A2'
    ws_blur.row_dimensions[1].height = 32

    blur_cols = [
        ('filename',        'Filename',        28),
        ('folder',          'Folder',          45),
        ('blur_score',      'Blur Score',      11),
        ('quality_rating',  'Quality',         12),
        ('quality_score',   'Quality %',       10),
        ('quality_issues',  'Issues',          30),
        ('width',           'Width (px)',      11),
        ('height',          'Height (px)',     11),
        ('size_mb',         'Size (MB)',       11),
        ('date_taken',      'Date Taken',      20),
        ('delete_flag',     'DELETE? (Yes/No)',15),
        ('full_path',       'Full Path',       55),
    ]

    for ci, (_, label, width) in enumerate(blur_cols, 1):
        c = ws_blur.cell(row=1, column=ci, value=label)
        h = _hdr('FF8C00')
        c.font, c.fill, c.alignment = h['font'], h['fill'], h['alignment']
        c.border = _border()
        ws_blur.column_dimensions[get_column_letter(ci)].width = width

    blurry_recs = sorted([r for r in records if r['is_blurry'] == True],
                         key=lambda x: x.get('blur_score', 0))

    for ri, rec in enumerate(blurry_recs, 2):
        fill = PatternFill('solid', start_color='FFE8B6')
        for ci, (key, _, _) in enumerate(blur_cols, 1):
            val = rec.get(key, '')
            if isinstance(val, datetime): val = val.strftime('%Y-%m-%d %H:%M:%S')
            c = ws_blur.cell(row=ri, column=ci, value=val)
            c.border = _border()
            c.alignment = Alignment(vertical='center')
            c.fill = fill

    ws_blur.auto_filter.ref = f"A1:{get_column_letter(len(blur_cols))}1"

    # ── Sheet 3: Duplicates ──────────────────────
    ws2 = wb.create_sheet('Duplicates')
    ws2.freeze_panes = 'A2'
    ws2.row_dimensions[1].height = 32

    dup_cols = [
        ('duplicate_group', 'Group',    8),
        ('filename',        'Filename', 28),
        ('folder',          'Folder',   45),
        ('extension',       'Format',   10),
        ('size_mb',         'Size (MB)',11),
        ('quality_score',   'Quality %',10),
        ('delete_flag',     'DELETE? (Yes/No)',15),
        ('md5_hash',        'MD5 Hash', 34),
        ('full_path',       'Full Path',55),
    ]
    for ci, (_, label, width) in enumerate(dup_cols, 1):
        c = ws2.cell(row=1, column=ci, value=label)
        h = _hdr('8B0000')
        c.font, c.fill, c.alignment = h['font'], h['fill'], h['alignment']
        c.border = _border()
        ws2.column_dimensions[get_column_letter(ci)].width = width

    dups = sorted([r for r in records if r['is_duplicate'] == 'YES'],
                  key=lambda x: (x.get('duplicate_group', 0), x['full_path']))

    prev, alt_flag = None, False
    f1 = PatternFill('solid', start_color='FFE8E8')
    f2 = PatternFill('solid', start_color='FFF5F5')
    for ri, rec in enumerate(dups, 2):
        grp = rec.get('duplicate_group')
        if grp != prev:
            alt_flag = not alt_flag
            prev = grp
        fill = f1 if alt_flag else f2
        for ci, (key, _, _) in enumerate(dup_cols, 1):
            val = rec.get(key, '')
            if isinstance(val, datetime): val = val.strftime('%Y-%m-%d %H:%M:%S')
            c = ws2.cell(row=ri, column=ci, value=val)
            c.border = _border()
            c.alignment = Alignment(vertical='center')
            c.fill = fill
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(dup_cols))}1"

    # ── Sheet 4: Summary ─────────────────────────
    ws3 = wb.create_sheet('Summary')
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 20

    ws3['A1'] = 'Image Scan Summary'
    ws3['A1'].font = Font(bold=True, size=14, color='2E4057')
    ws3['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws3['A2'].font = Font(italic=True, color='888888')
    ws3['A3'] = f"Scanned: {SCAN_FOLDER}"
    ws3['A3'].font = Font(italic=True, color='888888')

    total     = len(records)
    dup_count = sum(1 for r in records if r['is_duplicate'] == 'YES')
    dup_grps  = len(set(r['duplicate_group'] for r in records if r['duplicate_group']))
    blur_count = sum(1 for r in records if r['is_blurry'] == True)
    with_exif = sum(1 for r in records if r.get('has_exif'))
    with_gps  = sum(1 for r in records if r.get('gps_lat'))
    ext_cnts  = Counter(r['extension'] for r in records)
    avg_quality = round(sum(r.get('quality_score', 0) for r in records) / total, 1) if total > 0 else 0

    rows = [
        ('', ''),
        ('GENERAL', ''),
        ('Total Images Found',    total),
        ('Total Folders Scanned', len(set(r['folder'] for r in records))),
        ('Total Size (MB)',        round(sum(r.get('size_mb', 0) for r in records) , 1)),
        ('Average Quality Score',  f"{avg_quality}%"),
        ('', ''),
        ('QUALITY ISSUES', ''),
        ('Blurry Images',   blur_count),
        ('Duplicate Files', dup_count),
        ('Duplicate Groups', dup_grps),
        ('', ''),
        ('METADATA', ''),
        ('Files with EXIF', with_exif),
        ('Files with GPS',  with_gps),
        ('', ''),
        ('BY FORMAT', ''),
        *[(f'  .{ext.lower()}', cnt) for ext, cnt in sorted(ext_cnts.items(), key=lambda x: -x[1])]
    ]

    hdr_fill = PatternFill('solid', start_color='2E4057')
    acc_fill = PatternFill('solid', start_color='EBF2FF')
    for ri, (label, value) in enumerate(rows, 5):
        a = ws3.cell(row=ri, column=1, value=label)
        b = ws3.cell(row=ri, column=2, value=value)
        if label in ('GENERAL', 'QUALITY ISSUES', 'METADATA', 'BY FORMAT'):
            a.font = Font(bold=True, color='FFFFFF')
            a.fill = hdr_fill
            b.fill = hdr_fill
        elif label:
            a.font = Font(bold=True, size=11)
            if ri % 2 == 0:
                a.fill = acc_fill
                b.fill = acc_fill

    wb.save(out_path)
    print(f"✓ Saved -> {out_path}")
    return str(out_path)


# ── TASK 2: Delete marked files ────────────────
def delete_marked_files(excel_path):
    """
    Read Excel file and delete files marked with 'Yes' in delete_flag column
    """
    if not OPENPYXL_OK:
        print("ERROR: openpyxl required")
        return

    print("\n" + "="*60)
    print("TASK 2: DELETE MARKED FILES")
    print("="*60)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb['All Images']

    # Find delete_flag column
    delete_col = None
    for ci, cell in enumerate(ws[1], 1):
        if cell.value == 'DELETE? (Yes/No)':
            delete_col = ci
            break

    if not delete_col:
        print("ERROR: 'DELETE? (Yes/No)' column not found")
        return

    # Find filename and full_path columns
    filename_col = None
    fullpath_col = None
    for ci, cell in enumerate(ws[1], 1):
        if cell.value == 'Filename':
            filename_col = ci
        elif cell.value == 'Full Path':
            fullpath_col = ci

    deleted_files = []
    for ri in range(2, ws.max_row + 1):
        delete_val = ws.cell(row=ri, column=delete_col).value
        if delete_val and str(delete_val).strip().upper() == 'YES':
            filepath = ws.cell(row=ri, column=fullpath_col).value
            filename = ws.cell(row=ri, column=filename_col).value

            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
                    deleted_files.append((filename, filepath, "Deleted"))
                    print(f"✓ Deleted: {filename}")
                else:
                    deleted_files.append((filename, filepath, "File not found"))
                    print(f"⚠ Not found: {filename}")
            except Exception as e:
                deleted_files.append((filename, filepath, f"Error: {str(e)[:50]}"))
                print(f"✗ Error deleting {filename}: {e}")

    # Create deletion report
    if deleted_files:
        report_path = Path(excel_path).parent / f"deletion_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb_report = openpyxl.Workbook()
        ws_report = wb_report.active
        ws_report.title = 'Deletion Report'

        headers = ['Filename', 'Full Path', 'Status']
        for ci, header in enumerate(headers, 1):
            c = ws_report.cell(row=1, column=ci, value=header)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', start_color='2E4057')

        for ri, (fname, fpath, status) in enumerate(deleted_files, 2):
            ws_report.cell(row=ri, column=1, value=fname)
            ws_report.cell(row=ri, column=2, value=fpath)
            ws_report.cell(row=ri, column=3, value=status)

        wb_report.save(report_path)
        print(f"\n✓ Deletion report saved: {report_path}")

    print(f"\nTotal files processed for deletion: {len(deleted_files)}")


# ── TASK 3: Organize by date ──────────────────
def organize_by_date(records, organized_folder):
    """
    Create folder structure YYYY/YYYYMM and copy files
    Returns: list of (source, destination) tuples
    """
    print("\n" + "="*60)
    print("TASK 3: ORGANIZE BY DATE")
    print("="*60)

    organized_folder = Path(organized_folder)
    organized_folder.mkdir(parents=True, exist_ok=True)

    movements = []

    for rec in records:
        # Skip if marked for deletion
        if rec.get('delete_flag', '').upper() == 'YES':
            continue

        src_path = Path(rec['full_path'])

        # Get date from metadata or file modified date
        date_taken = rec.get('date_taken')
        if isinstance(date_taken, datetime):
            year_month = date_taken.strftime('%Y%m')
            year = date_taken.strftime('%Y')
        else:
            # Fallback to file modified date
            try:
                file_dt = datetime.strptime(rec['file_modified'], '%Y-%m-%d %H:%M:%S')
                year_month = file_dt.strftime('%Y%m')
                year = file_dt.strftime('%Y')
            except:
                year = '9999_Undated'
                year_month = '999999_Undated'

        # Create folder structure
        dest_folder = organized_folder / year / year_month
        dest_folder.mkdir(parents=True, exist_ok=True)

        dest_path = dest_folder / src_path.name

        # Handle filename conflicts
        counter = 1
        while dest_path.exists():
            stem = src_path.stem
            suffix = src_path.suffix
            dest_path = dest_folder / f"{stem}_{counter}{suffix}"
            counter += 1

        try:
            shutil.copy2(src_path, dest_path)
            movements.append({
                'source_filename': src_path.name,
                'source_path': str(src_path),
                'destination_path': str(dest_path),
                'year_folder': year,
                'month_folder': year_month,
                'status': 'Copied'
            })
            print(f"✓ {src_path.name} -> {year}/{year_month}/")
        except Exception as e:
            movements.append({
                'source_filename': src_path.name,
                'source_path': str(src_path),
                'destination_path': '',
                'year_folder': year,
                'month_folder': year_month,
                'status': f'Error: {str(e)[:50]}'
            })
            print(f"✗ Error copying {src_path.name}: {e}")

    # Create movement report
    if movements:
        report_path = Path(organized_folder).parent / f"organization_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb_report = openpyxl.Workbook()
        ws_report = wb_report.active
        ws_report.title = 'Organization Report'

        headers = ['Source Filename', 'Source Path', 'Year', 'Month', 'Destination Path', 'Status']
        for ci, header in enumerate(headers, 1):
            c = ws_report.cell(row=1, column=ci, value=header)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', start_color='2E4057')
            ws_report.column_dimensions[get_column_letter(ci)].width = 40

        for ri, mov in enumerate(movements, 2):
            ws_report.cell(row=ri, column=1, value=mov['source_filename'])
            ws_report.cell(row=ri, column=2, value=mov['source_path'])
            ws_report.cell(row=ri, column=3, value=mov['year_folder'])
            ws_report.cell(row=ri, column=4, value=mov['month_folder'])
            ws_report.cell(row=ri, column=5, value=mov['destination_path'])
            ws_report.cell(row=ri, column=6, value=mov['status'])

        wb_report.save(report_path)
        print(f"\n✓ Organization report saved: {report_path}")

    print(f"\nTotal files organized: {len(movements)}")
    return movements


# ── Main ───────────────────────────────────────
if __name__ == '__main__':

    if not PIL_OK or not OPENPYXL_OK:
        print("\nPlease install missing packages and re-run.")
        print("Run: pip install pillow openpyxl opencv-python")
        exit(1)

    # TASK 1: Scan and generate metadata
    print("\n" + "="*60)
    print("TASK 1: SCAN & EXTRACT METADATA")
    print("="*60)

    records = scan(SCAN_FOLDER)

    if not records:
        print("No images found. Check your SCAN_FOLDER path.")
        exit(0)

    out = Path(__file__).parent / OUTPUT_FILE
    excel_path = write_excel(records, str(out))

    dup_count = sum(1 for r in records if r['is_duplicate'] == 'YES')
    dup_grps  = len(set(r['duplicate_group'] for r in records if r['duplicate_group']))
    blur_count = sum(1 for r in records if r['is_blurry'] == True)

    print(f"""
+--------------------------------------+
|          SCAN COMPLETE               |
+--------------------------------------+
  Total images    : {len(records)}
  Blurry images   : {blur_count}
  Duplicate files : {dup_count}
  Duplicate groups: {dup_grps}
  Output file     : {excel_path}
+--------------------------------------+

NEXT STEPS:
1. Open the Excel file above
2. Review "Blurry Images" sheet
3. Review "Duplicates" sheet
4. Mark files with 'Yes' in "DELETE? (Yes/No)" column
5. Run Task 2 to delete marked files
6. Run Task 3 to organize remaining files

""")

# TASK 2: Delete marked files (uncomment after marking files in Excel)
# delete_marked_files(excel_path)

# TASK 3: Organize by date (uncomment after deletions)
# organize_by_date(records, ORGANIZED_FOLDER)

