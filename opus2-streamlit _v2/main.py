# ============================================================
# FILE: main.py
# ============================================================
"""Image Scanner v2.1 - Main Entry Point"""

import sys
import logging
import pickle
from pathlib import Path
from typing import Optional

from src.config_manager import ConfigManager
from src.scanner import ImageScanner
from src.duplicate_handler import DuplicateHandler
from src.organizer import ImageOrganizer
from src.excel_writer import ExcelWriter

BACKUP_FILE = 'records-backup.pkl'


def setup_logging(config):
    lv = config.get('logging.level', 'INFO')
    lf = config.get('logging.file', './logs/image-scanner.log')
    Path(lf).parent.mkdir(parents=True, exist_ok=True)
    handlers = [logging.FileHandler(lf, encoding='utf-8')]
    if config.get('logging.console', True):
        handlers.append(logging.StreamHandler())
    logging.basicConfig(
        level=getattr(logging, str(lv).upper(), logging.INFO),
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=handlers)
    return logging.getLogger(__name__)


def save_backup(records, fn=BACKUP_FILE):
    try:
        with open(fn, 'wb') as f:
            pickle.dump(records, f, protocol=pickle.HIGHEST_PROTOCOL)
        print(f"  ✓ Backup: {fn} ({len(records)} records)")
    except Exception as e:
        print(f"  ⚠ Backup failed: {e}")


def load_backup(fn=BACKUP_FILE):
    try:
        if not Path(fn).exists():
            print(f"  ✗ Not found: {fn}")
            return None
        with open(fn, 'rb') as f:
            r = pickle.load(f)
        print(f"  ✓ Loaded {len(r)} records")
        return r
    except Exception as e:
        print(f"  ✗ Load error: {e}")
        return None


def print_banner():
    print("""
    ╔═══════════════════════════════════════════════════════╗
    ║      IMAGE SCANNER - Professional Edition v2.1        ║
    ║   Metadata · Blur · Duplicates · Faces · Tags · AI    ║
    ╚═══════════════════════════════════════════════════════╝""")


def print_menu():
    print("""
    ┌─────────────────────────────────────────────────────┐
    │ 1.  Scan & Extract Metadata                        │
    │ 1b. Resume Excel from Backup                       │
    │ 2.  Delete Marked Files                            │
    │ 3.  Organize by Date                               │
    │ 4.  Full Workflow (1→2→3)                          │
    │ 5.  Launch Web Dashboard (Streamlit)               │
    │ 6.  Generate Comparison Pages                      │
    │ 0.  Exit                                           │
    └─────────────────────────────────────────────────────┘""")


def box(title, rows):
    w = 55
    print(f"\n    ╔{'═'*w}╗\n    ║  {title:<{w-2}}║\n    ╠{'═'*w}╣")
    for l, v in rows:
        line = f"  {l}: {v}"
        print(f"    ║{line:<{w}}║")
    print(f"    ╚{'═'*w}╝")


def task_1(config, logger):
    print("\n" + "="*60 + "\n  TASK 1: SCAN & EXTRACT METADATA\n" + "="*60)
    records = None
    try:
        scanner = ImageScanner(config.to_dict())
        sf = config.get('scan.folder_path')
        print(f"\n  Scanning: {sf}")
        if config.get('processing.fast_mode'):
            print("  ⚡ FAST MODE: blur, face, tags, thumbnails disabled")

        records = scanner.scan(sf)
        if not records:
            print("  ⚠ No files found!")
            return None

        img_count = sum(1 for r in records if r.get('file_type') == 'image')
        vid_count = sum(1 for r in records if r.get('file_type') == 'video')
        print(f"\n  ✓ {len(records)} files (Images: {img_count}, Videos: {vid_count})")
        save_backup(records)

        # Duplicates
        if config.get('duplicates.enabled', True):
            print("\n  Detecting duplicates...")
            dh = DuplicateHandler(
                hash_algorithm=config.get('duplicates.hash_algorithm', 'md5'),
                selection_criteria=config.get('duplicates.selection_criteria',
                                               ['quality', 'resolution', 'date', 'size']),
                match_mode=config.get('duplicates.match_mode', 'exact'),
                similarity_threshold=config.get('duplicates.similarity_threshold', 90))
            records = dh.mark_duplicates(records)

        # Clustering
        if config.get('clustering.enabled', False):
            print("  Clustering images...")
            try:
                from src.image_clusterer import ImageClusterer
                cl = ImageClusterer(
                    n_clusters=config.get('clustering.n_clusters', 10),
                    method=config.get('clustering.method', 'color_histogram'),
                    min_cluster_size=config.get('clustering.min_cluster_size', 3))
                records = cl.cluster(records)
            except Exception as e:
                logger.warning(f"Clustering failed: {e}")

        save_backup(records)

        # Analytics
        analytics_data = None
        if config.get('analytics.enabled', True):
            try:
                from src.analytics import StorageAnalytics
                analytics_data = StorageAnalytics().analyze(records)
            except Exception as e:
                logger.warning(f"Analytics failed: {e}")

        # Stats
        dc = sum(1 for r in records if str(r.get('is_duplicate', '')).upper() == 'YES')
        dg = len(set(r.get('duplicate_group') for r in records
                      if r.get('duplicate_group') and str(r.get('duplicate_group')).strip()))
        bc = sum(1 for r in records if r.get('is_blurry') is True)
        dl = sum(1 for r in records if str(r.get('delete_flag', '')).strip().lower() in ('yes', 'true', '1'))
        fc = sum(r.get('face_count', 0) for r in records if isinstance(r.get('face_count'), int))
        gc = sum(1 for r in records if r.get('location_name'))
        tc = sum(1 for r in records if r.get('auto_tags'))
        cc = len(set(r.get('cluster_label') for r in records if r.get('cluster_label')))
        full_exif = sum(1 for r in records if r.get('metadata_status') == 'Full EXIF')
        no_exif = sum(1 for r in records if r.get('metadata_status') == 'No EXIF')

        print(f"  ✓ {bc} blurry | {dc} duplicates ({dg} groups) | {dl} delete-marked")
        print(f"  ✓ Metadata: {full_exif} full EXIF, {no_exif} no EXIF")
        if fc:
            print(f"  ✓ {fc} faces detected")
        if gc:
            print(f"  ✓ {gc} geolocated")
        if tc:
            print(f"  ✓ {tc} auto-tagged")
        if cc:
            print(f"  ✓ {cc} clusters")

        # Comparisons
        if config.get('comparison.enabled', True):
            try:
                from src.comparison_generator import ComparisonGenerator
                cg = ComparisonGenerator(config.get('comparison.output_folder', './comparisons'))
                pages = cg.generate(records)
                if pages:
                    print(f"  ✓ {len(pages)} comparison pages")
            except Exception as e:
                logger.warning(f"Comparison gen failed: {e}")

        # Excel
        print("\n  Generating Excel...")
        ew = ExcelWriter(config.to_dict())
        ep = ew.write(records, sf, analytics_data=analytics_data)

        box("TASK 1 COMPLETE", [
            ("Files", len(records)),
            ("Images", img_count),
            ("Videos", vid_count),
            ("Full EXIF", full_exif),
            ("No EXIF", no_exif),
            ("Blurry", bc),
            ("Duplicates", dc),
            ("Faces", fc),
            ("Geolocated", gc),
            ("Excel", ep),
        ])
        return ep

    except Exception as e:
        logger.error(f"Task 1: {e}", exc_info=True)
        print(f"\n  ✗ {e}")
        if records:
            save_backup(records)
        return None


def task_1b(config, logger):
    print("\n  TASK 1B: RESUME EXCEL")
    records = load_backup()
    if not records:
        return None
    try:
        analytics_data = None
        try:
            from src.analytics import StorageAnalytics
            analytics_data = StorageAnalytics().analyze(records)
        except Exception:
            pass
        ew = ExcelWriter(config.to_dict())
        ep = ew.write(records, config.get('scan.folder_path'), analytics_data=analytics_data)
        box("TASK 1B COMPLETE", [("Records", len(records)), ("Excel", ep)])
        return ep
    except Exception as e:
        logger.error(f"Task 1B: {e}", exc_info=True)
        print(f"  ✗ {e}")
        return None


def task_2(excel_path, logger):
    print("\n  TASK 2: DELETE MARKED FILES")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        sn = next((n for n in ['Duplicates', 'All Images'] if n in wb.sheetnames), None)
        if not sn:
            print("  ✗ No sheet")
            return
        ws = wb[sn]
        dc = fc = nc = None
        for ci, c in enumerate(ws[1], 1):
            if not c.value:
                continue
            h = str(c.value).strip().lower()
            if 'delete' in h:
                dc = ci
            elif h == 'full path':
                fc = ci
            elif h == 'filename':
                nc = ci
        if not dc or not fc:
            print("  ✗ Columns not found")
            return

        to_del = []
        for ri in range(2, ws.max_row + 1):
            v = ws.cell(row=ri, column=dc).value
            if v and str(v).strip().lower() in ('yes', 'true', '1'):
                to_del.append((
                    ws.cell(row=ri, column=fc).value,
                    ws.cell(row=ri, column=nc).value if nc else '?'))

        if not to_del:
            print("  ✓ Nothing to delete")
            return

        print(f"  {len(to_del)} files marked.")
        if input("  Confirm? (yes/no): ").strip().lower() != 'yes':
            print("  Cancelled")
            return

        ok = err = nf = 0
        for fp, fn in to_del:
            if not fp:
                err += 1
                continue
            try:
                p = Path(fp)
                if p.exists():
                    p.unlink()
                    ok += 1
                    print(f"  ✓ {fn}")
                else:
                    nf += 1
            except Exception:
                err += 1

        box("TASK 2 DONE", [("Deleted", ok), ("Not found", nf), ("Errors", err)])

    except Exception as e:
        logger.error(f"Task 2: {e}", exc_info=True)
        print(f"  ✗ {e}")


def task_3(excel_path, config, logger):
    print("\n  TASK 3: ORGANIZE BY DATE")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        sn = next((n for n in ['All Images', 'Duplicates'] if n in wb.sheetnames), None)
        if not sn:
            print("  ✗ No sheet")
            return
        ws = wb[sn]
        cm = {str(c.value).strip(): ci for ci, c in enumerate(ws[1], 1) if c.value}

        h2k = {
            'Filename': 'filename', 'Folder': 'folder', 'Full Path': 'full_path',
            'Date Taken': 'date_taken', 'File Modified': 'file_modified',
            'DELETE? (Yes/No)': 'delete_flag', 'Duplicate?': 'is_duplicate',
            'Best?': 'is_best_in_group', 'Dup Group': 'duplicate_group',
            'Group': 'duplicate_group', 'Recommendation': 'recommendation',
            'Type': 'file_type', 'Format': 'extension',
            'Size (MB)': 'size_mb', 'Quality %': 'quality_score',
            'Metadata Status': 'metadata_status', 'Date Source': 'date_source',
        }

        records = []
        for ri in range(2, ws.max_row + 1):
            rec = {}
            for h, ci in cm.items():
                v = ws.cell(row=ri, column=ci).value
                rec[h2k.get(h, h.lower().replace(' ', '_'))] = v
                rec[h] = v
            records.append(rec)

        print(f"  Loaded {len(records)} records")
        if config.get('organization.reuse_existing_folders'):
            print("  ✓ Reuse existing folders: ON")
        if config.get('organization.video_subfolder'):
            print("  ✓ Video subfolder: ON")

        org = ImageOrganizer(config.to_dict())
        mvs = org.organize(records)

        s = sum(1 for m in mvs if m['status'] == 'Success')
        e = sum(1 for m in mvs if 'Error' in m['status'])

        box("TASK 3 DONE", [
            ("Organized", s),
            ("Errors", e),
            ("Output", config.get('organization.output_folder', '')),
        ])

    except Exception as e:
        logger.error(f"Task 3: {e}", exc_info=True)
        print(f"  ✗ {e}")


def task_5_streamlit():
    print("\n  Launching Streamlit Dashboard...")
    try:
        import subprocess
        app_path = Path(__file__).parent / 'web' / 'streamlit_app.py'
        if not app_path.exists():
            print(f"  ✗ Not found: {app_path}")
            return
        subprocess.Popen([sys.executable, '-m', 'streamlit', 'run', str(app_path),
                          '--server.headless', 'true'])
        print("  ✓ Dashboard at http://localhost:8501")
    except Exception as e:
        print(f"  ✗ {e}")


def task_6_comparisons(config, logger):
    print("\n  TASK 6: GENERATE COMPARISONS")
    records = load_backup()
    if not records:
        return
    try:
        from src.comparison_generator import ComparisonGenerator
        cg = ComparisonGenerator(config.get('comparison.output_folder', './comparisons'))
        pages = cg.generate(records)
        print(f"  ✓ {len(pages)} pages in {config.get('comparison.output_folder', './comparisons')}")
    except Exception as e:
        logger.error(f"Comparison: {e}")
        print(f"  ✗ {e}")


def main():
    try:
        print_banner()
        config = ConfigManager()
        logger = setup_logging(config)
        logger.info("Image Scanner v2.1 started")
        print(f"  Config: {config.config_path}")
        print(f"  Scan: {config.get('scan.folder_path')}")

        features = []
        if config.get('face_detection.enabled'):
            features.append('faces')
        if config.get('auto_tagging.enabled'):
            features.append('tags')
        if config.get('geocoding.enabled'):
            features.append('geo')
        if config.get('clustering.enabled'):
            features.append('clusters')
        if config.get('thumbnails.enabled'):
            features.append('thumbs')
        if config.get('processing.fast_mode'):
            features.append('FAST-MODE')
        if features:
            print(f"  Features: {', '.join(features)}")

        if config.get('organization.reuse_existing_folders'):
            print("  Organization: reuse existing folders ON")
        if config.get('organization.video_subfolder'):
            print("  Organization: video subfolder ON")

        if not config.validate():
            if input("  Continue? (yes/no): ").strip().lower() != 'yes':
                return

        last = None
        while True:
            print_menu()
            ch = input("  Choice: ").strip().lower()

            if ch == '1':
                last = task_1(config, logger)
            elif ch == '1b':
                last = task_1b(config, logger)
            elif ch == '2':
                p = input("  Excel (Enter=last): ").strip() or last
                if p and Path(p).exists():
                    task_2(p, logger)
                else:
                    print(f"  ✗ Not found: {p}")
            elif ch == '3':
                p = input("  Excel (Enter=last): ").strip() or last
                if p and Path(p).exists():
                    task_3(p, config, logger)
                else:
                    print(f"  ✗ Not found: {p}")
            elif ch == '4':
                last = task_1(config, logger)
                if last:
                    if input("  Delete? (yes/no): ").strip().lower() == 'yes':
                        task_2(last, logger)
                    if input("  Organize? (yes/no): ").strip().lower() == 'yes':
                        task_3(last, config, logger)
            elif ch == '5':
                task_5_streamlit()
            elif ch == '6':
                task_6_comparisons(config, logger)
            elif ch == '0':
                print("\n  ✓ Goodbye!")
                break
            else:
                print("  ✗ Invalid")

            input("\n  Enter to continue...")

    except KeyboardInterrupt:
        print("\n  ✗ Interrupted")
        sys.exit(1)
    except Exception as e:
        print(f"\n  ✗ Fatal: {e}")
        logging.error(f"Fatal: {e}", exc_info=True)
        sys.exit(1)


if __name__ == '__main__':
    main()
