"""Image Scanner v4.1"""

import sys
import logging
import pickle
from pathlib import Path

from src.config_manager import ConfigManager
from src.scanner import ImageScanner
from src.duplicate_handler import DuplicateHandler
from src.organizer import ImageOrganizer
from src.excel_writer import ExcelWriter
from src.face_indexer import FaceIndexer
from src.metadata_store import MetadataStore
from src.people_sync import sync_people_tags
from src.metadata_reconcile import reconcile_vault_paths, auto_reconcile_if_enabled
from src.vault_maintenance import (
    dedupe_vault,
    delete_media_cascade,
    cleanup_untagged_orphans,
    save_last_excel_path,
    load_last_excel_path,
    rebuild_vault_index,
)


from src.workspace_paths import records_backup_path


def setup_log(config):
    lf = config.get('logging.file')
    if not lf:
        raise ValueError('logging.file not resolved; set workspace.root in config.yaml')
    Path(lf).parent.mkdir(parents=True, exist_ok=True)
    handlers = [logging.FileHandler(lf, encoding='utf-8')]
    if config.get('logging.console', True):
        handlers.append(logging.StreamHandler())
    logging.basicConfig(
        level=getattr(logging, str(config.get('logging.level', 'INFO')).upper(), logging.INFO),
        format='%(asctime)s-%(name)s-%(levelname)s-%(message)s',
        handlers=handlers
    )
    return logging.getLogger(__name__)


def save_bk(records, config):
    path = records_backup_path(config)
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, 'wb') as f:
            pickle.dump(records, f, protocol=pickle.HIGHEST_PROTOCOL)
        print('  Backup: ' + str(len(records)) + ' records -> ' + str(path))
    except Exception as e:
        print('  Backup failed: ' + str(e))


def load_bk(config):
    path = records_backup_path(config)
    if not path.exists():
        print('  No backup')
        return None
    try:
        with open(path, 'rb') as f:
            r = pickle.load(f)
        print('  Loaded ' + str(len(r)) + ' records from ' + str(path))
        return r
    except Exception as e:
        print('  Load error: ' + str(e))
        return None


def task_1(config, logger):
    print('\n  STEP 1: GENERATE / REFRESH METADATA')
    print('  ' + '=' * 50)
    records = None
    try:
        scanner = ImageScanner(config.to_dict())
        sf = config.get('scan.folder_path')
        print('  Scanning: ' + str(sf))
        records = scanner.scan(sf)
        if not records:
            print('  No files!')
            return None
        print('  Found ' + str(len(records)) + ' files')
        for r in records:
            if not r.get('delete_flag'):
                r['delete_flag'] = 'No'
        if config.get('duplicates.enabled', True):
            print('  Detecting duplicates...')
            records = DuplicateHandler(
                hash_algorithm=config.get('duplicates.hash_algorithm', 'md5'),
                selection_criteria=config.get(
                    'duplicates.selection_criteria',
                    ['quality', 'resolution', 'date', 'size']
                )
            ).mark_duplicates(records)
        else:
            for r in records:
                r['is_duplicate'] = 'No'
                r['duplicate_group'] = ''
                r['is_best_in_group'] = ''
                r['recommendation'] = ''
        if config.get('similar_detection.enabled', False):
            print('  Detecting similar...')
            from src.similar_detector import SimilarDetector
            sd = SimilarDetector(config.to_dict())
            records = sd.compute_hashes(records)
            sg = sd.find_similar(records)
            records = sd.mark_similar(records, sg)
        else:
            for r in records:
                r['is_similar'] = 'No'
                r['similar_group'] = ''
                r['similar_score'] = ''
                r['similar_methods'] = ''
        md_store = MetadataStore(config.to_dict())
        records = md_store.upsert_records(records)
        if config.get('metadata.dedupe_after_scan', True):
            dedupe_vault(config.to_dict(), quiet=True)
            records = MetadataStore(config.to_dict()).load_records()
        print('  Metadata JSON: ' + str(len(records)) + ' records')
        save_bk(records, config)
        return str(md_store.root)
    except Exception as e:
        logger.error('Task 1: %s', e, exc_info=True)
        print('  Error: ' + str(e))
        if records:
            save_bk(records, config)
        return None


def task_1b(config, logger):
    print('\n  STEP 2: GENERATE EXCEL FROM METADATA')
    cfg = config.to_dict()
    if config.get('metadata.dedupe_before_excel', True):
        dedupe_vault(cfg, quiet=True)
    records = MetadataStore(cfg).load_records()
    if not records:
        records = load_bk(config)
    if not records:
        return None
    missing = sum(1 for r in records if str(r.get('file_exists', '')).lower() != 'yes')
    if missing:
        print('  Note: ' + str(missing) + ' vault row(s) have no file on disk (see File Exists? column)')
    if config.get('workflow.reset_dup_sim_for_excel', False):
        for r in records:
            r['is_duplicate'] = 'No'
            r['is_similar'] = 'No'
    for r in records:
        if not r.get('delete_flag'):
            r['delete_flag'] = 'No'
    ad = None
    try:
        from src.analytics import StorageAnalytics
        ad = StorageAnalytics().analyze(records)
    except Exception:
        pass
    ep = ExcelWriter(cfg).write(
        records, config.get('scan.folder_path'), analytics_data=ad
    )
    if ep:
        save_last_excel_path(config, ep)
        print('  Excel: ' + str(ep))
        if config.get('comparison.generate_after_excel', False):
            try:
                from src.comparison_generator import ComparisonGenerator
                out = config.get('comparison.output_folder')
                if out:
                    ComparisonGenerator(out).generate(records)
                    print('  Comparisons: ' + str(out))
            except Exception as e:
                print('  Comparison skipped: ' + str(e))
    return ep


def task_reconcile_paths(config, logger):
    print('\n  UPDATE VAULT FULL PATHS (fast reconcile)')
    print('  ' + '=' * 50)
    try:
        stats = reconcile_vault_paths(config.to_dict())
        if stats.get('dedupe_removed'):
            print('  Dedupe removed: ' + str(stats['dedupe_removed']) + ' duplicate json')
    except Exception as e:
        logger.error('Reconcile: %s', e, exc_info=True)
        print('  Error: ' + str(e))


def task_dedupe_vault(config, logger):
    print('\n  DEDUPE METADATA VAULT (same file on disk)')
    print('  ' + '=' * 50)
    try:
        dedupe_vault(config.to_dict())
    except Exception as e:
        print('  Error: ' + str(e))


def task_cleanup_untagged(config, logger):
    print('\n  CLEANUP UNTAGGED SAMPLE FOLDERS')
    print('  ' + '=' * 50)
    try:
        cleanup_untagged_orphans(config.to_dict())
    except Exception as e:
        print('  Error: ' + str(e))


def task_2(ep, config, logger):
    print('\n  STEP 3: APPLY EXCEL DELETE ACTIONS')
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ep)
        targets = []
        for sn in ['All Images', 'Duplicates', 'Similar Images']:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            hdr = {str(c.value).strip().lower(): ci for ci, c in enumerate(ws[1], 1) if c.value}
            fci = hdr.get('full path') or hdr.get('path')
            if not fci:
                continue
            dci = None
            sci = None
            dupi = None
            mci = hdr.get('metadata json path')
            midci = hdr.get('media id')
            for h, ci in hdr.items():
                if 'delete' in h:
                    dci = ci
                if h == 'similar?' or 'similar?' in h:
                    sci = ci
                if h == 'duplicate?' or 'dup?' in h:
                    dupi = ci
            for ri in range(2, ws.max_row + 1):
                fp = ws.cell(row=ri, column=fci).value
                if not fp:
                    continue
                delete_flag = ws.cell(row=ri, column=dci).value if dci else ''
                similar_flag = ws.cell(row=ri, column=sci).value if sci else ''
                dup_flag = ws.cell(row=ri, column=dupi).value if dupi else ''
                delete_me = any(
                    str(v).strip().lower() in ('yes', 'true', '1')
                    for v in [delete_flag, similar_flag, dup_flag]
                )
                if not delete_me:
                    continue
                mp = ws.cell(row=ri, column=mci).value if mci else ''
                mid = ws.cell(row=ri, column=midci).value if midci else ''
                targets.append({
                    'full_path': str(fp),
                    'metadata_json_path': str(mp or '').strip(),
                    'media_id': str(mid or '').strip(),
                })

        if not targets:
            print('  Nothing to delete')
            return
        uniq = {}
        for t in targets:
            uniq[t['full_path']] = t
        print('  ' + str(len(uniq)) + ' files marked')
        if input('  Confirm? (yes/no): ').strip().lower() != 'yes':
            return
        stats = delete_media_cascade(config.to_dict(), list(uniq.values()), delete_files=True)
        print(
            '  Images deleted: ' + str(stats['files_deleted'])
            + ' | missing: ' + str(stats['files_missing'])
        )
        print(
            '  Metadata deleted: ' + str(stats['json_deleted'])
            + ' | untagged dirs: ' + str(stats['untagged_removed'])
            + ' | face index rows: ' + str(stats['face_rows_removed'])
        )
        dedupe_vault(config.to_dict(), quiet=True)
        auto_reconcile_if_enabled(config.to_dict(), 'after delete')
        records = MetadataStore(config.to_dict()).load_records()
        if records:
            save_bk(records, config)
    except Exception as e:
        print('  Error: ' + str(e))


def task_3(ep, config, logger):
    print('\n  STEP 4: ORGANIZE IMAGES + METADATA')
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ep)
        sn = next((n for n in ['All Images', 'Duplicates'] if n in wb.sheetnames), None)
        if not sn:
            print('  No sheet')
            return
        ws = wb[sn]
        cm = {str(c.value).strip(): ci for ci, c in enumerate(ws[1], 1) if c.value}
        h2k = {
            'Media ID': 'media_id',
            'Filename': 'filename', 'File': 'filename', 'Folder': 'folder',
            'Full Path': 'full_path', 'Path': 'full_path',
            'Date Taken': 'date_taken', 'Date': 'date_taken',
            'File Modified': 'file_modified',
            'Manual Date Override': 'manual_date_override',
            'Effective Organize Date': 'effective_organize_date',
            'Effective Date Source': 'effective_date_source',
            'Metadata JSON Path': 'metadata_json_path',
            'DELETE?': 'delete_flag', 'Dup?': 'is_duplicate',
            'Type': 'file_type', 'Format': 'extension', 'Fmt': 'extension',
            'Size MB': 'size_mb', 'Size': 'size_mb',
            'Width': 'width', 'W': 'width',
            'Height': 'height', 'H': 'height',
            'EXIF': 'has_exif', 'Metadata': 'metadata_status',
        }
        records = []
        for ri in range(2, ws.max_row + 1):
            rec = {}
            for h, ci in cm.items():
                v = ws.cell(row=ri, column=ci).value
                rec[h2k.get(h, h.lower().replace(' ', '_'))] = v
                rec[h] = v
            records.append(rec)
        print('  Loaded ' + str(len(records)) + ' records')
        mvs = ImageOrganizer(config.to_dict()).organize(records)
        s = sum(1 for m in mvs if m['status'] == 'Success')
        e = sum(1 for m in mvs if 'Error' in m['status'])
        print('  Done: ' + str(s) + ' success, ' + str(e) + ' errors')
        auto_reconcile_if_enabled(config.to_dict(), 'after organize')
        dedupe_vault(config.to_dict(), quiet=True)
        records = MetadataStore(config.to_dict()).load_records()
        if records:
            save_bk(records, config)
    except Exception as e:
        print('  Error: ' + str(e))


def task_5(config, logger):
    print('\n  STEP 5: BUILD / UPDATE FACE INDEX')
    print('  ' + '=' * 50)
    try:
        fd = str(config.get('faces.data_folder', '') or '').strip()
        if fd:
            Path(fd).mkdir(parents=True, exist_ok=True)
        fi = FaceIndexer(config.to_dict())
        files, faces = fi.build_or_update_index(recursive=True)
        print(f'  Updated: {files} files | {faces} faces')
        print(f'  DB: {fi.index_db}')
    except Exception as e:
        print('  Error: ' + str(e))


def task_6(config, logger):
    print('\n  STEP 6: PEOPLE TAG SYNC + UNTAGGED SAMPLES')
    print('  ' + '=' * 50)
    try:
        md_store = MetadataStore(config.to_dict())
        records = md_store.load_records()
        if not records:
            records = load_bk(config)
        if not records:
            print('  No records. Run Step 1 first.')
            return None

        fi = FaceIndexer(config.to_dict())
        matches = fi.find_person()
        untagged_root = Path(config.get('faces.untagged_root'))
        try:
            ums = int(config.get('faces.untagged_max_samples', 1) or 1)
        except (TypeError, ValueError):
            ums = 1
        ums = max(1, min(20, ums))
        pick_q = config.get('faces.untagged_pick_best_quality', True)
        if isinstance(pick_q, str):
            pick_q = pick_q.strip().lower() in ('1', 'true', 'yes', 'on')
        export_ut = config.get('faces.export_untagged', True)
        if isinstance(export_ut, str):
            export_ut = export_ut.strip().lower() in ('1', 'true', 'yes', 'on')
        known, unknown = sync_people_tags(
            records,
            matches,
            untagged_root,
            export_untagged=bool(export_ut),
            seed_only_refresh=False,
            untagged_max_samples=ums,
            untagged_pick_best_quality=bool(pick_q),
            untagged_export_mode=str(config.get('faces.untagged_export_mode', 'full') or 'full'),
            config=config.to_dict(),
        )
        save_bk(records, config)
        print(f'  Known tagged: {known} | Unknown tagged: {unknown}')
        print(f'  Untagged samples folder: {untagged_root}')
        return True
    except Exception as e:
        print('  Error: ' + str(e))
        return None


def task_7(config, logger):
    print('\n  STEP 7: SEED FEEDBACK REFRESH (known matches only, no untagged export)')
    print('  ' + '=' * 50)
    try:
        md_store = MetadataStore(config.to_dict())
        records = md_store.load_records()
        if not records:
            records = load_bk(config)
        if not records:
            print('  No records. Run Step 1 first.')
            return None
        fi = FaceIndexer(config.to_dict())
        matches = fi.find_person()
        untagged_root = Path(config.get('faces.untagged_root'))
        known, unknown = sync_people_tags(
            records, matches, untagged_root, export_untagged=False, seed_only_refresh=True,
            config=config.to_dict(),
        )
        save_bk(records, config)
        print(f'  Known re-tagged: {known}')
        return True
    except Exception as e:
        print('  Error: ' + str(e))
        return None


def task_convert_structure(config, logger):
    print('\n  TOOL: CONVERT FOLDER STRUCTURE')
    print('  ' + '=' * 50)
    source = config.get('organization.output_folder', './organized_images')
    print('  Source folder: ' + str(source))
    if not Path(source).exists():
        print('  Error: Folder not found!')
        return
    print('')
    print('  1. flat   2. year   3. year-month-date')
    ch = input('  Select target (1/2/3): ').strip()
    targets = {'1': 'flat', '2': 'year', '3': 'year-month-date'}
    target = targets.get(ch)
    if not target:
        print('  Invalid choice')
        return
    out = input('  Output folder (Enter=in-place): ').strip()
    target_folder = out if out else None
    if input('  Confirm? (yes/no): ').strip().lower() != 'yes':
        print('  Cancelled')
        return
    ImageOrganizer.convert_structure(source, target, target_folder)
    auto_reconcile_if_enabled(config.to_dict(), 'after convert structure')


def task_merge_dates(config, logger):
    print('\n  TOOL: MERGE SAME-DATE FOLDERS')
    print('  ' + '=' * 50)
    source = config.get('organization.output_folder', './organized_images')
    print('  Source folder: ' + str(source))
    if not Path(source).exists():
        print('  Error: Folder not found!')
        return
    detected = ImageOrganizer.detect_structure(source)
    print('  Detected structure: ' + detected)
    override = input('  Structure override [flat/year/year-month-date] (Enter=detected): ').strip().lower()
    structure = override if override in ('flat', 'year', 'year-month-date') else detected
    print('  Using structure: ' + structure)
    if input('  Confirm? (yes/no): ').strip().lower() != 'yes':
        print('  Cancelled')
        return
    ImageOrganizer.merge_duplicate_dates(source, structure)
    auto_reconcile_if_enabled(config.to_dict(), 'after merge same-date')


def task_8(config, logger):
    print('\n  STEP 8: REFRESH FINAL EXCEL FROM METADATA')
    return task_1b(config, logger)


def _vault_line(config) -> str:
    lines = []
    art = config.artifact_summary() if hasattr(config, 'artifact_summary') else {}
    ws = art.get('workspace') or str(config.get('workspace.root', '') or '').strip()
    if ws:
        lines.append('  Workspace: ' + ws)
    rf = art.get('metadata') or str(config.get('metadata.root_folder', '') or '').strip()
    if rf:
        lines.append('  Metadata vault: ' + rf)
    lr = str(config.get('metadata.library_root', '') or '').strip()
    if lr:
        lines.append('  Library root: ' + lr)
    return '\n'.join(lines)


def menu_metadata(config, logger, last_excel):
    while True:
        print('')
        print('  --- DATA MANAGEMENT ---')
        print(_vault_line(config))
        print('  1.  New Scan (Generate Metadata)')
        print('  2.  Generate Excel Report')
        print('  3.  Execute Deletions (From Excel)')
        print('  4.  Vault Maintenance (Update Paths & Dedupe)')
        print('  5.  Cleanup Untagged Folders')
        print('  0.  Back')
        ch = input('  Choice: ').strip().lower()
        if ch == '0':
            return last_excel
        if ch == '1':
            task_1(config, logger)
        elif ch == '2':
            last_excel = task_1b(config, logger) or last_excel
        elif ch == '3':
            p = input('  Excel (Enter=last): ').strip() or last_excel or load_last_excel_path(config) or ''
            if p and Path(p).exists():
                task_2(p, config, logger)
            else:
                print('  Not found')
        elif ch == '4':
            task_reconcile_paths(config, logger)
            task_dedupe_vault(config, logger)
        elif ch == '5':
            task_cleanup_untagged(config, logger)
        else:
            print('  Invalid')
        input('\n  Enter to continue...')


def menu_organize(config, logger, last_excel):
    while True:
        print('')
        print('  --- Organize library ---')
        print('  Output: ' + str(config.get('organization.output_folder', '')))
        print('  1.  Organize from Excel')
        print('  2.  Convert folder structure')
        print('  3.  Merge same-date folders')
        print('  0.  Back')
        ch = input('  Choice: ').strip().lower()
        if ch == '0':
            return
        if ch == '1':
            p = input('  Excel (Enter=last): ').strip() or last_excel or load_last_excel_path(config) or ''
            if p and Path(p).exists():
                task_3(p, config, logger)
            else:
                print('  Not found')
        elif ch == '2':
            task_convert_structure(config, logger)
        elif ch == '3':
            task_merge_dates(config, logger)
        else:
            print('  Invalid')
        input('\n  Enter to continue...')


def menu_faces(config, logger):
    while True:
        print('')
        print('  --- Faces ---')
        print('  1.  Build / Update Face Index')
        print('  2.  People Tag Sync + Untagged Samples')
        print('  3.  Seed Feedback Refresh')
        print('  0.  Back')
        ch = input('  Choice: ').strip().lower()
        if ch == '0':
            return
        if ch == '1':
            task_5(config, logger)
        elif ch == '2':
            task_6(config, logger)
        elif ch == '3':
            task_7(config, logger)
        else:
            print('  Invalid')
        input('\n  Enter to continue...')


def main():
    try:
        print('')
        print('  ========================================================')
        print('       IMAGE SCANNER v5.0')
        print('  ========================================================')
        config = ConfigManager()
        if not config.validate():
            print('\n  Fix config.yaml and restart.')
            sys.exit(1)
        logger = setup_log(config)
        print('  Config: ' + str(config.config_path))
        art = config.artifact_summary()
        print('  Workspace: ' + str(art.get('workspace', '')))
        print('  Artifacts: metadata, reports, face_data, ... under workspace')
        print('  Scan: ' + str(config.get('scan.folder_path')))
        print('  Structure: ' + str(config.get('organization.folder_structure', 'year')))
        ft = []
        for k, n in [
            ('duplicates.enabled', 'duplicates'),
            ('similar_detection.enabled', 'similar'),
            ('blur_detection.enabled', 'blur'),
            ('face_detection.enabled', 'faces'),
            ('processing.fast_mode', 'FAST'),
        ]:
            if config.get(k):
                ft.append(n)
        if ft:
            print('  Features: ' + ', '.join(ft))
        last_excel = load_last_excel_path(config)
        while True:
            print('')
            print(_vault_line(config))
            print('  1.  FULL AUTOMATIC FLOW (Scan + Delete + Organize)')
            print('  2.  Data Management (Scanning, Excel, Deletion)')
            print('  3.  Library Tools (Organize, Convert, Merge)')
            print('  4.  AI & Face Discovery (Indexing, Tagging)')
            print('  0.  Exit')
            print('')
            ch = input('  Choice: ').strip().lower()
            if ch == '1':
                print('\n  --- FULL AUTOMATIC FLOW ---')
                task_1(config, logger)
                last_excel = task_1b(config, logger)
                if last_excel and input('\n  Apply Excel Delete Actions? (yes/no): ').strip().lower() == 'yes':
                    task_2(last_excel, config, logger)
                if last_excel and input('\n  Organize into Library? (yes/no): ').strip().lower() == 'yes':
                    task_3(last_excel, config, logger)
            elif ch == '2':
                last_excel = menu_metadata(config, logger, last_excel) or last_excel
            elif ch == '3':
                menu_organize(config, logger, last_excel)
            elif ch == '4':
                menu_faces(config, logger)
            elif ch == '0':
                print('\n  Bye!')
                break
            else:
                print('  Invalid')
            input('\n  Enter to continue...')
    except KeyboardInterrupt:
        print('\n  Interrupted')
        sys.exit(1)
    except Exception as e:
        print('\n  Fatal: ' + str(e))
        sys.exit(1)


if __name__ == '__main__':
    main()

