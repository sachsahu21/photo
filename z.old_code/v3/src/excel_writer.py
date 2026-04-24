"""
Excel Report Generation
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Generate Excel reports"""

    COLUMNS = [
        ('filename', 'Filename', 28),
        ('folder', 'Folder', 45),
        ('extension', 'Format', 10),
        ('size_mb', 'Size (MB)', 11),
        ('width', 'Width (px)', 11),
        ('height', 'Height (px)', 11),
        ('mode', 'Color Mode', 12),
        ('dpi', 'DPI', 10),
        ('date_taken', 'Date Taken', 20),
        ('camera_make', 'Camera Make', 16),
        ('camera_model', 'Camera Model', 18),
        ('focal_length', 'Focal Length', 13),
        ('aperture', 'Aperture', 11),
        ('iso', 'ISO', 8),
        ('exposure_time', 'Exposure', 11),
        ('gps_lat', 'GPS Lat', 11),
        ('gps_lon', 'GPS Lon', 11),
        ('has_exif', 'Has EXIF', 10),
        ('blur_score', 'Blur Score', 11),
        ('quality_rating', 'Quality', 12),
        ('quality_score', 'Quality %', 10),
        ('quality_issues', 'Issues', 30),
        ('is_blurry', 'Blurry?', 10),
        ('is_duplicate', 'Duplicate?', 12),
        ('duplicate_group', 'Dup. Group', 10),
        ('is_best_in_group', 'Best?', 8),
        ('recommendation', 'Recommendation', 18),
        ('delete_flag', 'DELETE? (Yes/No)', 15),
        ('md5_hash', 'MD5 Hash', 34),
        ('file_modified', 'File Modified', 20),
        ('full_path', 'Full Path', 55),
        ('error', 'Read Error', 25),
    ]

    def __init__(self, config: Dict):
        """Initialize Excel writer"""
        self.config = config
        self.output_folder = Path(config.get('output.output_folder', './output'))
        self.output_folder.mkdir(parents=True, exist_ok=True)

    def _sanitize_excel_value(self, val):
        """Convert unsupported Python types to Excel-safe values"""

        # tuple → string or flattened value
        if isinstance(val, tuple):
            return "x".join(str(v) for v in val)

        # list → string
        if isinstance(val, list):
            return ",".join(str(v) for v in val)

        # dict → string
        if isinstance(val, dict):
            return str(val)

        return val
    
    def write(self, records: List[Dict], scan_folder: str) -> str:
        """Write Excel report with incremental saves"""
        parent = Path(scan_folder).name
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = self.output_folder / f"image_scan_{parent}_{ts}.xlsx"

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        try:
            print("Writing All Images sheet...")
            self._write_all_images(wb, records)
            wb.save(output_path)
            print("✓ Saved All Images")

            print("Writing Blurry Images sheet...")
            self._write_blurry_images(wb, records)
            wb.save(output_path)
            print("✓ Saved Blurry Images")

            print("Writing Duplicates sheet...")
            self._write_duplicates(wb, records)
            wb.save(output_path)
            print("✓ Saved Duplicates")

            print("Writing Quality Report sheet...")
            self._write_quality_report(wb, records)
            wb.save(output_path)
            print("✓ Saved Quality Report")

            print("Writing Summary sheet...")
            self._write_summary(wb, records, scan_folder)
            wb.save(output_path)
            print("✓ Saved Summary")

            logger.info(f"Excel report saved: {output_path}")
            return str(output_path)

        except Exception as e:
            logger.error(f"Error writing Excel: {e}")
            print(f"✗ Error: {e}")
            raise


    def _write_all_images(self, wb: openpyxl.Workbook, records: List[Dict]):
        """Write all images sheet"""
        ws = wb.create_sheet('All Images', 0)
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 32

        # Headers
        for ci, (_, label, width) in enumerate(self.COLUMNS, 1):
            c = ws.cell(row=1, column=ci, value=label)
            self._style_header(c)
            ws.column_dimensions[get_column_letter(ci)].width = width

        # Data
        alt_fill = PatternFill('solid', start_color='F5F7FA')
        dup_fill = PatternFill('solid', start_color='FFD6D6')
        blur_fill = PatternFill('solid', start_color='FFE8B6')

        for ri, rec in enumerate(records, 2):
            if rec.get('is_blurry') == True:
                fill = blur_fill
            elif rec.get('is_duplicate') == 'YES':
                fill = dup_fill
            else:
                fill = alt_fill if ri % 2 == 0 else None

            for ci, (key, _, _) in enumerate(self.COLUMNS, 1):
                val = self._format_value(rec.get(key, ''))
                val = self._sanitize_excel_value(val)
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = self._border()
                c.alignment = Alignment(vertical='center')
                if fill:
                    c.fill = fill

        ws.auto_filter.ref = f"A1:{get_column_letter(len(self.COLUMNS))}1"

    def _write_blurry_images(self, wb: openpyxl.Workbook, records: List[Dict]):
        """Write blurry images sheet"""
        ws = wb.create_sheet('Blurry Images')
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 32

        blur_cols = [
            ('filename', 'Filename', 28),
            ('folder', 'Folder', 45),
            ('blur_score', 'Blur Score', 11),
            ('quality_rating', 'Quality', 12),
            ('quality_score', 'Quality %', 10),
            ('quality_issues', 'Issues', 30),
            ('width', 'Width (px)', 11),
            ('height', 'Height (px)', 11),
            ('size_mb', 'Size (MB)', 11),
            ('date_taken', 'Date Taken', 20),
            ('delete_flag', 'DELETE? (Yes/No)', 15),
            ('full_path', 'Full Path', 55),
        ]

        # Headers
        for ci, (_, label, width) in enumerate(blur_cols, 1):
            c = ws.cell(row=1, column=ci, value=label)
            self._style_header(c, 'FF8C00')
            ws.column_dimensions[get_column_letter(ci)].width = width

        # Data
        blurry_recs = sorted(
            [r for r in records if r.get('is_blurry') == True],
            key=lambda x: x.get('blur_score', 0)
        )

        fill = PatternFill('solid', start_color='FFE8B6')
        for ri, rec in enumerate(blurry_recs, 2):
            for ci, (key, _, _) in enumerate(blur_cols, 1):
                val = self._format_value(rec.get(key, ''))
                val = self._sanitize_excel_value(val)
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = self._border()
                c.alignment = Alignment(vertical='center')
                c.fill = fill

        ws.auto_filter.ref = f"A1:{get_column_letter(len(blur_cols))}1"

    def _write_duplicates(self, wb: openpyxl.Workbook, records: List[Dict]):
        """Write duplicates sheet"""
        ws = wb.create_sheet('Duplicates')
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 32

        dup_cols = [
            ('duplicate_group', 'Group', 8),
            ('is_best_in_group', 'Best?', 8),
            ('recommendation', 'Recommendation', 18),
            ('filename', 'Filename', 28),
            ('folder', 'Folder', 45),
            ('extension', 'Format', 10),
            ('size_mb', 'Size (MB)', 11),
            ('quality_score', 'Quality %', 10),
            ('delete_flag', 'DELETE? (Yes/No)', 15),
            ('md5_hash', 'MD5 Hash', 34),
            ('full_path', 'Full Path', 55),
        ]

        # Headers
        for ci, (_, label, width) in enumerate(dup_cols, 1):
            c = ws.cell(row=1, column=ci, value=label)
            self._style_header(c, '8B0000')
            ws.column_dimensions[get_column_letter(ci)].width = width

        # Data
        dups = sorted(
            [r for r in records if r.get('is_duplicate') == 'YES'],
            key=lambda x: (x.get('duplicate_group', 0), x['full_path'])
        )

        prev, alt_flag = None, False
        f1 = PatternFill('solid', start_color='FFE8E8')
        f2 = PatternFill('solid', start_color='FFF5F5')

        for ri, rec in enumerate(dups, 2):
            grp = rec.get('duplicate_group')
            if grp != prev:
                alt_flag = not alt_flag
                prev = grp
            fill = f1 if alt_flag else f2

            for ci, (key, _, _) in enumerate(dup_cols, 1):
                val = self._format_value(rec.get(key, ''))
                val = self._sanitize_excel_value(val)
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = self._border()
                c.alignment = Alignment(vertical='center')
                c.fill = fill

        ws.auto_filter.ref = f"A1:{get_column_letter(len(dup_cols))}1"

    def _write_quality_report(self, wb: openpyxl.Workbook, records: List[Dict]):
        """Write quality report sheet"""
        ws = wb.create_sheet('Quality Report')
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        ws['A1'] = 'Quality Analysis Report'
        ws['A1'].font = Font(bold=True, size=14, color='2E4057')

        # Quality distribution
        quality_scores = [r.get('quality_score', 0) for r in records]
        avg_quality = sum(quality_scores) / len(quality_scores) if quality_scores else 0

        rows = [
            ('', ''),
            ('QUALITY STATISTICS', ''),
            ('Average Quality Score', f"{avg_quality:.1f}%"),
            ('Highest Quality', f"{max(quality_scores):.1f}%"),
            ('Lowest Quality', f"{min(quality_scores):.1f}%"),
            ('', ''),
            ('QUALITY DISTRIBUTION', ''),
        ]

        # Count by quality ranges
        excellent = sum(1 for s in quality_scores if s >= 80)
        good = sum(1 for s in quality_scores if 60 <= s < 80)
        fair = sum(1 for s in quality_scores if 40 <= s < 60)
        poor = sum(1 for s in quality_scores if s < 40)

        rows.extend([
            ('Excellent (80-100%)', excellent),
            ('Good (60-79%)', good),
            ('Fair (40-59%)', fair),
            ('Poor (0-39%)', poor),
        ])

        hdr_fill = PatternFill('solid', start_color='2E4057')
        for ri, (label, value) in enumerate(rows, 3):
            a = ws.cell(row=ri, column=1, value=label)
            b = ws.cell(row=ri, column=2, value=value)

            if label in ('QUALITY STATISTICS', 'QUALITY DISTRIBUTION'):
                a.font = Font(bold=True, color='FFFFFF')
                a.fill = hdr_fill
                b.fill = hdr_fill
            elif label:
                a.font = Font(bold=True)

    def _write_summary(self, wb: openpyxl.Workbook, records: List[Dict], scan_folder: str):
        """Write summary sheet"""
        ws = wb.create_sheet('Summary', 0)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        ws['A1'] = 'Image Scan Summary'
        ws['A1'].font = Font(bold=True, size=14, color='2E4057')
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, color='888888')
        ws['A3'] = f"Scanned: {scan_folder}"
        ws['A3'].font = Font(italic=True, color='888888')

        # Statistics
        total = len(records)
        dup_count = sum(1 for r in records if r.get('is_duplicate') == 'YES')
        dup_grps = len(set(r.get('duplicate_group') for r in records if r.get('duplicate_group')))
        blur_count = sum(1 for r in records if r.get('is_blurry') == True)
        with_exif = sum(1 for r in records if r.get('has_exif'))
        with_gps = sum(1 for r in records if r.get('gps_lat'))
        ext_cnts = Counter(r['extension'] for r in records)
        avg_quality = sum(r.get('quality_score', 0) for r in records) / total if total > 0 else 0
        total_size = sum(r.get('size_mb', 0) for r in records)

        rows = [
            ('', ''),
            ('GENERAL', ''),
            ('Total Images Found', total),
            ('Total Folders Scanned', len(set(r['folder'] for r in records))),
            ('Total Size (MB)', round(total_size, 1)),
            ('Average Quality Score', f"{avg_quality:.1f}%"),
            ('', ''),
            ('QUALITY ISSUES', ''),
            ('Blurry Images', blur_count),
            ('Duplicate Files', dup_count),
            ('Duplicate Groups', dup_grps),
            ('', ''),
            ('METADATA', ''),
            ('Files with EXIF', with_exif),
            ('Files with GPS', with_gps),
            ('', ''),
            ('BY FORMAT', ''),
            *[(f'  .{ext.lower()}', cnt) for ext, cnt in sorted(ext_cnts.items(), key=lambda x: -x[1])]
        ]

        hdr_fill = PatternFill('solid', start_color='2E4057')
        acc_fill = PatternFill('solid', start_color='EBF2FF')

        for ri, (label, value) in enumerate(rows, 5):
            a = ws.cell(row=ri, column=1, value=label)
            b = ws.cell(row=ri, column=2, value=value)

            if label in ('GENERAL', 'QUALITY ISSUES', 'METADATA', 'BY FORMAT'):
                a.font = Font(bold=True, color='FFFFFF')
                a.fill = hdr_fill
                b.fill = hdr_fill
            elif label:
                a.font = Font(bold=True, size=11)
                if ri % 2 == 0:
                    a.fill = acc_fill
                    b.fill = acc_fill

    def _style_header(self, cell, color: str = '2E4057'):
        """Style header cell"""
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill('solid', start_color=color)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = self._border()

    def _border(self):
        """Get border style"""
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    def _format_value(self, value):
        """Format value for Excel"""
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, bool):
            return 'Yes' if value else 'No'
        if isinstance(value, str):
            return ''.join(ch for ch in value if ord(ch) >= 32)
        return value
    

# # ============================================================================
# # FILE: src/excel_writer.py
# # ============================================================================
# """
# Excel Writer - Generates Excel reports with image metadata
# """

# import logging
# from pathlib import Path
# from typing import List, Dict, Optional
# from datetime import datetime
# from collections import defaultdict

# try:
#     from openpyxl import Workbook
#     from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
#     from openpyxl.utils import get_column_letter
#     from openpyxl.worksheet.datavalidation import DataValidation
# except ImportError:
#     Workbook = None

# from src.utils import ensure_directory, get_timestamp_string

# logger = logging.getLogger(__name__)


# class ExcelWriter:
#     """Generates Excel reports with image metadata"""

#     def __init__(self, config: Dict):
#         """
#         Initialize Excel writer

#         Args:
#             config: Configuration dictionary
#         """
#         self.config = config
#         self.output_folder = Path(config.get('excel', {}).get('output_folder', './reports'))
#         self.freeze_rows = config.get('excel', {}).get('freeze_rows', 1)
#         self.auto_filter = config.get('excel', {}).get('auto_filter', True)

#         self.color_duplicate = "FFFF0000"
#         self.color_blurry = "FFFFA500"
#         self.color_good = "FF00B050"

#     def write(self, records: List[Dict], scan_folder: str) -> str:
#         """
#         Write Excel report

#         Args:
#             records: List of image records
#             scan_folder: Source folder path

#         Returns:
#             Path to generated Excel file
#         """
#         logger.info(f"Starting Excel generation for {len(records)} records")
#         ensure_directory(self.output_folder)

#         timestamp = get_timestamp_string()
#         filename = f"image_scan_{timestamp}.xlsx"
#         filepath = self.output_folder / filename

#         try:
#             wb = Workbook()
#             wb.remove(wb.active)

#             # Create sheets
#             self._create_all_images_sheet(wb, records)
#             self._create_blurry_images_sheet(wb, records)
#             self._create_duplicates_sheet(wb, records)
#             self._create_quality_report_sheet(wb, records)
#             self._create_summary_sheet(wb, records, scan_folder)

#             wb.save(str(filepath))
#             logger.info(f"Excel file saved: {filepath}")
#             return str(filepath)

#         except Exception as e:
#             logger.error(f"Error writing Excel: {e}", exc_info=True)
#             raise

#     def _create_all_images_sheet(self, wb: Workbook, records: List[Dict]) -> None:
#         """Create 'All Images' sheet"""
#         ws = wb.create_sheet("All Images")

#         headers = [
#             'Filename', 'Full Path', 'File Size (MB)', 'Dimensions', 'Megapixels',
#             'Camera Make', 'Camera Model', 'Focal Length', 'Aperture', 'ISO',
#             'Exposure Time', 'Date Taken', 'File Modified', 'GPS Latitude', 'GPS Longitude',
#             'Blur Score', 'Quality Class', 'Quality Score (%)', 'Is Blurry',
#             'Is Duplicate', 'Duplicate Group', 'File Hash', 'Color Mode', 'DPI',
#             'DELETE? (Yes/No)'
#         ]

#         ws.append(headers)
#         self._style_header(ws)

#         for record in records:
#             row = [
#                 record.get('filename'),
#                 record.get('full_path'),
#                 record.get('file_size_mb'),
#                 f"{record.get('width', 0)}x{record.get('height', 0)}",
#                 record.get('megapixels'),
#                 record.get('camera_make'),
#                 record.get('camera_model'),
#                 record.get('focal_length'),
#                 record.get('aperture'),
#                 record.get('iso'),
#                 record.get('exposure_time'),
#                 record.get('date_taken'),
#                 record.get('file_modified_date'),
#                 record.get('gps_latitude'),
#                 record.get('gps_longitude'),
#                 record.get('blur_score'),
#                 record.get('quality_class'),
#                 record.get('quality_score'),
#                 'Yes' if record.get('is_blurry') else 'No',
#                 record.get('is_duplicate', 'NO'),
#                 record.get('duplicate_group'),
#                 record.get('file_hash'),
#                 record.get('color_mode'),
#                 record.get('dpi'),
#                 ''
#             ]

#             ws.append(row)

#             # Color code rows
#             row_num = ws.max_row
#             if record.get('is_duplicate') == 'YES':
#                 self._color_row(ws, row_num, self.color_duplicate)
#             elif record.get('is_blurry'):
#                 self._color_row(ws, row_num, self.color_blurry)

#         self._auto_fit_columns(ws)
#         self._freeze_header(ws)
#         if self.auto_filter:
#             self._add_filters(ws)

#     def _create_blurry_images_sheet(self, wb: Workbook, records: List[Dict]) -> None:
#         """Create 'Blurry Images' sheet"""
#         blurry_records = [r for r in records if r.get('is_blurry')]
#         blurry_records.sort(key=lambda x: x.get('blur_score', 0))

#         ws = wb.create_sheet("Blurry Images")

#         headers = [
#             'Filename', 'Full Path', 'Blur Score', 'Quality Class', 'Quality Score (%)',
#             'Dimensions', 'Date Taken', 'File Size (MB)', 'DELETE? (Yes/No)'
#         ]

#         ws.append(headers)
#         self._style_header(ws)

#         for record in blurry_records:
#             row = [
#                 record.get('filename'),
#                 record.get('full_path'),
#                 record.get('blur_score'),
#                 record.get('quality_class'),
#                 record.get('quality_score'),
#                 f"{record.get('width', 0)}x{record.get('height', 0)}",
#                 record.get('date_taken'),
#                 record.get('file_size_mb'),
#                 ''
#             ]
#             ws.append(row)
#             self._color_row(ws, ws.max_row, self.color_blurry)

#         self._auto_fit_columns(ws)
#         self._freeze_header(ws)
#         if self.auto_filter:
#             self._add_filters(ws)

#     def _create_duplicates_sheet(self, wb: Workbook, records: List[Dict]) -> None:
#         """Create 'Duplicates' sheet"""
#         ws = wb.create_sheet("Duplicates")

#         headers = [
#             'Filename', 'Full Path', 'File Size (MB)', 'Dimensions', 'Megapixels',
#             'Quality Score (%)', 'Date Taken', 'Duplicate Group', 'Is Best',
#             'Is Duplicate', 'File Hash', 'DELETE? (Yes/No)'
#         ]

#         ws.append(headers)
#         self._style_header(ws)

#         # Group by duplicate group
#         dup_groups = defaultdict(list)
#         for record in records:
#             if record.get('duplicate_group'):
#                 dup_groups[record.get('duplicate_group')].append(record)

#         for group_id in sorted(dup_groups.keys()):
#             group_records = dup_groups[group_id]

#             for record in group_records:
#                 row = [
#                     record.get('filename'),
#                     record.get('full_path'),
#                     record.get('file_size_mb'),
#                     f"{record.get('width', 0)}x{record.get('height', 0)}",
#                     record.get('megapixels'),
#                     record.get('quality_score'),
#                     record.get('date_taken'),
#                     record.get('duplicate_group'),
#                     'Yes' if record.get('is_best_of_duplicates') else 'No',
#                     record.get('is_duplicate', 'NO'),
#                     record.get('file_hash'),
#                     ''
#                 ]
#                 ws.append(row)

#                 row_num = ws.max_row
#                 if record.get('is_duplicate') == 'YES':
#                     self._color_row(ws, row_num, self.color_duplicate)

#         self._auto_fit_columns(ws)
#         self._freeze_header(ws)
#         if self.auto_filter:
#             self._add_filters(ws)

#     def _create_quality_report_sheet(self, wb: Workbook, records: List[Dict]) -> None:
#         """Create 'Quality Report' sheet"""
#         ws = wb.create_sheet("Quality Report")

#         # Quality distribution
#         quality_dist = defaultdict(int)
#         for record in records:
#             quality_class = record.get('quality_class', 'Unknown')
#             quality_dist[quality_class] += 1

#         ws.append(['Quality Class', 'Count', 'Percentage'])
#         self._style_header(ws)

#         total = len(records)
#         for quality_class in ['Very Blurry', 'Blurry', 'Fair', 'Sharp', 'Unknown', 'Error']:
#             count = quality_dist.get(quality_class, 0)
#             percentage = (count / total * 100) if total > 0 else 0
#             ws.append([quality_class, count, f"{percentage:.1f}%"])

#         # Statistics
#         ws.append([])
#         ws.append(['Statistic', 'Value'])
#         self._style_header(ws)

#         avg_quality = sum(r.get('quality_score', 0) for r in records) / len(records) if records else 0
#         avg_blur = sum(r.get('blur_score', 0) for r in records) / len(records) if records else 0
#         blurry_count = sum(1 for r in records if r.get('is_blurry'))
#         dup_count = sum(1 for r in records if r.get('is_duplicate') == 'YES')

#         ws.append(['Total Images', len(records)])
#         ws.append(['Average Quality Score', f"{avg_quality:.1f}%"])
#         ws.append(['Average Blur Score', f"{avg_blur:.2f}"])
#         ws.append(['Blurry Images', blurry_count])
#         ws.append(['Duplicate Files', dup_count])

#         self._auto_fit_columns(ws)

#     def _create_summary_sheet(self, wb: Workbook, records: List[Dict], scan_folder: str) -> None:
#         """Create 'Summary' sheet"""
#         ws = wb.create_sheet("Summary", 0)

#         ws.append(['IMAGE SCAN SUMMARY REPORT'])
#         ws.append([])

#         ws.append(['Scan Information', ''])
#         ws.append(['Scan Folder', scan_folder])
#         ws.append(['Scan Date', datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
#         ws.append([])

#         ws.append(['Statistics', ''])
#         ws.append(['Total Images', len(records)])

#         blurry_count = sum(1 for r in records if r.get('is_blurry'))
#         dup_count = sum(1 for r in records if r.get('is_duplicate') == 'YES')
#         dup_groups = len(set(r.get('duplicate_group') for r in records if r.get('duplicate_group')))

#         ws.append(['Blurry Images', blurry_count])
#         ws.append(['Duplicate Files', dup_count])
#         ws.append(['Duplicate Groups', dup_groups])
#         ws.append([])

#         avg_quality = sum(r.get('quality_score', 0) for r in records) / len(records) if records else 0
#         total_size = sum(r.get('file_size_mb', 0) for r in records)

#         ws.append(['Average Quality Score', f"{avg_quality:.1f}%"])
#         ws.append(['Total Size (MB)', f"{total_size:.2f}"])
#         ws.append([])

#         ws.append(['Quality Distribution', ''])
#         quality_dist = defaultdict(int)
#         for record in records:
#             quality_class = record.get('quality_class', 'Unknown')
#             quality_dist[quality_class] += 1

#         for quality_class in ['Very Blurry', 'Blurry', 'Fair', 'Sharp']:
#             count = quality_dist.get(quality_class, 0)
#             percentage = (count / len(records) * 100) if records else 0
#             ws.append([quality_class, f"{count} ({percentage:.1f}%)"])

#         self._auto_fit_columns(ws)

#     def _style_header(self, ws) -> None:
#         """Style header row"""
#         header_fill = PatternFill(start_color="FF366092", end_color="FF366092", fill_type="solid")
#         header_font = Font(bold=True, color="FFFFFFFF")
#         header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

#         for cell in ws[1]:
#             cell.fill = header_fill
#             cell.font = header_font
#             cell.alignment = header_alignment

#     def _color_row(self, ws, row_num: int, color: str) -> None:
#         """Color entire row"""
#         fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
#         for cell in ws[row_num]:
#             cell.fill = fill

#     def _auto_fit_columns(self, ws) -> None:
#         """Auto-fit column widths"""
#         for column in ws.columns:
#             max_length = 0
#             column_letter = get_column_letter(column[0].column)

#             for cell in column:
#                 try:
#                     if cell.value:
#                         max_length = max(max_length, len(str(cell.value)))
#                 except:
#                     pass

#             adjusted_width = min(max_length + 2, 50)
#             ws.column_dimensions[column_letter].width = adjusted_width

#     def _freeze_header(self, ws) -> None:
#         """Freeze header rows"""
#         if self.freeze_rows > 0:
#             ws.freeze_panes = f"A{self.freeze_rows + 1}"

#     def _add_filters(self, ws) -> None:
#         """Add auto-filters to header row"""
#         if ws.max_row > 0:
#             ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
