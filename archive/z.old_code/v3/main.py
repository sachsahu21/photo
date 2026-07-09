

# ============================================================================
# FILE: main.py
# ============================================================================
"""
Image Scanner - Main Entry Point
"""

import sys
import logging
import pickle
from pathlib import Path
from typing import Optional

from src.config_manager import ConfigManager
from src.scanner import ImageScanner
from src.duplicate_handler import DuplicateHandler
from src.organizer import ImageOrganizer
from src.excel_writer import ExcelWriter


def setup_logging(config: ConfigManager):
    """Setup logging configuration"""
    log_level = config.get('logging.level', 'INFO')
    log_file = config.get('logging.file', './logs/image_scanner.log')

    Path(log_file).parent.mkdir(parents=True, exist_ok=True)

    logging.basicConfig(
        level=getattr(logging, log_level),
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)


def print_banner():
    """Print application banner"""
    print("""
    ╔═══════════════════════════════════════════════════════╗
    ║        IMAGE SCANNER - Professional Edition           ║
    ║     Metadata Extraction & Image Organization          ║
    ║                   v1.0.0                              ║
    ╚═══════════════════════════════════════════════════════╝
    """)


def print_menu():
    """Print main menu"""
    print("""
    ┌─────────────────────────────────────────────────────┐
    │ MAIN MENU                                           │
    ├─────────────────────────────────────────────────────┤
    │ 1. Task 1: Scan & Extract Metadata                 │
    │ 1b. Task 1B: Resume Excel Write (from backup)      │
    │ 2. Task 2: Delete Marked Files                     │
    │ 3. Task 3: Organize Images by Date                 │
    │ 4. Full Workflow (1 → 2 → 3)                       │
    │ 5. Exit                                            │
    └─────────────────────────────────────────────────────┘
    """)


def save_records_backup(records, filename='records_backup.pkl'):
    """Save records to pickle file for recovery"""
    try:
        pickle.dump(records, open(filename, 'wb'))
        print(f"✓ Backup saved: {filename}")
        return filename
    except Exception as e:
        print(f"⚠ Could not save backup: {e}")
        return None


def load_records_backup(filename='records_backup.pkl'):
    """Load records from pickle backup file"""
    try:
        if not Path(filename).exists():
            print(f"✗ Backup not found: {filename}")
            return None

        records = pickle.load(open(filename, 'rb'))
        print(f"✓ Loaded {len(records)} records from backup")
        return records
    except Exception as e:
        print(f"✗ Error loading backup: {e}")
        return None


def task_1_scan(config: ConfigManager, logger) -> Optional[str]:
    """Task 1: Scan and extract metadata"""
    print("\n" + "="*60)
    print("TASK 1: SCAN & EXTRACT METADATA")
    print("="*60)

    print(config.to_dict())

    try:
        scanner = ImageScanner(config.to_dict())
        scan_folder = config.get('scan.folder_path')
        records = scanner.scan(scan_folder)

        if not records:
            print("⚠ No images found")
            return None

        print(f"\n✓ Found {len(records)} images")

        # Save backup before processing
        save_records_backup(records)

        # Mark duplicates
        dup_handler = DuplicateHandler(
            hash_algorithm=config.get('duplicates.hash_algorithm', 'md5'),
            selection_criteria=config.get('duplicates.selection_criteria', ['quality', 'resolution', 'date', 'size'])
        )
        records = dup_handler.mark_duplicates(records)

        # Save backup after processing
        save_records_backup(records)

        dup_count = sum(1 for r in records if r.get('is_duplicate') == 'YES')
        dup_grps = len(set(r.get('duplicate_group') for r in records if r.get('duplicate_group')))
        blur_count = sum(1 for r in records if r.get('is_blurry') == True)

        print(f"✓ Detected {blur_count} blurry images")
        print(f"✓ Found {dup_count} duplicates in {dup_grps} groups")

        # Generate Excel
        print("\nGenerating Excel report (this may take a while)...")
        excel_writer = ExcelWriter(config.to_dict())
        excel_path = excel_writer.write(records, scan_folder)

        print(f"""
    ╔═══════════════════════════════════════════════════╗
    ║          TASK 1 COMPLETE                          ║
    ╠═══════════════════════════════════════════════════╣
    ║ Total images:       {len(records):>30} ║
    ║ Blurry images:      {blur_count:>30} ║
    ║ Duplicate files:    {dup_count:>30} ║
    ║ Duplicate groups:   {dup_grps:>30} ║
    ╚═══════════════════════════════════════════════════╝

    Excel file: {excel_path}
    Backup file: records_backup.pkl
        """)

        return excel_path

    except Exception as e:
        logger.error(f"Error in Task 1: {e}", exc_info=True)
        print(f"\n✗ Error: {e}")
        print("Attempting to save backup...")
        try:
            save_records_backup(records)
            print("✓ Partial data saved to records_backup.pkl")
        except:
            pass
        return None


def task_1b_resume_excel(config: ConfigManager, logger) -> Optional[str]:
    """Task 1B: Resume Excel write from backup"""
    print("\n" + "="*60)
    print("TASK 1B: RESUME EXCEL WRITE")
    print("="*60)

    try:
        # Load backup
        records = load_records_backup('records_backup.pkl')
        if not records:
            return None

        # Get scan folder
        scan_folder = config.get('scan.folder_path')

        # Write Excel
        print("\nGenerating Excel report...")
        excel_writer = ExcelWriter(config.to_dict())
        excel_path = excel_writer.write(records, scan_folder)

        print(f"""
    ╔═══════════════════════════════════════════════════╗
    ║          TASK 1B COMPLETE                         ║
    ╠═══════════════════════════════════════════════════╣
    ║ Excel file: {excel_path}
    ╚═══════════════════════════════════════════════════╝
        """)

        return excel_path

    except Exception as e:
        logger.error(f"Error in Task 1B: {e}", exc_info=True)
        print(f"\n✗ Error: {e}")
        return None


def task_2_delete(excel_path: str, logger):
    """Task 2: Delete marked files"""
    print("\n" + "="*60)
    print("TASK 2: DELETE MARKED FILES")
    print("="*60)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(excel_path)
        ws = wb['Duplicates']

        # Find columns
        delete_col = None
        filename_col = None
        fullpath_col = None

        for ci, cell in enumerate(ws[1], 1):
            if not cell.value:
                continue

            header = str(cell.value).strip().lower()

            if header == 'delete? (yes/no)':
                delete_col = ci
            elif header == 'filename':
                filename_col = ci
            elif header == 'full path':
                fullpath_col = ci

        if not delete_col:
            print("✗ DELETE column not found")
            return

        # Delete marked files
        deleted_count = 0
        error_count = 0

        for ri in range(2, ws.max_row + 1):

            delete_val = ws.cell(row=ri, column=delete_col).value
            filepath = ws.cell(row=ri, column=fullpath_col).value if fullpath_col else None
            filename = ws.cell(row=ri, column=filename_col).value if filename_col else None

            # Normalize delete flag
            delete_flag = str(delete_val).strip().lower() if delete_val is not None else ""

            if delete_flag in ("yes", "true", "1"):

                if not filepath:
                    print(f"⚠ Missing path for: {filename}")
                    continue

                try:
                    p = Path(filepath)

                    if p.exists():
                        p.unlink()
                        deleted_count += 1
                        print(f"✓ Deleted: {filename}")
                    else:
                        print(f"⚠ Not found: {filename}")

                except Exception as e:
                    error_count += 1
                    logger.error(f"Error deleting {filename}: {e}")
                    print(f"✗ Error deleting: {filename}")

        print(f"""
    ╔═══════════════════════════════════════════════════╗
    ║          TASK 2 COMPLETE                          ║
    ╠═══════════════════════════════════════════════════╣
    ║ Files deleted:      {deleted_count:>30} ║
    ║ Errors:             {error_count:>30} ║
    ╚═══════════════════════════════════════════════════╝
        """)

    except Exception as e:
        logger.error(f"Error in Task 2: {e}", exc_info=True)
        print(f"\n✗ Error: {e}")


def task_3_organize(excel_path: str, config: ConfigManager, logger):
    """Task 3: Organize images by date"""
    print("\n" + "="*60)
    print("TASK 3: ORGANIZE IMAGES BY DATE")
    print("="*60)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(excel_path)
        ws = wb['All Images']

        # Parse records from Excel
        records = []
        col_map = {}

        for ci, cell in enumerate(ws[1], 1):
            col_map[cell.value] = ci

        for ri in range(2, ws.max_row + 1):
            record = {}
            for header, col_idx in col_map.items():
                value = ws.cell(row=ri, column=col_idx).value
                record[header] = value
            records.append(record)

        # Organize files
        organizer = ImageOrganizer(config.to_dict())
        movements = organizer.organize(records)

        success_count = sum(1 for m in movements if 'Success' in m['status'])
        error_count = sum(1 for m in movements if 'Error' in m['status'])

        print(f"""
    ╔═══════════════════════════════════════════════════╗
    ║          TASK 3 COMPLETE                          ║
    ╠═══════════════════════════════════════════════════╣
    ║ Files organized:    {success_count:>30} ║
    ║ Errors:             {error_count:>30} ║
    ║ Output: {config.get('organization.output_folder'):>45} ║
    ╚═══════════════════════════════════════════════════╝
        """)

    except Exception as e:
        logger.error(f"Error in Task 3: {e}", exc_info=True)
        print(f"\n✗ Error: {e}")


def main():
    """Main entry point"""
    try:
        print_banner()

        config = ConfigManager()
        logger = setup_logging(config)
        logger.info("Application started")

        if not config.validate():
            print("✗ Configuration validation failed")
            return

        while True:
            print_menu()
            choice = input("Enter your choice (1-5): ").strip()

            if choice == '1':
                task_1_scan(config, logger)

            elif choice == '1b':
                task_1b_resume_excel(config, logger)

            elif choice == '2':
                excel_path = input("Enter Excel file path: ").strip()
                if Path(excel_path).exists():
                    task_2_delete(excel_path, logger)
                else:
                    print(f"✗ File not found: {excel_path}")

            elif choice == '3':
                excel_path = input("Enter Excel file path: ").strip()
                if Path(excel_path).exists():
                    task_3_organize(excel_path, config, logger)
                else:
                    print(f"✗ File not found: {excel_path}")

            elif choice == '4':
                excel_path = task_1_scan(config, logger)
                if excel_path:
                    input("\nPress Enter for Task 2...")
                    task_2_delete(excel_path, logger)
                    input("\nPress Enter for Task 3...")
                    task_3_organize(excel_path, config, logger)

            elif choice == '5':
                print("\n✓ Goodbye!")
                logger.info("Application closed")
                break

            else:
                print("✗ Invalid choice")

            input("\nPress Enter to continue...")

    except KeyboardInterrupt:
        print("\n✗ Interrupted")
        sys.exit(1)
    except Exception as e:
        print(f"\n✗ Fatal error: {e}")
        logging.error(f"Fatal error: {e}", exc_info=True)
        sys.exit(1)


if __name__ == '__main__':
    main()
