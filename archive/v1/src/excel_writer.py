
# ============================================================
# FILE: src/excel_writer.py
# ============================================================
"""
Excel Report Generation - Multi-sheet workbook with formatting.
Handles both image and video metadata columns.
"""

import csv
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
from collections import Counter

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Generate professional Excel reports."""

    # ── All Images columns (images + videos) ──
    IMAGE_COLUMNS = [
        ('filename', 'Filename', 30),
        ('folder', 'Folder', 50),
        ('extension', 'Format', 10),
        ('file_type', 'Type', 8),
        ('size_mb', 'Size (MB)', 12),
        # Image dimensions
        ('width', 'Width (px)', 12),
        ('height', 'Height (px)', 12),
        ('mode', 'Color Mode', 12),
        ('dpi', 'DPI', 10),
        # EXIF
        ('date_taken', 'Date Taken', 20),
        ('camera_make', 'Camera Make', 16),
        ('camera_model', 'Camera Model', 18),
        ('focal_length', 'Focal Length', 14),
        ('aperture', 'Aperture', 12),
        ('iso', 'ISO', 8),
        ('exposure_time', 'Exposure', 12),
        ('gps_lat', 'GPS Lat', 12),
        ('gps_lon', 'GPS Lon', 12),
        ('has_exif', 'Has EXIF', 10),
        # Quality / Blur
        ('blur_score', 'Blur Score', 12),
        ('quality_rating', 'Quality', 12),
        ('quality_score', 'Quality %', 11),
        ('quality_issues', 'Issues', 35),
        ('is_blurry', 'Blurry?', 10),
        # Video-specific
        ('video_duration_fmt', 'Duration', 12),
        ('video_duration_sec', 'Duration (s)', 12),
        ('video_fps', 'FPS', 8),
        ('video_codec', 'Codec', 12),
        ('video_bitrate_kbps', 'Bitrate (kbps)', 14),
        # Duplicates
        ('is_duplicate', 'Duplicate?', 12),
        ('duplicate_group', 'Dup Group', 12),
        ('is_best_in_group', 'Best?', 8),
        ('recommendation', 'Recommendation', 20),
        ('delete_flag', 'DELETE? (Yes/No)', 16),
        # Hashes / Paths
        ('md5_hash', 'MD5 Hash', 36),
        ('file_modified', 'File Modified', 20),
        ('full_path', 'Full Path', 60),
        ('error', 'Read Error', 30),
    ]

    # ── Duplicate sheet columns ──
    DUP_COLUMNS = [
        ('duplicate_group', 'Group', 12),
        ('is_best_in_group', 'Best?', 8),
        ('recommendation', 'Recommendation', 20),
        ('filename', 'Filename', 30),
        ('folder', 'Folder', 50),
        ('extension', 'Format', 10),
        ('file_type', 'Type', 8),
        ('size_mb', 'Size (MB)', 12),
        ('quality_score', 'Quality %', 11),
        ('blur_score', 'Blur Score', 12),
        ('width', 'Width (px)', 12),
        ('height', 'Height (px)', 12),
        ('video_duration_fmt', 'Duration', 12),
        ('video_codec', 'Codec', 12),
        ('date_taken', 'Date Taken', 20),
        ('delete_flag', 'DELETE? (Yes/No)', 16),
        ('md5_hash', 'MD5 Hash', 36),
        ('full_path', 'Full Path', 60),
    ]

    # ── Blurry sheet columns ──
    BLUR_COLUMNS = [
        ('filename', 'Filename', 30),
        ('folder', 'Folder', 50),
        ('blur_score', 'Blur Score', 12),
        ('quality_rating', 'Quality', 12),
        ('quality_score', 'Quality %', 11),
        ('quality_issues', 'Issues', 35),
        ('width', 'Width (px)', 12),
        ('height', 'Height (px)', 12),
        ('size_mb', 'Size (MB)', 12),
        ('date_taken', 'Date Taken', 20),
        ('delete_flag', 'DELETE? (Yes/No)', 16),
        ('full_path', 'Full Path', 60),
    ]

    # ── Colors ──
    CLR_HEADER = '2E4057'
    CLR_HEADER_DUP = '8B0000'
    CLR_HEADER_BLUR = 'CC6600'
    CLR_WHITE_FONT = 'FFFFFF'
    CLR_ALT_ROW = 'F2F4F7'
    CLR_DUP_ROW = 'FFD6D6'
    CLR_BLUR_ROW = 'FFF0D0'
    CLR_DUP_GRP1 = 'FFE0E0'
    CLR_DUP_GRP2 = 'FFF0F0'
    CLR_BEST_FONT = '006400'
    CLR_SECTION_HDR = '2E4057'
    CLR_ACCENT_ROW = 'EBF2FF'
    CLR_BORDER = 'CCCCCC'

    def __init__(self, config):
        self.config = config
        out_cfg = config.get('output', {})
        self.output_folder = Path(out_cfg.get('output_folder', './reports'))
        self.output_folder.mkdir(parents=True, exist_ok=True)
        self.prefix = out_cfg.get('filename_prefix', 'image_scan')

        sheets_cfg = out_cfg.get('sheets', {})
        self.do_all = sheets_cfg.get('all_images', True)
        self.do_blurry = sheets_cfg.get('blurry_images', True)
        self.do_dups = sheets_cfg.get('duplicates', True)
        self.do_summary = sheets_cfg.get('summary', True)
        self.do_quality = sheets_cfg.get('quality_report', True)

    # ──────────────────────────────────────────────
    # PUBLIC
    # ───────��──────────────────────────────────────

    def write(self, records, scan_folder):
        """Write complete Excel report with incremental saves."""
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl not installed")
            csv_path = self._csv_fallback(records)
            if csv_path:
                return csv_path
            raise ImportError("openpyxl required for Excel")

        parent_name = Path(scan_folder).name
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = self.output_folder / f"{self.prefix}_{parent_name}_{ts}.xlsx"

        wb = openpyxl.Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)

        try:
            if self.do_summary:
                print("  Writing Summary sheet...")
                self._sheet_summary(wb, records, scan_folder)
                wb.save(output_path)
                print("  ✓ Summary saved")

            if self.do_all:
                print("  Writing All Images sheet...")
                self._sheet_all_images(wb, records)
                wb.save(output_path)
                print("  ✓ All Images saved")

            if self.do_blurry:
                print("  Writing Blurry Images sheet...")
                self._sheet_blurry(wb, records)
                wb.save(output_path)
                print("  ✓ Blurry Images saved")

            if self.do_dups:
                print("  Writing Duplicates sheet...")
                self._sheet_duplicates(wb, records)
                wb.save(output_path)
                print("  ✓ Duplicates saved")

            if self.do_quality:
                print("  Writing Quality Report sheet...")
                self._sheet_quality(wb, records)
                wb.save(output_path)
                print("  ✓ Quality Report saved")

            logger.info(f"Excel saved: {output_path}")
            return str(output_path)

        except Exception as e:
            logger.error(f"Excel write error: {e}", exc_info=True)
            print(f"  ✗ Excel error: {e}")

            try:
                wb.save(output_path)
                print(f"  ✓ Partial Excel saved: {output_path}")
            except Exception:
                pass

            csv_path = self._csv_fallback(records)
            if csv_path:
                print(f"  ✓ CSV fallback: {csv_path}")

            if output_path.exists():
                return str(output_path)

            raise

    # ──────────────────────────────────────────────
    # SHEET: SUMMARY
    # ──────────────────────────────────────────────

    def _sheet_summary(self, wb, records, scan_folder):
        ws = wb.create_sheet('Summary', 0)
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 30

        ws['A1'] = 'Image Scanner - Summary Report'
        ws['A1'].font = Font(bold=True, size=16, color=self.CLR_HEADER)
        ws.merge_cells('A1:B1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, color='888888', size=10)

        ws['A3'] = f"Scanned: {scan_folder}"
        ws['A3'].font = Font(italic=True, color='888888', size=10)

        total = len(records)
        images = sum(1 for r in records if r.get('file_type') == 'image')
        videos = sum(1 for r in records if r.get('file_type') == 'video')
        dup_count = sum(1 for r in records if str(r.get('is_duplicate', '')).upper() == 'YES')
        dup_grps = len(set(
            r.get('duplicate_group') for r in records
            if r.get('duplicate_group') and str(r.get('duplicate_group')).strip()
        ))
        blur_count = sum(1 for r in records if r.get('is_blurry') is True)
        with_exif = sum(1 for r in records if r.get('has_exif'))
        with_gps = sum(1 for r in records if r.get('gps_lat'))
        delete_count = sum(
            1 for r in records
            if str(r.get('delete_flag', '')).strip().lower() in ('yes', 'true', '1')
        )

        q_scores = [r['quality_score'] for r in records
                     if isinstance(r.get('quality_score'), (int, float))]
        avg_q = sum(q_scores) / len(q_scores) if q_scores else 0

        total_size = sum(r.get('size_mb', 0) or 0 for r in records)
        unique_folders = len(set(r.get('folder', '') for r in records))
        ext_cnts = Counter(r.get('extension', '?') for r in records)

        # Video stats
        vid_durations = [r.get('video_duration_sec') for r in records
                         if isinstance(r.get('video_duration_sec'), (int, float)) and r.get('video_duration_sec', 0) > 0]
        total_vid_sec = sum(vid_durations) if vid_durations else 0

        from .utils import format_duration
        total_vid_fmt = format_duration(total_vid_sec) if total_vid_sec > 0 else 'N/A'

        rows = [
            ('', ''),
            ('GENERAL', ''),
            ('Total Files', total),
            ('Images', images),
            ('Videos', videos),
            ('Folders Scanned', unique_folders),
            ('Total Size', f"{total_size:.1f} MB ({total_size/1024:.2f} GB)"),
            ('Average Quality', f"{avg_q:.1f}%"),
            ('', ''),
            ('VIDEO STATS', ''),
            ('Total Video Duration', total_vid_fmt),
            ('Videos with Duration', len(vid_durations)),
            ('', ''),
            ('QUALITY ISSUES', ''),
            ('Blurry Images', blur_count),
            ('Duplicate Files', dup_count),
            ('Duplicate Groups', dup_grps),
            ('Marked for Deletion', delete_count),
            ('', ''),
            ('METADATA', ''),
            ('With EXIF Data', with_exif),
            ('With GPS Data', with_gps),
            ('Without EXIF', total - with_exif),
            ('', ''),
            ('BY FORMAT', ''),
        ]

        for ext, cnt in sorted(ext_cnts.items(), key=lambda x: -x[1]):
            rows.append((f"  .{ext.lower()}", cnt))

        section_headers = {'GENERAL', 'VIDEO STATS', 'QUALITY ISSUES', 'METADATA', 'BY FORMAT'}
        hdr_fill = PatternFill('solid', fgColor=self.CLR_SECTION_HDR)
        acc_fill = PatternFill('solid', fgColor=self.CLR_ACCENT_ROW)

        for ri, (label, value) in enumerate(rows, 5):
            a = ws.cell(row=ri, column=1, value=label)
            b = ws.cell(row=ri, column=2, value=value)

            if label in section_headers:
                for cell in (a, b):
                    cell.font = Font(bold=True, color=self.CLR_WHITE_FONT, size=11)
                    cell.fill = hdr_fill
                    cell.border = self._border()
            elif label:
                a.font = Font(bold=True, size=11)
                b.font = Font(size=11)
                if ri % 2 == 0:
                    a.fill = acc_fill
                    b.fill = acc_fill

    # ──────────────────────────────────────────────
    # SHEET: ALL IMAGES
    # ──────────────────────────────────────────────

    def _sheet_all_images(self, wb, records):
        ws = wb.create_sheet('All Images')
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 35

        for ci, (_, label, width) in enumerate(self.IMAGE_COLUMNS, 1):
            cell = ws.cell(row=1, column=ci, value=label)
            self._apply_header(cell, self.CLR_HEADER)
            ws.column_dimensions[get_column_letter(ci)].width = width

        alt_fill = PatternFill('solid', fgColor=self.CLR_ALT_ROW)
        dup_fill = PatternFill('solid', fgColor=self.CLR_DUP_ROW)
        blur_fill = PatternFill('solid', fgColor=self.CLR_BLUR_ROW)

        for ri, rec in enumerate(records, 2):
            if rec.get('is_blurry') is True:
                fill = blur_fill
            elif str(rec.get('is_duplicate', '')).upper() == 'YES':
                fill = dup_fill
            elif ri % 2 == 0:
                fill = alt_fill
            else:
                fill = None

            for ci, (key, _, _) in enumerate(self.IMAGE_COLUMNS, 1):
                val = self._safe_val(rec.get(key, ''))
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = self._border()
                cell.alignment = Alignment(vertical='center', wrap_text=False)
                if fill:
                    cell.fill = fill

        last_col = get_column_letter(len(self.IMAGE_COLUMNS))
        last_row = len(records) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # ──────────────────────────────────────────────
    # SHEET: BLURRY IMAGES
    # ──────────────────────────────────────────────

    def _sheet_blurry(self, wb, records):
        ws = wb.create_sheet('Blurry Images')
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 35

        for ci, (_, label, width) in enumerate(self.BLUR_COLUMNS, 1):
            cell = ws.cell(row=1, column=ci, value=label)
            self._apply_header(cell, self.CLR_HEADER_BLUR)
            ws.column_dimensions[get_column_letter(ci)].width = width

        blurry = sorted(
            [r for r in records if r.get('is_blurry') is True],
            key=lambda x: (x.get('blur_score') if isinstance(x.get('blur_score'), (int, float)) else 0)
        )

        fill = PatternFill('solid', fgColor=self.CLR_BLUR_ROW)

        for ri, rec in enumerate(blurry, 2):
            for ci, (key, _, _) in enumerate(self.BLUR_COLUMNS, 1):
                val = self._safe_val(rec.get(key, ''))
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = self._border()
                cell.alignment = Alignment(vertical='center')
                cell.fill = fill

        if blurry:
            last_col = get_column_letter(len(self.BLUR_COLUMNS))
            ws.auto_filter.ref = f"A1:{last_col}{len(blurry) + 1}"

    # ──────────────────────────────────────────────
    # SHEET: DUPLICATES
    # ──────────────────────────────────────────────

    def _sheet_duplicates(self, wb, records):
        """ALL members of each duplicate group shown."""
        ws = wb.create_sheet('Duplicates')
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 35

        for ci, (_, label, width) in enumerate(self.DUP_COLUMNS, 1):
            cell = ws.cell(row=1, column=ci, value=label)
            self._apply_header(cell, self.CLR_HEADER_DUP)
            ws.column_dimensions[get_column_letter(ci)].width = width

        dups = sorted(
            [r for r in records if str(r.get('is_duplicate', '')).upper() == 'YES'],
            key=lambda x: (str(x.get('duplicate_group', '')), x.get('full_path', ''))
        )

        prev_group = None
        alt_flag = False
        fill_1 = PatternFill('solid', fgColor=self.CLR_DUP_GRP1)
        fill_2 = PatternFill('solid', fgColor=self.CLR_DUP_GRP2)
        best_font = Font(bold=True, color=self.CLR_BEST_FONT)
        normal_font = Font()

        for ri, rec in enumerate(dups, 2):
            grp = rec.get('duplicate_group', '')
            if grp != prev_group:
                alt_flag = not alt_flag
                prev_group = grp

            fill = fill_1 if alt_flag else fill_2
            is_best = str(rec.get('is_best_in_group', '')).lower() == 'yes'

            for ci, (key, _, _) in enumerate(self.DUP_COLUMNS, 1):
                val = self._safe_val(rec.get(key, ''))
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = self._border()
                cell.alignment = Alignment(vertical='center')
                cell.fill = fill
                cell.font = best_font if is_best else normal_font

        if dups:
            last_col = get_column_letter(len(self.DUP_COLUMNS))
            ws.auto_filter.ref = f"A1:{last_col}{len(dups) + 1}"

    # ──────────────────────────────────────────────
    # SHEET: QUALITY REPORT
    # ──────────────────────────────────────────────

    def _sheet_quality(self, wb, records):
        ws = wb.create_sheet('Quality Report')
        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15

        ws['A1'] = 'Quality Analysis Report'
        ws['A1'].font = Font(bold=True, size=16, color=self.CLR_HEADER)
        ws.merge_cells('A1:C1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, color='888888', size=10)

        q_scores = [r['quality_score'] for r in records
                     if isinstance(r.get('quality_score'), (int, float))]
        b_scores = [r['blur_score'] for r in records
                     if isinstance(r.get('blur_score'), (int, float))]

        avg_q = sum(q_scores) / len(q_scores) if q_scores else 0
        avg_b = sum(b_scores) / len(b_scores) if b_scores else 0

        rows = [
            ('', '', ''),
            ('QUALITY STATISTICS', '', ''),
            ('Images Analyzed', len(q_scores), ''),
            ('Average Quality', f"{avg_q:.1f}%", ''),
            ('Highest Quality', f"{max(q_scores):.1f}%" if q_scores else 'N/A', ''),
            ('Lowest Quality', f"{min(q_scores):.1f}%" if q_scores else 'N/A', ''),
            ('', '', ''),
            ('BLUR STATISTICS', '', ''),
            ('Images with Blur Data', len(b_scores), ''),
            ('Average Blur Score', f"{avg_b:.1f}" if b_scores else 'N/A', ''),
            ('Sharpest', f"{max(b_scores):.1f}" if b_scores else 'N/A', ''),
            ('Blurriest', f"{min(b_scores):.1f}" if b_scores else 'N/A', ''),
            ('', '', ''),
            ('QUALITY DISTRIBUTION', 'Count', 'Percentage'),
        ]

        total_q = len(q_scores) or 1
        excellent = sum(1 for s in q_scores if s >= 80)
        good = sum(1 for s in q_scores if 60 <= s < 80)
        fair = sum(1 for s in q_scores if 40 <= s < 60)
        poor = sum(1 for s in q_scores if s < 40)

        rows.extend([
            ('Excellent (80-100%)', excellent, f"{excellent/total_q*100:.1f}%"),
            ('Good (60-79%)', good, f"{good/total_q*100:.1f}%"),
            ('Fair (40-59%)', fair, f"{fair/total_q*100:.1f}%"),
            ('Poor (0-39%)', poor, f"{poor/total_q*100:.1f}%"),
            ('', '', ''),
            ('BLUR DISTRIBUTION', 'Count', 'Percentage'),
        ])

        vb = sum(1 for r in records if r.get('quality_rating') == 'Very Blurry')
        bl = sum(1 for r in records if r.get('quality_rating') == 'Blurry')
        fa = sum(1 for r in records if r.get('quality_rating') == 'Fair')
        sh = sum(1 for r in records if r.get('quality_rating') == 'Sharp')
        total_b = (vb + bl + fa + sh) or 1

        rows.extend([
            ('Very Blurry (< 50)', vb, f"{vb/total_b*100:.1f}%"),
            ('Blurry (50-100)', bl, f"{bl/total_b*100:.1f}%"),
            ('Fair (100-200)', fa, f"{fa/total_b*100:.1f}%"),
            ('Sharp (> 200)', sh, f"{sh/total_b*100:.1f}%"),
        ])

        # Video stats
        vid_records = [r for r in records if r.get('file_type') == 'video']
        if vid_records:
            vid_with_dur = [r for r in vid_records
                            if isinstance(r.get('video_duration_sec'), (int, float))
                            and r.get('video_duration_sec', 0) > 0]
            vid_with_res = [r for r in vid_records
                            if r.get('video_width') and r.get('video_height')]

            from .utils import format_duration
            total_dur = sum(r['video_duration_sec'] for r in vid_with_dur)

            rows.extend([
                ('', '', ''),
                ('VIDEO STATISTICS', 'Count', ''),
                ('Total Videos', len(vid_records), ''),
                ('With Duration Data', len(vid_with_dur), ''),
                ('With Resolution Data', len(vid_with_res), ''),
                ('Total Duration', format_duration(total_dur), ''),
            ])

            # Codec distribution
            codecs = Counter(r.get('video_codec') for r in vid_records if r.get('video_codec'))
            if codecs:
                rows.extend([('', '', ''), ('VIDEO CODECS', 'Count', '')])
                for codec, cnt in codecs.most_common(10):
                    rows.append((codec, cnt, ''))

        # Camera distribution
        cameras = Counter()
        for r in records:
            make = r.get('camera_make', '') or ''
            model = r.get('camera_model', '') or ''
            cam = f"{make} {model}".strip()
            if cam:
                cameras[cam] += 1

        if cameras:
            rows.extend([('', '', ''), ('CAMERA DISTRIBUTION', 'Count', '')])
            for cam, cnt in cameras.most_common(15):
                rows.append((cam, cnt, ''))

        section_headers = {
            'QUALITY STATISTICS', 'BLUR STATISTICS', 'QUALITY DISTRIBUTION',
            'BLUR DISTRIBUTION', 'VIDEO STATISTICS', 'VIDEO CODECS',
            'CAMERA DISTRIBUTION'
        }
        hdr_fill = PatternFill('solid', fgColor=self.CLR_SECTION_HDR)
        acc_fill = PatternFill('solid', fgColor=self.CLR_ACCENT_ROW)

        for ri, row_data in enumerate(rows, 4):
            label = row_data[0] if len(row_data) > 0 else ''
            val1 = row_data[1] if len(row_data) > 1 else ''
            val2 = row_data[2] if len(row_data) > 2 else ''

            a = ws.cell(row=ri, column=1, value=label)
            b = ws.cell(row=ri, column=2, value=val1)
            c = ws.cell(row=ri, column=3, value=val2)

            if label in section_headers:
                for cell in (a, b, c):
                    cell.font = Font(bold=True, color=self.CLR_WHITE_FONT, size=11)
                    cell.fill = hdr_fill
                    cell.border = self._border()
            elif label:
                a.font = Font(bold=True, size=11)
                if ri % 2 == 0:
                    for cell in (a, b, c):
                        cell.fill = acc_fill

    # ──────────────────────────────────────────────
    # HELPERS
    # ──────────────────────────────────────────────

    def _apply_header(self, cell, bg_color):
        cell.font = Font(bold=True, color=self.CLR_WHITE_FONT, size=11)
        cell.fill = PatternFill('solid', fgColor=bg_color)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = self._border()

    def _border(self):
        s = Side(style='thin', color=self.CLR_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def _safe_val(self, value):
        """Convert value to Excel-safe format."""
        if value is None:
            return ''
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, bool):
            return 'Yes' if value else 'No'
        if isinstance(value, float):
            return round(value, 2)
        if isinstance(value, tuple):
            return 'x'.join(str(v) for v in value)
        if isinstance(value, (list, dict)):
            return str(value)
        if isinstance(value, str):
            return ''.join(ch for ch in value if ord(ch) >= 32 or ch in ('\n', '\t'))
        return value

    def _csv_fallback(self, records):
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            csv_path = self.output_folder / f"{self.prefix}_{ts}.csv"

            headers = [label for _, label, _ in self.IMAGE_COLUMNS]
            keys = [key for key, _, _ in self.IMAGE_COLUMNS]

            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for rec in records:
                    writer.writerow([self._safe_val(rec.get(k, '')) for k in keys])

            logger.info(f"CSV fallback saved: {csv_path}")
            return str(csv_path)
        except Exception as e:
            logger.error(f"CSV fallback failed: {e}")
            return None
