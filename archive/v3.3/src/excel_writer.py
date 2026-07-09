# ============================================================
# FILE: src/excel_writer.py
# ============================================================
"""Excel Writer v2.4 - Fixed similar sheet + file_modified column"""
import csv, logging
from pathlib import Path
from datetime import datetime
from collections import Counter
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_OK = True
except ImportError: XLSX_OK = False
logger = logging.getLogger(__name__)

class ExcelWriter:
    ALL_COLS = [
        ('filename','Filename',30),('folder','Folder',50),('extension','Format',10),('file_type','Type',8),
        ('size_mb','Size MB',12),('file_modified','File Modified',20),('metadata_status','Metadata',18),
        ('date_source','Date Source',12),('width','Width',12),('height','Height',12),
        ('date_taken','Date Taken',20),('camera_make','Camera',16),('camera_model','Model',18),
        ('focal_length','Focal',14),('aperture','Aperture',12),('iso','ISO',8),('exposure_time','Exposure',12),
        ('gps_lat','GPS Lat',12),('gps_lon','GPS Lon',12),('has_exif','EXIF',8),
        ('blur_score','Blur',10),('quality_rating','Quality',12),('quality_score','Quality %',11),
        ('is_blurry','Blurry?',8),('face_count','Faces',8),
        ('is_duplicate','Dup?',8),('duplicate_group','Dup Grp',12),
        ('is_similar','Similar?',10),('similar_group','Sim Grp',12),('similar_score','Sim %',10),('similar_methods','Sim Methods',30),
        ('recommendation','Action',20),('delete_flag','DELETE?',16),('md5_hash','MD5',36),('full_path','Full Path',60),
    ]
    DUP_COLS = [('duplicate_group','Group',12),('is_best_in_group','Best?',8),('recommendation','Action',20),('filename','Filename',30),('size_mb','Size',12),('quality_score','Q%',11),('width','W',12),('height','H',12),('date_taken','Date Taken',20),('file_modified','File Modified',20),('delete_flag','DELETE?',16),('full_path','Path',60)]
    BLUR_COLS = [('filename','Filename',30),('blur_score','Blur',12),('quality_rating','Quality',12),('quality_score','Q%',11),('width','W',12),('height','H',12),('date_taken','Date Taken',20),('file_modified','File Modified',20),('full_path','Path',60)]
    SIM_COLS = [('similar_group','Group',12),('similar_score','Sim%',12),('similar_methods','Methods',35),('filename','Filename',30),('folder','Folder',50),('extension','Fmt',10),('size_mb','Size',12),('width','W',12),('height','H',12),('date_taken','Date Taken',20),('file_modified','File Modified',20),('metadata_status','Meta',18),('quality_score','Q%',11),('is_duplicate','Exact?',12),('full_path','Path',60)]
    CH='2E4057'; CW='FFFFFF'; CA='F2F4F7'; CB='CCCCCC'

    def __init__(self, config):
        out = config.get('output', {}); self.output_folder = Path(out.get('output_folder','./reports'))
        self.output_folder.mkdir(parents=True, exist_ok=True); self.prefix = out.get('filename_prefix','image-scan')
        sh = out.get('sheets', {})
        self.do_all=sh.get('all_images',True); self.do_blur=sh.get('blurry_images',True); self.do_dup=sh.get('duplicates',True)
        self.do_sim=sh.get('similar_images',True); self.do_sum=sh.get('summary',True); self.do_qual=sh.get('quality_report',True)
        self.do_anl=sh.get('analytics',True); self.do_cls=sh.get('clusters',True)

    def write(self, records, scan_folder, analytics_data=None):
        if not XLSX_OK: return self._csv(records) or ''
        ts = datetime.now().strftime('%Y%m%d-%H%M%S')
        op = self.output_folder / (self.prefix + '-' + Path(scan_folder).name + '-' + ts + '.xlsx')
        wb = openpyxl.Workbook()
        if wb.sheetnames: wb.remove(wb.active)
        try:
            if self.do_sum: self._summary(wb, records, scan_folder); wb.save(op)
            if self.do_all: self._sheet(wb, 'All Images', self.ALL_COLS, records); wb.save(op)
            if self.do_blur:
                bl = sorted([r for r in records if r.get('is_blurry') is True], key=lambda x: x.get('blur_score',0) if isinstance(x.get('blur_score'),(int,float)) else 0)
                if bl: self._sheet(wb, 'Blurry', self.BLUR_COLS, bl, 'CC6600'); wb.save(op)
            if self.do_dup:
                dr = sorted([r for r in records if str(r.get('is_duplicate','')).upper()=='YES'], key=lambda x: (str(x.get('duplicate_group','')), x.get('full_path','')))
                if dr: self._sheet(wb, 'Duplicates', self.DUP_COLS, dr, '8B0000'); wb.save(op)
            if self.do_sim:
                sr = [r for r in records if str(r.get('is_similar','')).upper() == 'YES']
                if sr:
                    sr = sorted(sr, key=lambda x: (str(x.get('similar_group','')), x.get('full_path','')))
                    self._sheet(wb, 'Similar Images', self.SIM_COLS, sr, '4A148C')
                    ws = wb['Similar Images']; pg = None; alt = False
                    f1, f2 = PatternFill('solid', fgColor='E8D5F5'), PatternFill('solid', fgColor='F3E5F5')
                    for ri, rec in enumerate(sr, 2):
                        g = rec.get('similar_group', '')
                        if g != pg: alt = not alt; pg = g
                        for ci in range(1, len(self.SIM_COLS)+1): ws.cell(row=ri, column=ci).fill = f1 if alt else f2
                    wb.save(op)
                    print('  Similar Images: ' + str(len(sr)) + ' files written')
                else:
                    print('  Similar Images: 0 (sheet skipped)')
            if self.do_qual: self._qual(wb, records); wb.save(op)
            if self.do_anl and analytics_data: self._anl(wb, analytics_data); wb.save(op)
            if self.do_cls:
                cr = [r for r in records if r.get('cluster_label')]
                if cr: self._sheet(wb, 'Clusters', [('cluster_label','Cluster',15),('filename','File',30),('quality_score','Q%',12),('full_path','Path',50)], sorted(cr, key=lambda x: x.get('cluster_label',''))); wb.save(op)
            return str(op)
        except Exception as e:
            logger.error('Excel: %s', e)
            try: wb.save(op)
            except: pass
            return str(op) if op.exists() else (self._csv(records) or '')

    def _hdr(self, c, bg=None):
        c.font=Font(bold=True,color=self.CW,size=11); c.fill=PatternFill('solid',fgColor=bg or self.CH)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); c.border=self._brd()
    def _brd(self):
        s=Side(style='thin',color=self.CB); return Border(left=s,right=s,top=s,bottom=s)
    def _sv(self, v):
        if v is None: return ''
        if isinstance(v, datetime): return v.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(v, bool): return 'Yes' if v else 'No'
        if isinstance(v, float): return round(v, 2)
        if isinstance(v, (list,tuple,dict)): return str(v)
        if isinstance(v, str): return ''.join(ch for ch in v if ord(ch)>=32 or ch in '\n\t')
        return v

    def _sheet(self, wb, name, cols, data, hc=None):
        ws=wb.create_sheet(name); ws.freeze_panes='A2'; ws.row_dimensions[1].height=35
        for ci,(_, lb, wd) in enumerate(cols,1):
            c=ws.cell(row=1,column=ci,value=lb); self._hdr(c,hc); ws.column_dimensions[get_column_letter(ci)].width=wd
        af=PatternFill('solid',fgColor=self.CA)
        for ri, rec in enumerate(data,2):
            for ci,(key,_,_) in enumerate(cols,1):
                c=ws.cell(row=ri,column=ci,value=self._sv(rec.get(key,''))); c.border=self._brd(); c.alignment=Alignment(vertical='center')
                if ri%2==0: c.fill=af
        if data: ws.auto_filter.ref='A1:' + get_column_letter(len(cols)) + str(len(data)+1)

    def _qual(self, wb, records):
        ws=wb.create_sheet('Quality Report'); ws.column_dimensions['A'].width=38; ws.column_dimensions['B'].width=20
        ws['A1']='Quality Analysis'; ws['A1'].font=Font(bold=True,size=16,color=self.CH)
        qs=[r['quality_score'] for r in records if isinstance(r.get('quality_score'),(int,float))]
        aq=sum(qs)/len(qs) if qs else 0
        rows=[('',''),('STATS',''),('Analyzed',len(qs)),('Average',str(round(aq,1))+'%'),('',''),('METADATA','')]
        for st,cnt in Counter(r.get('metadata_status','?') for r in records).most_common(): rows.append((st,cnt))
        rows.extend([('',''),('DATE SOURCE','')])
        for st,cnt in Counter(r.get('date_source','?') for r in records).most_common(): rows.append((st,cnt))
        sc=sum(1 for r in records if str(r.get('is_similar','')).upper()=='YES')
        sg=len(set(r.get('similar_group') for r in records if r.get('similar_group') and str(r.get('similar_group','')).strip()))
        rows.extend([('',''),('SIMILAR',''),('Similar Files',sc),('Similar Groups',sg)])
        secs={'STATS','METADATA','DATE SOURCE','SIMILAR'}; hf=PatternFill('solid',fgColor=self.CH)
        for ri,(l,v) in enumerate(rows,4):
            a,b=ws.cell(row=ri,column=1,value=l),ws.cell(row=ri,column=2,value=v)
            if l in secs:
                for x in (a,b): x.font=Font(bold=True,color=self.CW); x.fill=hf
            elif l: a.font=Font(bold=True)

    def _summary(self, wb, records, sf):
        ws=wb.create_sheet('Summary',0); ws.column_dimensions['A'].width=35; ws.column_dimensions['B'].width=30
        ws['A1']='Image Scanner v2.4'; ws['A1'].font=Font(bold=True,size=16,color=self.CH); ws['A2']='Scanned: '+str(sf)
        t=len(records)
        rows=[('',''),('GENERAL',''),('Total',t),('Images',sum(1 for r in records if r.get('file_type')=='image')),('Videos',sum(1 for r in records if r.get('file_type')=='video')),('Size',str(round(sum(r.get('size_mb',0) or 0 for r in records),1))+' MB'),('',''),('METADATA',''),('Full EXIF',sum(1 for r in records if r.get('metadata_status')=='Full EXIF')),('No EXIF',sum(1 for r in records if r.get('metadata_status')=='No EXIF')),('',''),('QUALITY',''),('Blurry',sum(1 for r in records if r.get('is_blurry') is True)),('Duplicates',sum(1 for r in records if str(r.get('is_duplicate','')).upper()=='YES')),('Similar',sum(1 for r in records if str(r.get('is_similar','')).upper()=='YES'))]
        secs={'GENERAL','METADATA','QUALITY'}; hf=PatternFill('solid',fgColor=self.CH)
        for ri,(l,v) in enumerate(rows,5):
            a,b=ws.cell(row=ri,column=1,value=l),ws.cell(row=ri,column=2,value=v)
            if l in secs:
                for x in (a,b): x.font=Font(bold=True,color=self.CW,size=11); x.fill=hf
            elif l: a.font=Font(bold=True,size=11)

    def _anl(self, wb, data):
        ws=wb.create_sheet('Analytics'); ws.column_dimensions['A'].width=40; ws.column_dimensions['B'].width=20
        ws['A1']='Storage Analytics'; ws['A1'].font=Font(bold=True,size=16,color=self.CH)
        for ri,(l,v) in enumerate([('',''),('Total Files',data.get('total_files',0)),('Total Size',data.get('total_size_human',''))],3):
            ws.cell(row=ri,column=1,value=l).font=Font(bold=True); ws.cell(row=ri,column=2,value=v)

    def _csv(self, records):
        try:
            p=self.output_folder/(self.prefix+'-'+datetime.now().strftime('%Y%m%d-%H%M%S')+'.csv')
            with open(p,'w',newline='',encoding='utf-8-sig') as f:
                w=csv.writer(f); w.writerow([l for _,l,_ in self.ALL_COLS])
                for r in records: w.writerow([self._sv(r.get(k,'')) for k,_,_ in self.ALL_COLS])
            return str(p)
        except: return None
