

# ============================================================
# FILE: main.py
# ============================================================
"""
Image Scanner - Main Entry Point
"""

import sys
import logging
import pickle
from pathlib import Path
from typing import Optional

from src.config_manager import ConfigManager
from src.scanner import ImageScanner
from src.duplicate_handler import DuplicateHandler
from src.organizer import ImageOrganizer
from src.excel_writer import ExcelWriter


BACKUP_FILE = 'records_backup.pkl'


def setup_logging(config):
    log_level = config.get('logging.level', 'INFO')
    log_file = config.get('logging.file', './logs/image_scanner.log')
    log_console = config.get('logging.console', True)

    Path(log_file).parent.mkdir(parents=True, exist_ok=True)

    handlers = [logging.FileHandler(log_file, encoding='utf-8')]
    if log_console:
        handlers.append(logging.StreamHandler())

    logging.basicConfig(
        level=getattr(logging, str(log_level).upper(), logging.INFO),
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=handlers,
    )

    return logging.getLogger(__name__)


def save_backup(records, filename=BACKUP_FILE):
    try:
        with open(filename, 'wb') as f:
            pickle.dump(records, f, protocol=pickle.HIGHEST_PROTOCOL)
        print(f"  ✓ Backup saved: {filename} ({len(records)} records)")
        return filename
    except Exception as e:
        print(f"  ⚠ Backup failed: {e}")
        return None


def load_backup(filename=BACKUP_FILE):
    try:
        if not Path(filename).exists():
            print(f"  ✗ Backup not found: {filename}")
            return None
        with open(filename, 'rb') as f:
            records = pickle.load(f)
        print(f"  ✓ Loaded {len(records)} records from backup")
        return records
    except Exception as e:
        print(f"  ✗ Backup load error: {e}")
        return None


def print_banner():
    print("""
    ╔═══════════════════════════════════════════════════════╗
    ║        IMAGE SCANNER - Professional Edition           ║
    ║     Metadata Extraction & Image Organization          ║
    ║                   v1.0.0                              ║
    ╚═══════════════════════════════════════════════════════╝
    """)


def print_menu():
    print("""
    ┌─────────────────────────────────────────────────────┐
    │ MAIN MENU                                           │
    ├─────────────────────────────────────────────────────┤
    │ 1.  Scan & Extract Metadata                        │
    │ 1b. Resume Excel Write (from backup)               │
    │ 2.  Delete Marked Files                            │
    │ 3.  Organize Images by Date                        │
    │ 4.  Full Workflow (1 -> 2 -> 3)                    │
    │ 5.  Exit                                           │
    └─────────────────────────────────────────────────────┘
    """)


def print_box(title, rows):
    w = 55
    print(f"\n    ╔{'═' * w}╗")
    print(f"    ║  {title:<{w - 2}}║")
    print(f"    ╠{'═' * w}╣")
    for label, value in rows:
        line = f"  {label}: {value}"
        print(f"    ║{line:<{w}}║")
    print(f"    ╚{'═' * w}╝")


def task_1_scan(config, logger):
    """Scan folder, extract metadata, detect blur/duplicates, generate Excel."""
    print("\n" + "=" * 60)
    print("  TASK 1: SCAN & EXTRACT METADATA")
    print("=" * 60)

    records = None

    try:
        scanner = ImageScanner(config.to_dict())
        scan_folder = config.get('scan.folder_path')

        print(f"\n  Scanning: {scan_folder}")
        records = scanner.scan(scan_folder)

        if not records:
            print("  ⚠ No files found!")
            print("  Tips: Check scan.folder_path and extensions in config.yaml")
            return None

        print(f"\n  ✓ Found {len(records)} files")

        # Count types
        img_count = sum(1 for r in records if r.get('file_type') == 'image')
        vid_count = sum(1 for r in records if r.get('file_type') == 'video')
        print(f"    Images: {img_count}, Videos: {vid_count}")

        save_backup(records)

        # Duplicates
        if config.get('duplicates.enabled', True):
            print("\n  Detecting duplicates...")
            dup_handler = DuplicateHandler(
                hash_algorithm=config.get('duplicates.hash_algorithm', 'md5'),
                selection_criteria=config.get('duplicates.selection_criteria',
                                               ['quality', 'resolution', 'date', 'size']),
                match_mode=config.get('duplicates.match_mode', 'exact'),
                similarity_threshold=config.get('duplicates.similarity_threshold', 90),
            )
            records = dup_handler.mark_duplicates(records)

        save_backup(records)

        dup_count = sum(1 for r in records if str(r.get('is_duplicate', '')).upper() == 'YES')
        dup_grps = len(set(
            r.get('duplicate_group') for r in records
            if r.get('duplicate_group') and str(r.get('duplicate_group')).strip()
        ))
        blur_count = sum(1 for r in records if r.get('is_blurry') is True)
        delete_count = sum(
            1 for r in records
            if str(r.get('delete_flag', '')).strip().lower() in ('yes', 'true', '1')
        )

        print(f"  ✓ {blur_count} blurry images")
        print(f"  ✓ {dup_count} duplicates in {dup_grps} groups")
        print(f"  ✓ {delete_count} auto-marked for deletion")

        print("\n  Generating Excel report...")
        excel_writer = ExcelWriter(config.to_dict())
        excel_path = excel_writer.write(records, scan_folder)

        print_box("TASK 1 COMPLETE", [
            ("Total files", len(records)),
            ("Images", img_count),
            ("Videos", vid_count),
            ("Blurry", blur_count),
            ("Duplicates", dup_count),
            ("Groups", dup_grps),
            ("Delete marked", delete_count),
            ("Excel", excel_path),
            ("Backup", BACKUP_FILE),
        ])

        return excel_path

    except Exception as e:
        logger.error(f"Task 1 error: {e}", exc_info=True)
        print(f"\n  ✗ Error: {e}")
        if records:
            save_backup(records)
        return None


def task_1b_resume(config, logger):
    """Resume Excel generation from backup."""
    print("\n" + "=" * 60)
    print("  TASK 1B: RESUME EXCEL FROM BACKUP")
    print("=" * 60)

    try:
        records = load_backup()
        if not records:
            return None

        scan_folder = config.get('scan.folder_path')

        print("\n  Generating Excel...")
        writer = ExcelWriter(config.to_dict())
        path = writer.write(records, scan_folder)

        print_box("TASK 1B COMPLETE", [
            ("Records", len(records)),
            ("Excel", path),
        ])
        return path

    except Exception as e:
        logger.error(f"Task 1B error: {e}", exc_info=True)
        print(f"\n  ✗ Error: {e}")
        return None


def task_2_delete(excel_path, logger):
    """Delete files marked 'Yes' in the Duplicates sheet."""
    print("\n" + "=" * 60)
    print("  TASK 2: DELETE MARKED FILES")
    print("=" * 60)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(excel_path)

        sheet_name = None
        for name in ['Duplicates', 'All Images']:
            if name in wb.sheetnames:
                sheet_name = name
                break

        if not sheet_name:
            print("  ✗ No suitable sheet found")
            return

        ws = wb[sheet_name]
        print(f"  Reading: {sheet_name}")

        delete_col = fullpath_col = filename_col = None

        for ci, cell in enumerate(ws[1], 1):
            if not cell.value:
                continue
            h = str(cell.value).strip().lower()
            if 'delete' in h:
                delete_col = ci
            elif h == 'full path':
                fullpath_col = ci
            elif h == 'filename':
                filename_col = ci

        if not delete_col:
            print("  ✗ DELETE column not found")
            print(f"  Headers: {[c.value for c in ws[1] if c.value]}")
            return

        if not fullpath_col:
            print("  ✗ Full Path column not found")
            return

        to_delete = []
        for ri in range(2, ws.max_row + 1):
            val = ws.cell(row=ri, column=delete_col).value
            if val and str(val).strip().lower() in ('yes', 'true', '1'):
                fp = ws.cell(row=ri, column=fullpath_col).value
                fn = ws.cell(row=ri, column=filename_col).value if filename_col else 'unknown'
                to_delete.append((fp, fn))

        if not to_delete:
            print("  ✓ No files marked for deletion")
            return

        print(f"\n  {len(to_delete)} files marked for deletion.")
        confirm = input("  Confirm delete? (yes/no): ").strip().lower()
        if confirm != 'yes':
            print("  ✗ Cancelled")
            return

        deleted = errors = not_found = 0

        for fp, fn in to_delete:
            if not fp:
                print(f"  ⚠ No path for: {fn}")
                errors += 1
                continue
            try:
                p = Path(fp)
                if p.exists():
                    p.unlink()
                    deleted += 1
                    print(f"  ✓ Deleted: {fn}")
                else:
                    not_found += 1
                    print(f"  ⚠ Not found: {fn}")
            except Exception as e:
                errors += 1
                logger.error(f"Delete error {fn}: {e}")
                print(f"  ✗ Error: {fn}")

        print_box("TASK 2 COMPLETE", [
            ("Deleted", deleted),
            ("Not found", not_found),
            ("Errors", errors),
        ])

    except ImportError:
        print("  ✗ openpyxl not installed")
    except Exception as e:
        logger.error(f"Task 2 error: {e}", exc_info=True)
        print(f"\n  ✗ Error: {e}")


def task_3_organize(excel_path, config, logger):
    """Organize images by date from Excel data."""
    print("\n" + "=" * 60)
    print("  TASK 3: ORGANIZE IMAGES BY DATE")
    print("=" * 60)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(excel_path)

        sheet_name = None
        for name in ['All Images', 'Duplicates']:
            if name in wb.sheetnames:
                sheet_name = name
                break

        if not sheet_name:
            print("  ✗ No suitable sheet found")
            return

        ws = wb[sheet_name]
        print(f"  Reading: {sheet_name}")

        col_map = {}
        for ci, cell in enumerate(ws[1], 1):
            if cell.value:
                col_map[str(cell.value).strip()] = ci

        header_to_key = {
            'Filename': 'filename',
            'Folder': 'folder',
            'Full Path': 'full_path',
            'Format': 'extension',
            'Type': 'file_type',
            'Size (MB)': 'size_mb',
            'Date Taken': 'date_taken',
            'File Modified': 'file_modified',
            'DELETE? (Yes/No)': 'delete_flag',
            'Quality %': 'quality_score',
            'Blur Score': 'blur_score',
            'MD5 Hash': 'md5_hash',
            'Duplicate?': 'is_duplicate',
            'Best?': 'is_best_in_group',
            'Group': 'duplicate_group',
            'Dup Group': 'duplicate_group',
            'Recommendation': 'recommendation',
            'Duration': 'video_duration_fmt',
            'Duration (s)': 'video_duration_sec',
            'FPS': 'video_fps',
            'Codec': 'video_codec',
            'Bitrate (kbps)': 'video_bitrate_kbps',
        }

        records = []
        for ri in range(2, ws.max_row + 1):
            rec = {}
            for header, ci in col_map.items():
                val = ws.cell(row=ri, column=ci).value
                key = header_to_key.get(header, header.lower().replace(' ', '_'))
                rec[key] = val
                rec[header] = val
            records.append(rec)

        print(f"  Loaded {len(records)} records")

        organizer = ImageOrganizer(config.to_dict())
        movements = organizer.organize(records)

        success = sum(1 for m in movements if m['status'] == 'Success')
        errors = sum(1 for m in movements if 'Error' in m['status'])
        skipped = sum(1 for m in movements if 'Skip' in m['status'])

        print_box("TASK 3 COMPLETE", [
            ("Organized", success),
            ("Skipped", skipped),
            ("Errors", errors),
            ("Output", config.get('organization.output_folder', '')),
        ])

    except ImportError:
        print("  ✗ openpyxl not installed")
    except Exception as e:
        logger.error(f"Task 3 error: {e}", exc_info=True)
        print(f"\n  ✗ Error: {e}")


def main():
    """Main entry point with interactive menu."""
    try:
        print_banner()

        config = ConfigManager()
        logger = setup_logging(config)
        logger.info("=" * 50)
        logger.info("Image Scanner started")

        print(f"  Config: {config.config_path}")
        print(f"  Scan: {config.get('scan.folder_path')}")

        if not config.validate():
            print("\n  ⚠ Fix configuration issues before continuing.")
            resp = input("  Continue anyway? (yes/no): ").strip().lower()
            if resp != 'yes':
                return

        last_excel = None

        while True:
            print_menu()
            choice = input("  Choice (1, 1b, 2, 3, 4, 5): ").strip().lower()

            if choice == '1':
                last_excel = task_1_scan(config, logger)

            elif choice == '1b':
                last_excel = task_1b_resume(config, logger)

            elif choice == '2':
                path = input("  Excel path (Enter for last): ").strip()
                if not path and last_excel:
                    path = last_excel
                    print(f"  Using: {path}")
                if path and Path(path).exists():
                    task_2_delete(path, logger)
                else:
                    print(f"  ✗ Not found: {path}")

            elif choice == '3':
                path = input("  Excel path (Enter for last): ").strip()
                if not path and last_excel:
                    path = last_excel
                    print(f"  Using: {path}")
                if path and Path(path).exists():
                    task_3_organize(path, config, logger)
                else:
                    print(f"  ✗ Not found: {path}")

            elif choice == '4':
                print("\n  Full Workflow: Scan -> Delete -> Organize")
                last_excel = task_1_scan(config, logger)
                if last_excel:
                    print("\n  Review Excel, then continue.")
                    if input("  Proceed to Delete? (yes/no): ").strip().lower() == 'yes':
                        task_2_delete(last_excel, logger)
                        if input("  Proceed to Organize? (yes/no): ").strip().lower() == 'yes':
                            task_3_organize(last_excel, config, logger)

            elif choice == '5':
                print("\n  ✓ Goodbye!")
                logger.info("Application closed")
                break

            else:
                print("  ✗ Invalid choice")

            input("\n  Press Enter to continue...")

    except KeyboardInterrupt:
        print("\n\n  ✗ Interrupted")
        sys.exit(1)
    except Exception as e:
        print(f"\n  ✗ Fatal: {e}")
        logging.error(f"Fatal: {e}", exc_info=True)
        sys.exit(1)


if __name__ == '__main__':
    main()
